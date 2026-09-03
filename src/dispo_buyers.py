"""One deduped buyer registry for Knox and Blount, ready to skip trace and text.

The dispo side has never had what the acquisition side has: a list of PEOPLE.
`buyer_sweep.py` resolves deed-level buyers one zip at a time,
`dispo_flip_buyers.py` landed pending-flip PROPERTIES in the CRM, and the
buyer-prospector skill ships a nationwide CSV that is 176 rows here and mostly
Opendoor. None of them answer "who are the two thousand people who buy houses in
these two counties, and what is their phone number".

That is what this builds, in phases, each resumable and DRY by default:

    sweep       search rows per county x investor-transaction type  (cheap)
    hydrate     get_detail per row -> owner name + mailing address  (slow)
    aggregate   dedupe into buyer identities                        (free)
    unmask      LLC -> human principal by reverse address           (FREE)
    principals  the ones unmask missed, via Enformion BusinessV2    (BILLED)
    skipinput   project the registry into a SmartSkip upload        (free)
    phones      join the SmartSkip return back onto buyer_key         (free)
    score       Trestle dial tiers (--source registry|preset)         (BILLED)
    qa          the gate table

The order of the two principal phases is the whole cost story: an LLC's mailing
address is usually a person's house, so reversing it through the map we already
pay for names the principal for nothing. Only what that misses is worth $0.10.

THE SEMANTIC TRAP THIS FILE EXISTS TO SURVIVE. The SiftMap filter key labels the
LAST SALE, not the current owner. For `pending`, `wholesale`, `wholetail` and
`rental` the investor bought and still holds, so the CURRENT OWNER is the buyer.
For `flip` the exit already happened, so the current owner is a RETAIL HOMEBUYER
and the person we want is the last-sale SELLER. Reading the flip bucket the same
way as the others puts ordinary families on a cash-buyer call list.

Verified live 2026-08-28 before any of this was written:
  * `search` in the address block must be the bare county name. "Knox" returns
    153,675; "Knox County, TN" returns 0, which is indistinguishable from an
    empty segment.
  * An unknown filter key is SILENTLY IGNORED. A deliberately bogus key returned
    a byte-identical count, so a key is only proven by a count delta.
  * `extra_collapse_by_owner` is real: Knox pending 1,633 -> 1,140 rows.
  * Search rows carry NO owner name, so hydration is unavoidable. They do carry
    `saved_uuid`, the CRM record id, which the mirror phase reuses.
  * `sale_history` is ordered NEWEST FIRST (18 of 18 multi-sale properties came
    back strictly descending). The flip branch below reads `sale_history[0]`, so
    that ordering is load-bearing: reversed, every flipper we resolved would be
    whoever sold the house years before the investor ever touched it.

    python src/dispo_buyers.py --phase sweep
    python src/dispo_buyers.py --phase hydrate --workers 4
    python src/dispo_buyers.py --phase aggregate
    python src/dispo_buyers.py --phase qa
"""
from __future__ import annotations

import argparse
import io
import json
import logging
import os
import re
import sys
import threading
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import config  # noqa: F401  loads .env
    OUTPUT_ROOT = Path(getattr(config, "OUTPUT_DIR", "output"))
except Exception:
    OUTPUT_ROOT = Path("output")

from siftmap_standalone import SiftMapClient, SiftMapError  # noqa: E402
from enterprise_prospects import (  # noqa: E402
    ENTITY_RX, ORG_RX, SearchClient, WORLD, _match_candidate, classify_one,
    county_address, norm_name, strip_unit,
)

log = logging.getLogger("dispo_buyers")

OUT = OUTPUT_ROOT / "dispo_buyers"
STATE_PATH = OUT / "state.json"

K_TXN = "extra_last_sale_investor_transaction_type"

COUNTIES = [("47093", "Knox", "TN"), ("47009", "Blount", "TN")]

# The four types where the current owner IS the buyer, and the one where it is
# not. Kept as separate constants because every bug in this area comes from
# treating them the same.
HOLD_TYPES = ["pending", "wholesale", "wholetail", "rental"]
EXIT_TYPES = ["flip"]
ALL_TYPES = HOLD_TYPES + EXIT_TYPES

PAGE_SIZE = 250

# Suite markers force a name check before two buyers at one address are merged.
# Measured on the existing sweeps: 9111 Cross Park Dr Ste D200 hosts NS Homes
# LLC and R D Properties Group LLC, unrelated companies sharing an agent.
UNIT_MARK_RX = re.compile(r"\b(STE|SUITE|UNIT|APT|BLDG|FL|FLOOR|PMB|RM)\b", re.I)

STREET_WORD_RX = re.compile(
    r"\b(STREET|ST|ROAD|RD|DRIVE|DR|LANE|LN|AVENUE|AVE|BOULEVARD|BLVD|COURT|"
    r"CT|CIRCLE|CIR|PIKE|PKWY|PARKWAY|WAY|TRAIL|TRL|PLACE|PL)\b")


# ---------------------------------------------------------------- state ----

def _load(p: Path, default):
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return default


def _save(p: Path, obj) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_text(json.dumps(obj, indent=1, default=str), encoding="utf-8")
    tmp.replace(p)


def state() -> dict:
    return _load(STATE_PATH, {})


def mark_done(phase: str, **facts) -> None:
    s = state()
    s[phase] = {"when": datetime.now().isoformat(timespec="seconds"), **facts}
    _save(STATE_PATH, s)


def require(phase: str) -> None:
    if phase not in state():
        raise SystemExit("phase '" + phase + "' has not run yet; run it first")


def gate(ok: bool, msg: str) -> None:
    """A phase that produced nothing is a failure, not an empty market."""
    if not ok:
        raise SystemExit("GATE FAILED: " + msg)


# ---------------------------------------------------------------- sweep ----

def _body(fips: str, name: str, state_ab: str, txn: str, page: int,
          collapse: bool, since: str = "") -> dict:
    filters = {
        "type_single_family": True,
        K_TXN: [txn],
        # Gift deeds and intra-family paper are not investor purchases.
        "extra_last_sale_value_min": 1000,
        "extra_is_last_sale_interfamily": False,
    }
    if since:
        # Verified by count delta 2026-08-28: Knox 6,912 unbounded drops to
        # 1,151 at 12 months, while a deliberately bogus key returns the
        # unfiltered 6,912. Without this the sweep has NO date bound at all and
        # silently spans 2019 to 2026.
        filters["extra_last_sale_date_min"] = since
    if collapse:
        filters["extra_collapse_by_owner"] = True
    return {
        # result_index is a ROW OFFSET, not a page number.
        "result_index": 1 + (page - 1) * PAGE_SIZE,
        "with_boundaries": False,
        "filters": filters,
        "addresses": [county_address(fips, name, state_ab)],
        "polygon": WORLD,
    }


def phase_sweep(args) -> None:
    sc = SearchClient(min_interval=args.min_interval)
    rows_path = OUT / "rows.json"
    store = _load(rows_path, {})
    counts: dict = {}

    for fips, name, st in COUNTIES:
        for txn in ALL_TYPES:
            key = name + ":" + txn
            if key in store and not args.refresh:
                counts[key] = len(store[key])
                log.info("%s already swept (%s rows)", key, len(store[key]))
                continue
            # Hold types collapse to one row per owner, which is exactly the
            # aggregation. Exit types must NOT collapse: we need each sale so
            # the seller (the flipper) can be read off its history.
            collapse = txn in HOLD_TYPES
            seen: dict = {}
            page = 1
            total = None
            while True:
                r = sc.search(_body(fips, name, st, txn, page, collapse))
                if total is None:
                    total = r.get("total_results") or 0
                    log.info("%s: %s total%s", key, total,
                             " (collapsed)" if collapse else "")
                batch = r.get("data") or []
                if not batch:
                    break
                for row in batch:
                    rid = str(row.get("id") or "")
                    if rid:
                        seen[rid] = {
                            "id": rid,
                            "address": row.get("address"),
                            "county": row.get("county"),
                            "corporateOwned": row.get("corporateOwned"),
                            "absenteeOwner": row.get("absenteeOwner"),
                            "estimatedValue": row.get("estimatedValue"),
                            "saved_uuid": row.get("saved_uuid"),
                            "txn": txn,
                            # The physical buy box. A first version kept only
                            # the eight fields above and threw these away, which
                            # left no way to answer "what does this buyer
                            # actually buy" without re-pulling everything.
                            # 93 to 96% populated live. yearBuilt is NOT
                            # returned by this endpoint at all (0 of 250), so
                            # there is deliberately no property-age dimension,
                            # and propertyUse is constant because the sweep
                            # filters type_single_family.
                            "bedrooms": row.get("bedrooms"),
                            "bathrooms": row.get("bathrooms"),
                            "squareFeet": row.get("squareFeet"),
                            "lotAcres": row.get("lotAcres"),
                            "equityPercent": row.get("equityPercent"),
                            "investorScore": row.get("investor_off_market_score"),
                            "realtorScore": row.get("realtor_score"),
                            # Only the ACTIVE flags. The raw blob is a 27-key
                            # nested dict per property and would bloat rows.json
                            # for information that is almost all False.
                            "distress": sorted(
                                k for k, v in (row.get("distressors") or {}).items()
                                if isinstance(v, dict) and v.get("is_active")),
                        }
                page += 1
                if len(seen) >= (total or 0) or page > args.max_pages:
                    break
            store[key] = list(seen.values())
            counts[key] = len(seen)
            _save(rows_path, store)
            log.info("%s: kept %s rows", key, len(seen))

    total_rows = sum(counts.values())
    uniq = len({r["id"] for v in store.values() for r in v})
    gate(total_rows > 0, "sweep returned zero rows across both counties")
    mark_done("sweep", per_segment=counts, rows=total_rows, unique_ids=uniq)
    print(json.dumps(counts, indent=1))
    print("sweep: " + str(total_rows) + " rows, " + str(uniq) + " unique properties")


# ----------------------------------------------------------------- recent --
# WHO is still buying. The main sweep has no date bound, so it answers "who has
# ever bought here" and half that list stopped years ago: the median buyer last
# purchased 456 days ago and only 41% bought inside a year. This phase runs the
# same sweep with extra_last_sale_date_min and is the qualification layer.
# Recency decides who is on the list; the unbounded history decides what they
# buy. It is cheap because the recent set is a near-subset of what is already
# hydrated (1,101 of 1,129 on the first run).


def _since(days: int) -> str:
    from datetime import date, timedelta
    return (date.today() - timedelta(days=days)).isoformat()


def phase_recent(args) -> None:
    sc = SearchClient(min_interval=args.min_interval)
    since = _since(args.recent_days)
    details = _load(OUT / "details.json", {})
    rows_path = OUT / "recent_rows.json"
    store: dict = {}
    log.info("qualifying window: last sale on or after %s", since)

    for fips, name, st in COUNTIES:
        for txn in ALL_TYPES:
            collapse = txn in HOLD_TYPES
            seen: dict = {}
            page, total = 1, None
            while True:
                r = sc.search(_body(fips, name, st, txn, page, collapse, since))
                if total is None:
                    total = r.get("total_results") or 0
                batch = r.get("data") or []
                if not batch:
                    break
                for row in batch:
                    rid = str(row.get("id") or "")
                    if rid:
                        seen[rid] = {"id": rid, "address": row.get("address"),
                                     "county": row.get("county"), "txn": txn,
                                     "saved_uuid": row.get("saved_uuid")}
                page += 1
                if len(seen) >= (total or 0) or page > args.max_pages:
                    break
            store[name + ":" + txn] = list(seen.values())
            log.info("%s:%s -> %s in window", name, txn, len(seen))

    _save(rows_path, store)
    ids = {r["id"] for v in store.values() for r in v}
    unhydrated = [i for i in ids if i not in details]
    gate(len(ids) > 0, "the recent sweep returned nothing; check the date bound")

    # Resolve each qualifying property to the buyer, using the SAME hold-vs-exit
    # rule as the main sweep. For a flip the qualifying party is the last-sale
    # SELLER, and a recent exit is its own strong signal: that operator just
    # freed up capital and needs the next deal.
    active: dict = {}
    unresolved = 0
    for key, seg in store.items():
        county, txn = key.split(":", 1)
        for r in seg:
            d = details.get(r["id"]) or {}
            if d.get("miss") or not d.get("owner_name"):
                unresolved += 1
                continue
            sh = d.get("sale_history") or []
            last = sh[0] if sh else {}
            if txn in HOLD_TYPES:
                who = d["owner_name"]
            else:
                who = (last.get("seller") or "").strip()
            if not who:
                unresolved += 1
                continue
            k = norm_name(who)
            e = active.setdefault(k, {"name": who, "n_in_window": 0,
                                      "last_active_buy": "", "counties": set(),
                                      "txns": set()})
            e["n_in_window"] += 1
            e["counties"].add(county)
            e["txns"].add(txn)
            when = str(last.get("date") or "")
            if when > e["last_active_buy"]:
                e["last_active_buy"] = when
    for e in active.values():
        e["counties"] = sorted(e["counties"])
        e["txns"] = sorted(e["txns"])

    _save(OUT / "recent.json", active)
    mark_done("recent", days=args.recent_days, since=since,
              properties=len(ids), unhydrated=len(unhydrated),
              buyers=len(active), unresolved=unresolved)
    print("recent: %d properties sold in the last %d days"
          % (len(ids), args.recent_days))
    print("  already hydrated          : %d" % (len(ids) - len(unhydrated)))
    print("  NEW, need hydrating       : %d" % len(unhydrated))
    print("  buyers active in window   : %d" % len(active))
    print("  properties with no owner  : %d" % unresolved)
    if unhydrated:
        print("")
        print("  run --phase hydrate next, then --phase aggregate")


# -------------------------------------------------------------- hydrate ----

_HYDRATE_LOCK = threading.Lock()


def _owner_name(oi: dict) -> str:
    raw = (oi.get("owner_name")
           or " ".join(p for p in [oi.get("first_name"), oi.get("last_name")] if p)
           or oi.get("name") or "")
    return re.sub(r"\s+", " ", raw).strip()


def _mail_line(oi: dict) -> str:
    """owner_info carries the mailing address as ONE string, not a dict."""
    m = oi.get("owner_mail_address") or ""
    if isinstance(m, dict):
        m = ", ".join(p for p in [m.get("street"), m.get("city"),
                                  m.get("state"), m.get("zip")] if p)
    return (m or "").strip()


def _hydrate_one(smc: SiftMapClient, rid: str) -> dict:
    d = smc.get_detail(rid)
    oi = d.get("owner_info") or {}
    sh = d.get("sale_history") or []
    return {
        "id": rid,
        "owner_name": _owner_name(oi),
        "owner_mail": _mail_line(oi),
        "portfolio_n": oi.get("total_properties"),
        "portfolio_value": oi.get("portfolio_value"),
        "secondary_owners": oi.get("secondary_owner_names") or [],
        "sale_history": [{"date": s.get("sale_date"),
                          "price": s.get("sale_price"),
                          "buyer": s.get("buyer_name"),
                          "seller": s.get("seller_name"),
                          "cash": s.get("is_cash_sale")} for s in sh[:4]],
    }


def phase_hydrate(args) -> None:
    require("sweep")
    rows = _load(OUT / "rows.json", {})
    # The date-bounded qualification sweep finds properties the unbounded one
    # missed, and they need owners too. Reading only rows.json left 28 of them
    # permanently unhydrated while the phase reported "nothing to do".
    recent = _load(OUT / "recent_rows.json", {})
    det_path = OUT / "details.json"
    details = _load(det_path, {})

    todo = []
    for source in (rows, recent):
        for seg_rows in source.values():
            for r in seg_rows:
                if r["id"] not in details:
                    todo.append(r["id"])
    todo = sorted(set(todo))
    if args.limit:
        todo = todo[:args.limit]
    log.info("hydrating %s properties (%s already cached)", len(todo), len(details))
    if not todo:
        mark_done("hydrate", details=len(details))
        print("hydrate: nothing to do, " + str(len(details)) + " cached")
        return

    done = 0
    errors = 0
    clients: dict = {}

    def work(rid: str):
        tid = threading.get_ident()
        if tid not in clients:
            clients[tid] = SiftMapClient(min_interval=args.min_interval)
        return _hydrate_one(clients[tid], rid)

    # Detail throughput caps near 2-4 req/s AGGREGATE regardless of worker
    # count, with escalating Retry-After under sustained load. More workers
    # than that just buys 429s.
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(work, rid): rid for rid in todo}
        for fut in as_completed(futs):
            rid = futs[fut]
            try:
                details[rid] = fut.result()
            except SiftMapError as e:
                if "401" in str(e):
                    raise
                errors += 1
                details[rid] = {"id": rid, "miss": str(e)[:160]}
            except Exception as e:  # noqa: BLE001
                errors += 1
                details[rid] = {"id": rid, "miss": str(e)[:160]}
            done += 1
            # Checkpoint often. A stall on the far side of a long interval
            # throws away everything since the last write, and this phase has
            # already hung once on compounding 429 backoff.
            if done % 100 == 0:
                with _HYDRATE_LOCK:
                    _save(det_path, details)
                log.info("hydrated %s/%s (%s errors)", done, len(todo), errors)

    _save(det_path, details)
    resolved = sum(1 for d in details.values() if d.get("owner_name"))
    # The buyer_sweep precedent: a run that resolves zero of a non-empty target
    # list is an auth failure wearing an empty-market costume.
    gate(resolved > 0, "hydrate resolved zero owners; check REISIFT_API_KEY")
    mark_done("hydrate", details=len(details), resolved=resolved, errors=errors)
    print("hydrate: " + str(len(details)) + " cached, " + str(resolved)
          + " with an owner, " + str(errors) + " errors")


# ------------------------------------------------------------ aggregate ----

def _addr_key(mail: str) -> str:
    """Normalized mailing address, unit stripped, for clustering."""
    s = strip_unit(mail or "")
    s = re.sub(r"[^A-Z0-9 ]", " ", s.upper())
    s = STREET_WORD_RX.sub("", s)
    return re.sub(r"\s+", " ", s).strip()


def _name_tokens(name: str) -> set:
    stop = {"LLC", "LLP", "LP", "INC", "CORP", "THE", "COMPANY", "CO"}
    return {t for t in norm_name(name).split() if t not in stop and len(t) > 2}


def _similar(a: str, b: str) -> bool:
    """Fuzzy entity-name match for merging inside a shared-suite address."""
    ta, tb = _name_tokens(a), _name_tokens(b)
    if not ta or not tb:
        return False
    if ta <= tb or tb <= ta:
        return True
    overlap = len(ta & tb) / max(1, min(len(ta), len(tb)))
    return overlap >= 0.6


def _is_entity(name: str) -> bool:
    return bool(ENTITY_RX.search(norm_name(name)))


def local_tier(name: str, n_buys: int) -> tuple:
    """`classify_one` with two corrections this list cannot survive without.

    That taxonomy was calibrated for a NATIONWIDE enterprise prospect list,
    where volume means 6-month purchases across the country and "construction"
    means D R Horton. Applied unchanged to a county sweep it is wrong twice,
    and both were measured on the real registry:

    1. **It excludes the best dispo buyers we have.** Below its volume floor of
       10, any generic keyword match is an outright EXCLUDE. Nearly every local
       investor is below that floor, so CREEKSTONE CONSTRUCTION LLC (9 buys),
       MCCARLEY CONSTRUCTION LLC (6) and HEMBREE BUILDERS LLC were all dropped
       as "homebuilders". A local construction LLC buying 3 to 9 houses a year
       is a SELF-PERFORMER, which is precisely the buyer this team's own dispo
       doctrine says takes a heavy-rehab deal when GC-model flippers cannot.
       So a generic match becomes REVIEW here: kept, and flagged.
    2. **It matches keywords as substrings.** "BANK" inside "WILLBANKS" excluded
       a real buyer with 4 purchases. Any keyword that does not survive a
       word-boundary re-check is treated as a false positive.

    High-confidence brand exclusions (D R Horton, Fannie Mae, a city) are kept
    exactly as they are; those are right in both contexts.
    """
    tier, reason, conf = classify_one(name, n_buys)
    if tier != "EXCLUDE":
        return tier, reason, conf
    keyword = reason.split(":", 1)[-1].strip()
    if keyword and not re.search(r"\b" + re.escape(keyword) + r"\b", name):
        return "REVIEW", "kept: '" + keyword + "' matched only as a substring", "high"
    if "(generic)" in reason:
        return "REVIEW", "kept: generic '" + keyword + "' on a local buyer", "med"
    return tier, reason, conf


PHYS_FIELDS = ("bedrooms", "bathrooms", "squareFeet", "lotAcres",
               "equityPercent", "estimatedValue", "investorScore", "distress")


def _phys(row: dict) -> dict:
    """The physical facts of one property, carried onto its observation.

    Attached here rather than to the registry's `buys` list because that list is
    capped at 25 for display, and a 184-purchase buyer's box computed from 25
    sampled rows is a different number than one computed from all of them.
    """
    return {k: row.get(k) for k in PHYS_FIELDS}


def _observations(rows: dict, details: dict) -> list:
    obs = []
    for key, seg_rows in rows.items():
        county, txn = key.split(":", 1)
        for r in seg_rows:
            d = details.get(r["id"]) or {}
            if d.get("miss") or not d.get("owner_name"):
                continue
            sh = d.get("sale_history") or []
            last = sh[0] if sh else {}
            if txn in HOLD_TYPES:
                # The investor still holds it: the current owner is the buyer.
                obs.append({
                    "name": d["owner_name"], "mail": d.get("owner_mail") or "",
                    "county": county, "txn": txn, "role": "holder",
                    "prop": r.get("address"), "date": last.get("date"),
                    "price": last.get("price"), "cash": last.get("cash"),
                    "portfolio_n": d.get("portfolio_n"),
                    "portfolio_value": d.get("portfolio_value"),
                    "saved_uuid": r.get("saved_uuid"),
                    "secondary": d.get("secondary_owners") or [],
                    **_phys(r),
                })
            else:
                # The flip already exited: the current owner is the RETAIL
                # buyer. The investor is the SELLER on that last sale, and this
                # property gives us no mailing address for them.
                seller = (last.get("seller") or "").strip()
                if not seller:
                    continue
                obs.append({
                    "name": seller, "mail": "", "county": county, "txn": txn,
                    "role": "exited", "prop": r.get("address"),
                    "date": last.get("date"), "price": last.get("price"),
                    "cash": last.get("cash"), "portfolio_n": None,
                    "portfolio_value": None, "saved_uuid": None, "secondary": [],
                    **_phys(r),
                })
    return obs


def phase_aggregate(args) -> None:
    require("hydrate")
    rows = _load(OUT / "rows.json", {})
    details = _load(OUT / "details.json", {})

    obs = _observations(rows, details)
    gate(len(obs) > 0, "aggregate produced no observations")

    # Pass 1: exact normalized name. Pass 2: mailing-address cluster, guarded
    # so a shared office suite does not merge unrelated companies.
    by_name: dict = {}
    for o in obs:
        by_name.setdefault(norm_name(o["name"]), []).append(o)
    groups = [{"names": {n}, "obs": list(v)} for n, v in by_name.items()]

    addr_index: dict = defaultdict(list)
    for g in groups:
        mails = [o["mail"] for o in g["obs"] if o["mail"]]
        g["mail"] = mails[0] if mails else ""
        if g["mail"]:
            addr_index[_addr_key(g["mail"])].append(g)

    merges = []
    refused = []
    dropped = set()
    for akey, cluster in addr_index.items():
        if len(cluster) < 2 or not akey:
            continue
        anchor = cluster[0]
        suite = bool(UNIT_MARK_RX.search(anchor["mail"]))
        for other in cluster[1:]:
            a = sorted(anchor["names"])[0]
            b = sorted(other["names"])[0]
            if suite and not _similar(a, b):
                refused.append({"address": anchor["mail"], "kept_apart": [a, b],
                                "reason": "shared suite, names not similar"})
                continue
            anchor["names"] |= other["names"]
            anchor["obs"].extend(other["obs"])
            dropped.add(id(other))
            merges.append({"address": anchor["mail"], "merged": [a, b],
                           "reason": "same mailing address"
                                     + (" plus similar names" if suite else "")})
    groups = [g for g in groups if id(g) not in dropped]

    registry = []
    for g in groups:
        o_all = g["obs"]
        raw = Counter(o["name"] for o in o_all).most_common(1)[0][0]
        prices = sorted(o["price"] for o in o_all
                        if isinstance(o.get("price"), (int, float)) and o["price"])
        holders = [o for o in o_all if o["role"] == "holder"]
        tier, reason, _conf = local_tier(norm_name(raw), len(o_all))
        registry.append({
            "buyer_key": norm_name(raw),
            "name": raw,
            "aliases": sorted(n for n in g["names"] if n != norm_name(raw)),
            "is_entity": _is_entity(raw),
            "mail": g.get("mail") or "",
            "counties": sorted({o["county"] for o in o_all}),
            "txn_types": sorted({o["txn"] for o in o_all}),
            "n_buys": len(o_all),
            "n_held": len(holders),
            "n_exited": len(o_all) - len(holders),
            "cash_n": sum(1 for o in o_all if o.get("cash")),
            "price_min": prices[0] if prices else None,
            "price_med": prices[len(prices) // 2] if prices else None,
            "price_max": prices[-1] if prices else None,
            "portfolio_n": max((o.get("portfolio_n") or 0) for o in o_all) or None,
            "portfolio_value": max((o.get("portfolio_value") or 0)
                                   for o in o_all) or None,
            "saved_uuid": next((o["saved_uuid"] for o in o_all
                                if o.get("saved_uuid")), None),
            "secondary": sorted({s for o in o_all for s in (o.get("secondary") or [])}),
            "tier": tier,
            "tier_reason": reason,
            "buys": [{"address": o["prop"], "date": o["date"], "price": o["price"],
                      "county": o["county"], "txn": o["txn"], "role": o["role"]}
                     for o in sorted(o_all, key=lambda x: str(x.get("date") or ""),
                                     reverse=True)[:25]],
        })

    # ---- join the recency layer ------------------------------------------
    # The unbounded sweep says who has EVER bought here; recent.json says who
    # is still buying. Stamping it here keeps one source of truth, so the
    # profile phase does not have to re-derive activity.
    recent = _load(OUT / "recent.json", {})
    hot_cut = _since(182)
    matched = 0
    for b in registry:
        r = recent.get(b["buyer_key"])
        if not r:
            b["active_365d"] = False
            b["active_182d"] = False
            b["n_in_window"] = 0
            b["last_active_buy"] = ""
            continue
        matched += 1
        b["active_365d"] = True
        b["last_active_buy"] = r["last_active_buy"]
        b["n_in_window"] = r["n_in_window"]
        b["active_182d"] = bool(r["last_active_buy"] >= hot_cut)

    # A buyer who only shows up in the recent sweep is currently active and
    # would otherwise be invisible. Carry them in with what we know.
    known = {b["buyer_key"] for b in registry}
    added = 0
    for key, r in recent.items():
        if key in known:
            continue
        added += 1
        registry.append({
            "buyer_key": key, "name": r["name"], "aliases": [],
            "is_entity": _is_entity(r["name"]), "mail": "",
            "counties": r["counties"], "txn_types": r["txns"],
            "n_buys": r["n_in_window"], "n_held": 0, "n_exited": 0, "cash_n": 0,
            "price_min": None, "price_med": None, "price_max": None,
            "portfolio_n": None, "portfolio_value": None, "saved_uuid": None,
            "secondary": [], "tier": "REVIEW",
            "tier_reason": "seen only in the recency sweep",
            "buys": [], "active_365d": True, "active_182d":
                bool(r["last_active_buy"] >= hot_cut),
            "n_in_window": r["n_in_window"],
            "last_active_buy": r["last_active_buy"],
        })

    registry.sort(key=lambda b: (-b["n_buys"], -(b["portfolio_n"] or 0)))

    # An exited flipper with no mailing address cannot be skip traced by
    # address. Most active ones also hold something, so they are recovered by
    # the name merge above; the rest are reported, never silently dropped.
    no_mail = [b for b in registry if not b["mail"]]
    excluded = [b for b in registry if b["tier"] == "EXCLUDE"]

    _save(OUT / "registry.json", registry)
    _save(OUT / "merge_report.json", {"merges": merges, "refused": refused})

    act = sum(1 for b in registry if b.get("active_365d"))
    hot = sum(1 for b in registry if b.get("active_182d"))
    mark_done("aggregate", buyers=len(registry), observations=len(obs),
              merges=len(merges), refused=len(refused),
              no_mailing_address=len(no_mail), excluded=len(excluded),
              active_365d=act, active_182d=hot, recent_only_added=added)
    print("aggregate: " + str(len(obs)) + " observations -> "
          + str(len(registry)) + " unique buyers")
    print("  merged by mailing address : " + str(len(merges)))
    print("  refused by the suite guard: " + str(len(refused)))
    print("  no mailing address        : " + str(len(no_mail)) + " (exited flippers)")
    print("  institutional/EXCLUDE     : " + str(len(excluded)))
    print("  active in the last 365d   : " + str(act) + " ("
          + str(hot) + " within 182d)")
    print("  added from the recency sweep only: " + str(added))


# --------------------------------------------------------------- unmask ----
# The Harper move, and it is FREE. An LLC's mailing address is very often a
# human's house, so reversing that address through the same map we already pay
# for names the principal at no cost. Every entity that resolves here is one we
# do not send to Enformion at $0.10, which is why this phase runs first and on
# its own.

# Registered-agent and corporate-front names that must never become a
# "principal". Without this scrub the dial sheet fills up with C T Corporation
# System and US Bank Trust, which are 800-numbers, not buyers.
AGENT_FRONT_RX = re.compile(
    r"SERVICE OF PROCESS|CORPORATION SYSTEM|REGISTERED AGENT|NORTHWEST|"
    r"CORPORATION AGENTS|CORPORATE DIRECT|COGENCY|INCORP|NATIONAL ASSOC",
    re.I)


# A professional firm wearing a person-shaped name. The stock AGENT_FRONTS list
# only knows the national commercial agents (CT Corporation, Northwest), so it
# waves through the LOCAL law firms that actually appear on Knox filings.
# Measured live: "DRYER AND ASSOCIATES P C" and "CHAMBLISS BAHNER & STOPHEL"
# both came back as the "principal" of a real buying LLC.
#
# An ampersand ALONE is not the tell. Deed and filing names are full of married
# couples ("Oneal Brian A & Ethel L", "Ciminieri Louis & Colleen"), and a first
# pass that dropped every "&" threw away 7 real co-owning couples to catch 5
# firms. `clean_owner_name` already splits a couple down to one person, so the
# only thing worth rejecting is an explicit professional marker.
FIRM_RX = re.compile(
    r"\bAND ASSOCIATES\b|\bASSOCIATES\b|\bP\.?\s?C\.?\s*$|\bPLLC\b|\bLLP\b|"
    r"\bLAW\b|\bATTORNEY|\bLEGAL\b|\bFIRM\b|\bCPA\b|\bACCOUNTAN|"
    r"&\s*CO\b|\bLIMITED PARTNERSHIP\b|\bHOSPITAL\b|\bCLINIC\b|\bOFFICE\b",
    re.I)

# Titles that mean the person actually runs the company, as opposed to merely
# receiving its mail.
OWNER_TITLES = ("MANAGER", "MGR", "MEMBER", "ORGANIZER", "OFFICER", "PRESIDENT",
                "CEO", "SECRETARY", "TREASURER", "PARTNER", "DIRECTOR", "OWNER",
                "INCORPORATOR", "CONTACT")

# Legal status and credential tokens that a name parser can mistake for a
# surname. "Morales Family Trust" resolved to first=Morales last=Tr.
STATUS_TOKENS = {
    "TR", "TRUSTEE", "TTEE", "TRUST", "EST", "ESTATE", "LIFE", "ETAL", "ET",
    "AL", "UX", "JR", "SR", "II", "III", "IV", "V", "MD", "DC", "PC", "DDS",
    "DVM", "ESQ", "CPA", "REV", "LLC", "INC", "CORP", "CO", "LP", "LLP",
}


# Placeholder strings that arrive where a name should be. SiftMap returned the
# literal "UNKNOWN" as the owner of one LLC's mailing address, and it sailed
# through every entity and firm check because it is neither. Left alone it
# reaches a text as "Hi UNKNOWN".
PLACEHOLDER_NAMES = {
    "UNKNOWN", "UNKNOWN OWNER", "NONE", "N A", "NA", "TBD", "OWNER",
    "CURRENT OWNER", "OCCUPANT", "CURRENT RESIDENT", "RESIDENT", "TAXPAYER",
    "NOT AVAILABLE", "NO OWNER", "VACANT", "ESTATE OF",
}


def _is_human(name: str) -> bool:
    n = norm_name(name or "")
    if not n or n in PLACEHOLDER_NAMES:
        return False
    # A person has a first name and a surname. One bare token is a placeholder,
    # a company fragment, or a surname with no one attached to it.
    if len(n.split()) < 2:
        return False
    if ORG_RX.search(n) or AGENT_FRONT_RX.search(name or ""):
        return False
    return not FIRM_RX.search(name or "")


# BUSINESSV2 CANNOT BE CONSTRAINED GEOGRAPHICALLY, AND THAT IS THE WHOLE
# PROBLEM WITH IT. Verified live 2026-08-28: searching "SMITHBILT LLC" with no
# anchor, with Addresses[{AddressLine2: "Knoxville, TN"}], and with
# Addresses[{State: "TN"}] all return byte-identical results. It fuzzy-matches
# company names nationally, so a Knoxville LLC resolves to officers of a
# same-named company in another state. Measured on 12 cohort entities, only 2
# had ANY Tennessee officer, which means most of what the first pass bought was
# a stranger who shares a company name.
#
# The officer's own address is the only geographic control available, and it is
# only available because extract_officers() returns it. Filtering to the
# buyer's own state is therefore mandatory, not a refinement.
def _officer_in_state(officer: dict, state: str) -> bool:
    """Is this officer physically in the buyer's state.

    Tokenized rather than regexed on purpose. The first version used word
    boundaries and the escape did not survive being written to the file: the
    pattern compiled with literal backspace characters, matched nothing, and
    reported 0 of 153 entities verified as though it were a finding about the
    data. Splitting the address into alphanumeric tokens has no escaping to
    get wrong.
    """
    addr = (officer.get("address") or "").upper()
    if not addr:
        return False
    st = (state or "TN").upper()
    tokens = "".join(c if c.isalnum() else " " for c in addr).split()
    if st in tokens:
        return True
    return st == "TN" and "TENNESSEE" in tokens


def rank_officers(entity: str, officers: list, state: str) -> list:
    """State-verified officers first, then the ones who actually run the place.

    Ty's rule (2026-08-28): prefer an ownership title, fall back to any officer
    rather than giving up on the entity. But an officer in the wrong state is
    not a fallback, it is a different person, so those are dropped entirely.
    """
    keep = [o for o in officers if _officer_in_state(o, state)]
    def rank(o):
        conf = principal_confidence(entity, o.get("name") or "", o.get("title") or "")
        title = (o.get("title") or "").upper()
        owner_title = any(k in title for k in OWNER_TITLES)
        return (0 if owner_title else 1, 0 if conf == "high" else 1)
    return sorted(keep, key=rank)


def principal_confidence(entity: str, person: str, title: str) -> str:
    """How much to trust a BusinessV2 officer as the person to actually text.

    THE REGISTERED AGENT IS OFTEN THE COMPANY'S LAWYER, NOT ITS OWNER. On the
    live run 135 of 193 resolved officers carried the title AGENT or REGISTERED
    AGENT, and texting a firm's attorney a deal blast is the litigation bait
    this program is built to avoid.

    But an agent title is not automatically wrong: plenty of small operators are
    their own registered agent. The tell is the SURNAME. "Turner Homes LLC ->
    Michael L Turner" and "Smithbilt LLC -> Smith Kenneth" are obviously the
    owner; "New Season Properties LLC -> Dryer and Associates" is obviously not.

    So: an ownership title is trusted outright, an agent title is trusted only
    when a name token also appears in the company name, and everything else is
    low confidence and held back from the text list.
    """
    t = (title or "").upper()
    if any(k in t for k in OWNER_TITLES):
        return "high"
    ent_tokens = {w for w in norm_name(entity).split() if len(w) > 2}
    per_tokens = {w for w in norm_name(person).split() if len(w) > 2}
    if ent_tokens & per_tokens:
        return "high"
    return "low"


def _unmask_one(smc: SiftMapClient, buyer: dict) -> dict:
    out = {"principal": None, "principal_source": None, "miss": None}
    mail = buyer.get("mail") or ""
    if not mail:
        out["miss"] = "no mailing address"
        return out
    if re.search(r"\bP\.?\s?O\.?\s+BOX\b", mail, re.I):
        # A PO box reverses to nothing; it is not a residence.
        out["miss"] = "PO box mailing address"
        return out
    try:
        stripped = strip_unit(mail)
        hits = smc.autocomplete(stripped)
        hit = _match_candidate(hits, stripped)
        if not hit and stripped != mail:
            hits = smc.autocomplete(mail)
            hit = _match_candidate(hits, mail)
        if not hit:
            out["miss"] = "no autocomplete match on mailing address"
            return out
        d = smc.get_detail(hit["id"])
        oi = d.get("owner_info") or {}
        owner = _owner_name(oi)
        if owner and _is_human(owner):
            out["principal"] = owner
            out["principal_source"] = "siftmap-reverse-address"
            return out
        for s in (oi.get("secondary_owner_names") or []):
            if _is_human(str(s)):
                out["principal"] = str(s)
                out["principal_source"] = "siftmap-secondary-owner"
                return out
        out["miss"] = "mailing address owner is an organization"
    except SiftMapError as e:
        if "401" in str(e):
            raise
        out["miss"] = str(e)[:160]
    except Exception as e:  # noqa: BLE001
        out["miss"] = str(e)[:160]
    return out


def phase_unmask(args) -> None:
    require("aggregate")
    reg = _load(OUT / "registry.json", [])
    path = OUT / "principals.json"
    found = _load(path, {})

    targets = [b for b in reg
               if b["is_entity"] and b["tier"] not in ("EXCLUDE",)
               and b["buyer_key"] not in found]
    if args.limit:
        targets = targets[:args.limit]
    log.info("unmasking %s entities (%s already done)", len(targets), len(found))
    if not targets:
        mark_done("unmask", principals=len(found))
        print("unmask: nothing to do, " + str(len(found)) + " cached")
        return

    clients: dict = {}

    def work(b: dict):
        tid = threading.get_ident()
        if tid not in clients:
            clients[tid] = SiftMapClient(min_interval=args.min_interval)
        return b["buyer_key"], _unmask_one(clients[tid], b)

    done = 0
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = [ex.submit(work, b) for b in targets]
        for fut in as_completed(futs):
            key, res = fut.result()
            found[key] = res
            done += 1
            if done % 100 == 0:
                _save(path, found)
                log.info("unmasked %s/%s", done, len(targets))

    _save(path, found)
    hits = sum(1 for v in found.values() if v.get("principal"))
    gate(len(found) > 0, "unmask produced nothing")
    mark_done("unmask", attempted=len(found), resolved=hits)
    print("unmask: " + str(hits) + " principals from " + str(len(found))
          + " entities, all free")
    misses = Counter(v.get("miss") for v in found.values() if not v.get("principal"))
    for reason, n in misses.most_common(6):
        print("  miss: %-46s %s" % (str(reason)[:46], n))


# ----------------------------------------------------------- principals ----
# The paid fallback, and the ONLY source that can name an LLC's officers. It
# runs after unmask so we never buy what the map already gave us, and it needs
# --commit because every match is billed.

def phase_principals(args) -> None:
    require("unmask")
    import enformion_business as eb

    reg = _load(OUT / "registry.json", [])
    found = _load(OUT / "principals.json", {})
    by_key = {b["buyer_key"]: b for b in reg}

    # A buy floor, because the spend should follow the value. Measured on the
    # real registry: of the 1,257 unresolved entities that still carry a
    # mailing address, only 264 bought more than once in the window. The other
    # 993 are single-purchase LLCs, which are usually a holding vehicle or a
    # one-time buyer rather than a repeat dispo customer, so paying $0.10 each
    # to name them is $99 for the weakest part of the list (Ty, 2026-08-28).
    # No mailing-address prerequisite. BusinessV2 searches by COMPANY NAME, and
    # the entities with no mailing address (exited flippers) are exactly the
    # ones with no other route to a principal, so requiring an address excluded
    # the cases that needed this most.
    #
    # Re-attempt anything previously resolved by BusinessV2 as well: that pass
    # ran without state verification, so its answers are unverified until they
    # clear the TN filter.
    todo = [k for k, v in found.items()
            if k in by_key
            and (not v.get("principal")
                 or v.get("principal_source") == "enformion-businessv2")
            and by_key[k]["is_entity"] and by_key[k]["tier"] != "EXCLUDE"
            and by_key[k]["n_buys"] >= args.min_buys
            and (not args.active_only or by_key[k].get("active_365d"))]
    # Most valuable first, so an interrupted run has already bought the names
    # that matter.
    todo.sort(key=lambda k: -by_key[k]["n_buys"])
    if args.limit:
        todo = todo[:args.limit]

    if not eb.is_configured():
        print("principals: ENFORMION_AP_NAME / _PASSWORD not set, skipping")
        return
    cost = len(todo) * 0.10
    print("principals: " + str(len(todo)) + " entities with " + str(args.min_buys)
          + "+ buys still unresolved after the free unmask")
    print("            up to $%.2f at $0.10 per match (misses are free)" % cost)
    if not args.commit:
        print("DRY RUN. Re-run with --commit to spend.")
        return

    hits = 0
    cleared: list = []
    for i, key in enumerate(todo, 1):
        b = by_key[key]
        # Anchor on the mailing city/state so a common company name in another
        # state does not answer for this one.
        city_state = ""
        parts = [p.strip() for p in (b.get("mail") or "").split(",")]
        if len(parts) >= 3:
            city_state = parts[-2] + ", " + parts[-1].split()[0]
        try:
            officers = eb.find_principals(b["name"], city_state)
        except Exception as e:  # noqa: BLE001
            found[key]["miss"] = "businessv2: " + str(e)[:120]
            continue
        officers = [o for o in officers if _is_human(o.get("name") or "")]
        state = (b.get("mail", "").strip()[-8:].split()[0] if b.get("mail") else "") or "TN"
        if len(state) != 2 or not state.isalpha():
            state = "TN"
        ranked = rank_officers(b["name"], officers, state)
        if ranked:
            best = ranked[0]
            found[key]["principal"] = best["name"]
            found[key]["principal_source"] = "enformion-businessv2"
            found[key]["principal_title"] = best.get("title") or ""
            found[key]["principal_address"] = best.get("address") or ""
            # Keep the whole list. The first pass kept officers[0] and threw the
            # rest away, which is how a Registered Agent won whenever it
            # happened to be listed first.
            found[key]["officers"] = ranked
            found[key]["principal_confidence"] = principal_confidence(
                b["name"], best["name"], best.get("title") or "")
            found[key]["principal_state_verified"] = True
            found[key]["miss"] = None
            hits += 1
        else:
            # Clear an earlier unverified answer rather than leaving a stranger
            # from another state standing as this buyer's name.
            if found[key].get("principal_source") == "enformion-businessv2":
                cleared.append(found[key].get("principal"))
                found[key]["principal"] = None
                found[key]["principal_confidence"] = None
                found[key]["principal_address"] = ""
                # Clear the provenance too. Leaving principal_source set on a
                # row whose principal was withdrawn makes the sheet claim a
                # source for a name that is no longer there.
                found[key]["principal_source"] = None
                found[key]["principal_title"] = ""
            found[key]["principal_state_verified"] = False
            found[key]["miss"] = ("businessv2: %d officers, none in %s"
                                  % (len(officers), state)) if officers                 else "businessv2: no human officer"
        if i % 25 == 0:
            _save(OUT / "principals.json", found)
            log.info("businessv2 %s/%s (%s resolved)", i, len(todo), hits)

    _save(OUT / "principals.json", found)
    mark_done("principals", attempted=len(todo), resolved=hits,
              cleared_unverified=len(cleared))
    print("principals: %d of %d resolved and state verified" % (hits, len(todo)))
    print("  cleared as unverifiable (wrong state): %d" % len(cleared))
    for n in [c for c in cleared if c][:8]:
        print("      dropped: %s" % n)


# ------------------------------------------------------------ skipinput ----
# SmartSkip wants First Name, Last Name, Mailing Address, which is exactly the
# registry's natural key, so this is a straight projection. Nothing here is
# billed: submit/map/calculate are all free and print the exact quote, so the
# decision to spend happens after seeing the row count and price.

def split_mail(mail: str) -> tuple:
    """"1007 Andover View Ln, Knoxville, TN 37919" -> its four parts.

    SiftMap hands the mailing address back as ONE string, but SmartSkip's
    contract is `mailingAddress` = the STREET, with mailingCity / mailingState /
    mailingZip as separate optional fields (its own synonym list calls the
    field "mailingstreet"). Passing the whole line as the street matches worse,
    and on a 1,900-row batch that is real money spent on a degraded input.
    """
    raw = (mail or "").strip().rstrip(",")
    if not raw:
        return "", "", "", ""
    # TWO ADDRESS FORMATS REACH THIS FUNCTION. SiftMap writes
    # "1007 Andover View Ln, Knoxville, TN 37919" but Enformion officer records
    # write "4300 Hiawatha; Knoxville, TN 37919" with a SEMICOLON before the
    # city. Splitting on commas alone glued the city onto the street and left
    # Mailing City empty, and SmartSkip returned nothing for all 33 rows of a
    # paid batch as a result.
    # The semicolon is the street/city delimiter, so it takes precedence over
    # commas rather than being flattened into one. "2099 Thunderhead, Ste 204;
    # Knoxville, TN 37922" has a comma INSIDE the street, and treating both
    # separators alike made the city read "Ste 204".
    head = ""
    if ";" in raw:
        head, raw = raw.split(";", 1)
        head, raw = head.strip().rstrip(","), raw.strip()
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    if head:
        # Everything before the semicolon is the street; what follows is
        # city, state and zip.
        parts = [head] + parts
    street = parts[0] if parts else ""
    city = parts[1] if len(parts) > 2 else (parts[1] if len(parts) == 2 else "")
    tail = parts[-1] if len(parts) > 1 else ""
    m = re.match(r"^([A-Za-z]{2})\s+(\d{5})(?:-\d{4})?$", tail)
    if m:
        state, zip_code = m.group(1).upper(), m.group(2)
        if len(parts) < 3:
            city = ""
    else:
        # No trailing "ST 12345": fall back to whatever we can see.
        state, zip_code = "", ""
        m2 = re.search(r"\b(\d{5})(?:-\d{4})?\b", tail)
        if m2:
            zip_code = m2.group(1)
        m3 = re.search(r"\b([A-Za-z]{2})\b", tail)
        if m3 and not zip_code:
            state = m3.group(1).upper()
    return street, city, state, zip_code


def phase_skipinput(args) -> None:
    require("aggregate")
    import csv
    from enformion_ftm import clean_owner_name

    reg = _load(OUT / "registry.json", [])
    principals = _load(OUT / "principals.json", {})
    # Never re-buy a number we already own. SmartSkip bills per hit and does not
    # know we asked before: on the first pruned run 92 of 130 rows were people
    # already traced, which would have been a silent double charge.
    traced = _load(OUT / "phones.json", {})

    rows = []
    skipped = Counter()
    for b in reg:
        # Skip trace is the expensive step ($0.15 a head), so the floor is set
        # here rather than at the registry. 84% of the traceable set bought
        # exactly once in the window; a repeat buyer is the one provably still
        # in the market (Ty, 2026-08-28).
        if b["n_buys"] < args.trace_min_buys:
            skipped["under the trace buy floor"] += 1
            continue
        # Only trace buyers who are still buying. Paying $0.15 for a phone
        # number belonging to someone whose last purchase was in 2021 is the
        # whole problem this pass exists to fix.
        if args.require_active and not b.get("active_365d"):
            skipped["no purchase in the activity window"] += 1
            continue
        if b["buyer_key"] in traced and not args.retrace:
            skipped["already has phones (not re-buying)"] += 1
            continue
        if b["tier"] == "EXCLUDE":
            skipped["institutional/excluded"] += 1
            continue
        if b["tier"] == "3":
            # Institutional SFR funds buy through channels, not from a text.
            skipped["tier 3 SFR fund"] += 1
            continue
        mail = b.get("mail") or ""
        if not mail:
            skipped["no mailing address"] += 1
            continue
        person = b["name"]
        source = "owner"
        if b["is_entity"]:
            rec = principals.get(b["buyer_key"]) or {}
            p = rec.get("principal")
            if not p:
                skipped["entity, no principal resolved"] += 1
                continue
            # A LOW-confidence principal is an agent-titled officer whose name
            # has nothing to do with the company, which usually means the
            # company's attorney. Skip tracing and then TEXTING a firm's lawyer
            # about a deal is the litigation bait this program exists to avoid,
            # so they are held back rather than dropped: --include-low-confidence
            # puts them in when a human wants to review them.
            # Agent-titled principals are a FALLBACK, not a disqualifier
            # (Ty, 2026-08-28): trace them, but the workbook flags the row so a
            # caller asks for the principal by name instead of assuming they
            # have the owner.
            pass
            # A BUSINESSV2 PRINCIPAL CANNOT BE SKIP TRACED AT THE COMPANY'S
            # ADDRESS. SmartSkip matches First + Last + Mailing Address, and a
            # reverse-address principal satisfies that BY CONSTRUCTION (we
            # found their name living at that address). An officer lifted from
            # a corporate filing has no established link to it. Measured
            # 2026-08-28 on two real batches: reverse-address principals hit
            # 163 of 172 (95%), BusinessV2 principals hit 1 of 26 (4%). They
            # need a name plus city/state person search instead, so they are
            # held out of the batch rather than burned on a 4% chance.
            # The old blanket exclusion of BusinessV2 principals is gone: its
            # reason was the company-address mismatch, which is now fixed by
            # tracing the officer's own address. One condition remains, since a
            # principal with no address of their own still has nothing to match.
            if (rec.get("principal_source") == "enformion-businessv2"
                    and not rec.get("principal_address")):
                skipped["BusinessV2 principal with no address of their own"] += 1
                continue
            person, source = p, "principal"
        first, last = clean_owner_name(person)
        if not first or not last:
            skipped["unparseable person name"] += 1
            continue

        # NAME ORDER. County filings and corporate records write "SMITH KENNETH
        # HAYWARD", not "Kenneth Smith", so clean_owner_name reads the surname
        # as the first name. When the person came from an entity we have a free
        # oracle: whichever of their tokens also appears in the COMPANY name is
        # the surname. Smithbilt LLC -> Smith Kenneth means Kenneth Smith.
        # Measured: 45 of these on the live registry, and each one would have
        # bought a SmartSkip lookup that could not match.
        if b["is_entity"]:
            ent_tokens = {w for w in norm_name(b["name"]).split() if len(w) > 2}
            if first.upper() in ent_tokens and last.upper() not in ent_tokens:
                first, last = last, first

        # A legal-status token is not a surname. "Morales Family Trust" resolved
        # to first=Morales last=Tr, which is a trustee marker, not a person.
        if first.upper() in STATUS_TOKENS or last.upper() in STATUS_TOKENS:
            skipped["name is a legal-status token, not a person"] += 1
            continue
        # A bare initial is not a first name. SmartSkip matches on first + last
        # + mailing address, so "E Bourgeois" buys a $0.15 lookup that either
        # misses or returns the wrong household. Same positional rule the SMS
        # side already learned the hard way with "E A Henry".
        if len(first.strip(".")) < 2:
            skipped["first name is only an initial"] += 1
            continue
        # A BusinessV2 principal has no established link to the COMPANY's
        # mailing address, which is why those rows skip traced at 4% while
        # reverse-address principals hit 95%. Use the officer's own address
        # when we have it; the company address is only correct for a
        # reverse-address principal, where it IS that person's home.
        addr_for_trace = mail
        if source == "principal":
            own = (principals.get(b["buyer_key"]) or {}).get("principal_address")
            if own:
                addr_for_trace = own
        street, city, st, zip_code = split_mail(addr_for_trace)
        # A "street" with no house number and no PO Box is a city line that lost
        # its street, which cannot be matched and would still be billed.
        if not re.search(r"\d", street) and not re.match(r"^P\.?\s?O", street, re.I):
            skipped["mailing address has no street"] += 1
            continue
        rows.append({
            "buyer_key": b["buyer_key"],
            "First Name": first,
            "Last Name": last,
            "Mailing Address": street,
            "Mailing City": city,
            "Mailing State": st,
            "Mailing Zip": zip_code,
            "entity": b["name"] if b["is_entity"] else "",
            "source": source,
            "n_buys": b["n_buys"],
        })

    path = OUT / "smartskip_input.csv"
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["buyer_key", "First Name", "Last Name",
                                           "Mailing Address", "Mailing City",
                                           "Mailing State", "Mailing Zip",
                                           "entity", "source", "n_buys"])
        w.writeheader()
        w.writerows(rows)

    gate(len(rows) > 0, "skipinput produced no traceable people")
    mark_done("skipinput", people=len(rows), skipped=dict(skipped))
    print("skipinput: " + str(len(rows)) + " people -> " + str(path))
    for reason, n in skipped.most_common():
        print("  skipped %-32s %s" % (reason, n))
    print("")
    print("SmartSkip is FREE up to the pay step. Next:")
    print("  python skills/deep-prospecting-v5/scripts/smartskip_trace.py \\")
    print("      submit --csv " + str(path))
    print("  (read the quoted row count and price, THEN pay)")


# --------------------------------------------------------------- profiles --
# What does this buyer actually buy. The registry says who buys here and the
# skip trace says how to reach them; neither answers the question a dispo
# caller has in front of them, which is whether THIS deal is THIS person's kind
# of deal. Ty chose deal blasts over a qualification drip, so the buy box has
# to be inferred from deed history rather than asked.

def _band(values: list) -> dict:
    """p10 / median / p90, never min / max.

    Opendoor's raw price_min reads $1,304 because one stray deed sits in its
    history, and a band anchored on that is worse than no band at all. The
    trimmed band is what a caller can actually say out loud.
    """
    vals = sorted(v for v in values if isinstance(v, (int, float)) and v > 0)
    if not vals:
        return {"lo": None, "mid": None, "hi": None, "n": 0}
    def pct(p):
        if len(vals) == 1:
            return vals[0]
        i = min(len(vals) - 1, max(0, int(round(p * (len(vals) - 1)))))
        return vals[i]
    return {"lo": pct(0.10), "mid": pct(0.50), "hi": pct(0.90), "n": len(vals)}


# A house does not sell for $3,000. Sales below this, or at a tiny fraction of
# the property's own value, are quitclaims, deeds in lieu and intra-family
# paper. They still happened, so they stay in the purchase COUNT, but they must
# never touch the price band: with a small sample the percentile trim cannot
# save it, and 13 buyers came back with bands starting at $1,000 to $4,594.
# Same principle as NON_ARMS_LENGTH_RATIO in post_walkthrough.
NOMINAL_SALE_FLOOR = 10000
NOMINAL_VALUE_RATIO = 0.20

SELF_PERFORM_RX = re.compile(
    r"\bCONSTRUCTION\b|\bBUILDERS?\b|\bRENOVATION|\bREMODEL|\bCONTRACTING\b", re.I)


PRICE_RECENT_DAYS = 730          # two years
PRICE_MIN_FOR_BAND = 3           # below this, widen rather than publish noise


def _price_band(obs: list) -> tuple:
    """(band, basis, nominal_dropped) preferring recent sales.

    A 2019 purchase is not what this buyer would pay today, so the band is
    built from the last two years first and only widens to full history when
    that leaves too few priced sales to say anything. The basis is returned so
    the sheet can state which one produced the number, the same way tight_arv()
    reports its comp basis rather than hiding the widen.
    """
    cut = _since(PRICE_RECENT_DAYS)
    recent = [o for o in obs if str(o.get("date") or "") >= cut]
    vals, dropped = _real_prices(recent)
    if len(vals) >= PRICE_MIN_FOR_BAND:
        return _band(vals), "recent 24mo", dropped
    vals_all, dropped_all = _real_prices(obs)
    if not vals_all:
        return _band([]), "no priced sales", dropped_all
    if len(vals) and len(vals_all) == len(vals):
        return _band(vals), "recent 24mo", dropped
    return _band(vals_all), "widened to full history", dropped_all


def _real_prices(obs: list) -> tuple:
    """(prices worth banding, how many nominal transfers were dropped)."""
    keep, dropped = [], 0
    for o in obs:
        p = o.get("price")
        if not isinstance(p, (int, float)) or p <= 0:
            continue
        val = o.get("estimatedValue")
        nominal = p < NOMINAL_SALE_FLOOR or (
            isinstance(val, (int, float)) and val > 0
            and p < val * NOMINAL_VALUE_RATIO)
        if nominal:
            dropped += 1
        else:
            keep.append(p)
    return keep, dropped


def buyer_type(b: dict) -> tuple:
    """(type, the rule that fired). Ordered, deterministic, auditable."""
    if b.get("tier") == "EXCLUDE":
        # Kept in the deliverable as market intelligence, labelled so nobody
        # calls them. A national homebuilder or the City of Knoxville is not a
        # dispo customer.
        return "not a target", b.get("tier_reason") or "excluded by taxonomy"
    if b.get("tier") in ("2", "3"):
        return "institutional", "registry tier " + b["tier"]
    if SELF_PERFORM_RX.search(b["name"] or ""):
        return "self-performer", "name says construction or building trade"
    txns = set(b.get("txn_types") or [])
    held, exited = b.get("n_held", 0), b.get("n_exited", 0)
    if exited > held:
        return "exited flipper", "more exits than current holdings"
    # THE RENTAL LABEL LAGS BY OVER A YEAR. Measured 2026-08-28: `rental`
    # returns 1,930 properties in Knox unbounded but ZERO within 12 months,
    # because SiftMap cannot classify a purchase as a rental until it observes
    # the property being rented, so a recent buy-and-hold sits under `pending`.
    # The old rule required txn_types == {"rental"} exactly, which deleted
    # every landlord the moment recency entered the picture: 111 became 0.
    # Any rental history plus a hold-heavy record is the durable signal.
    if "rental" in txns and held >= exited:
        return "landlord", "has rental history and holds at least as many as it exits"
    if "pending" in txns and held:
        return "active holder", "holds at least one pending investor purchase"
    # Wholesale describes how they ACQUIRED, not their model, so it only
    # decides the label when nothing about holding or exiting did.
    if txns & {"wholesale", "wholetail"}:
        return "wholesaler", "acquired on a wholesale or wholetail transaction"
    if "pending" in txns:
        return "active holder", "holds at least one pending investor purchase"
    return "other", "no rule matched"


def _money(v) -> str:
    if not v:
        return ""
    v = int(v)
    if v >= 1_000_000:
        return "$%.1fM" % (v / 1_000_000.0)
    if v >= 1000:
        return "$%dK" % round(v / 1000.0)
    return "$%d" % v


def _narrative(p: dict) -> str:
    """One sentence, every figure of which is a column on the same row.

    Generated, not modelled. 698 rows of pure arithmetic is a poor use of a
    model, and a model asked to summarise a buyer will eventually invent a
    detail that a caller then repeats to that buyer on the phone.
    """
    who = p["display_name"]
    where = p["top_cities"][0][0] if p["top_cities"] else (p["counties"][0]
                                                           if p["counties"] else "the area")
    kind = {"landlord": "rentals", "exited flipper": "flips",
            "self-performer": "projects", "wholesaler": "wholesale deals",
            "institutional": "homes", "active holder": "homes",
            "not a target": "homes",
            "other": "homes"}.get(p["buyer_type"], "homes")
    bits = ["%s buys %s %s" % (who, where, kind)]
    # NOT "in 12 months". The SiftMap filter selects properties whose LAST SALE
    # was an investor transaction, with no date bound at all, so the observed
    # data actually spans 2019 to 2026 and the median buyer last bought 456
    # days ago. Saying 12 months put a false window on every one of these.
    span = ""
    if p["first_buy"] and p["last_buy"]:
        y0, y1 = p["first_buy"][:4], p["last_buy"][:4]
        span = (" in %s" % y0) if y0 == y1 else (" from %s to %s" % (y0, y1))
    bits.append("%d purchase%s%s" % (p["n_buys"],
                                     "" if p["n_buys"] == 1 else "s", span))
    price = p["box_price"]
    if price["lo"] and price["hi"] and price["n"] >= 3:
        bits.append("typically %s to %s" % (_money(price["lo"]), _money(price["hi"])))
    elif price["mid"]:
        # Under three real sales there is no band to speak of, so say the one
        # number we have rather than dress a single point up as a range.
        bits.append("only %d priced sale%s on record, around %s"
                    % (price["n"], "" if price["n"] == 1 else "s",
                       _money(price["mid"])))
    beds, baths, sqft = p["box_beds"], p["box_baths"], p["box_sqft"]
    shape = []
    if beds["mid"]:
        shape.append("%g bed" % beds["mid"])
    if baths["mid"]:
        shape.append("%g bath" % baths["mid"])
    if shape:
        s = " / ".join(shape)
        if sqft["mid"]:
            s += " around %s sqft" % format(int(sqft["mid"]), ",")
        bits.append(s)
    if p["cash_pct"] is not None:
        bits.append("%d%% cash" % p["cash_pct"])
    if p["n_held"] and p["n_exited"]:
        bits.append("holds %d and has exited %d" % (p["n_held"], p["n_exited"]))
    elif p["n_exited"] and not p["n_held"]:
        bits.append("has exited every one")
    elif p["n_held"] and not p["n_exited"]:
        bits.append("holds rather than flips")
    out = bits[0] + ": " + ", ".join(bits[1:]) + "."
    if p["days_since_last"] is not None:
        out += " Last bought %d days ago." % p["days_since_last"]
    return out


def phase_profiles(args) -> None:
    require("aggregate")
    from datetime import date

    reg = _load(OUT / "registry.json", [])
    principals = _load(OUT / "principals.json", {})
    phones = _load(OUT / "phones.json", {})
    scores = _load(OUT / "phone_scores.json", {})
    rows = _load(OUT / "rows.json", {})
    details = _load(OUT / "details.json", {})
    obs = _observations(rows, details)

    by_key: dict = defaultdict(list)
    for o in obs:
        by_key[norm_name(o["name"])].append(o)

    GOOD = ("Dial First", "Dial Second")
    today = date.today()
    profiles = []
    skipped_inactive = 0
    for b in reg:
        if b["n_buys"] < args.profile_min_buys:
            continue
        # Recency decides WHO is on the list; the unbounded history decides
        # what they buy. Without this the cohort spans 2019 to 2026 and half of
        # it stopped buying years ago.
        if args.require_active and not b.get("active_365d"):
            skipped_inactive += 1
            continue
        mine = by_key.get(b["buyer_key"]) or []
        for alias in b.get("aliases") or []:
            mine = mine + (by_key.get(alias) or [])
        if not mine:
            mine = []

        pr = principals.get(b["buyer_key"]) or {}
        ph = phones.get(b["buyer_key"]) or {}
        numbers = []
        for p in (ph.get("phones") or []):
            tier = (scores.get(p["number"]) or {}).get("tier")
            numbers.append({"number": p["number"], "type": p.get("type") or "",
                            "tier": tier or "Unscored"})
        srcs = {(q.get("source") or "smartskip") for q in (ph.get("phones") or [])}
        phone_src = "both" if any("+" in s for s in srcs) or len(srcs) > 1 else             (srcs.pop() if srcs else "")
        name_known = bool(ph.get("person")) or bool(pr.get("principal")) or not b["is_entity"]
        numbers.sort(key=lambda n: (GOOD + ("Dial Third", "Dial Fourth",
                                            "Drop", "Unscored")).index(n["tier"])
                     if n["tier"] in GOOD + ("Dial Third", "Dial Fourth", "Drop",
                                             "Unscored") else 9)
        best = numbers[0]["tier"] if numbers else ""

        cities = Counter()
        zips = Counter()
        for o in mine:
            _st, city, _s, zc = split_mail(o.get("prop") or "")
            if city:
                cities[city] += 1
            if zc:
                zips[zc] += 1
        dates = sorted(str(o.get("date") or "") for o in mine if o.get("date"))
        last = dates[-1] if dates else ""
        days = None
        if last:
            try:
                days = (today - datetime.fromisoformat(last[:10]).date()).days
            except ValueError:
                days = None
        cash_pct = None
        if mine:
            cash_pct = int(round(100.0 * sum(1 for o in mine if o.get("cash")) / len(mine)))
        distress = Counter(d for o in mine for d in (o.get("distress") or []))

        _pb = _price_band(mine)
        p = {
            "buyer_key": b["buyer_key"],
            "name": b["name"],
            # ONLY a high-confidence principal earns the headline. A
            # LOW-confidence one is an agent-titled officer whose name has
            # nothing to do with the company, i.e. usually their attorney, and
            # putting "Richard E Pardall buys Knoxville flips" on a call sheet
            # tells a caller to greet GDP Properties' lawyer as the buyer.
            "display_name": (pr.get("principal")
                             if (b["is_entity"]
                                 and pr.get("principal")
                                 and pr.get("principal_confidence") == "high")
                             else b["name"]),
            "is_entity": b["is_entity"],
            "entity_name": b["name"] if b["is_entity"] else "",
            "principal": pr.get("principal") or "",
            "principal_source": pr.get("principal_source") or "",
            "principal_confidence": pr.get("principal_confidence") or "",
            "mail": b["mail"],
            "counties": b["counties"],
            "tier": b["tier"],
            "n_buys": b["n_buys"],
            "n_held": b["n_held"],
            "n_exited": b["n_exited"],
            "cash_pct": cash_pct,
            "txn_types": b["txn_types"],
            "first_buy": dates[0] if dates else "",
            "last_buy": last,
            "days_since_last": days,
            "active_180d": bool(days is not None and days <= 180),
            "active_365d": bool(b.get("active_365d")),
            "active_182d": bool(b.get("active_182d")),
            "n_in_window": b.get("n_in_window") or 0,
            "last_active_buy": b.get("last_active_buy") or "",
            # Rate over the span actually observed for THIS buyer, not over an
            # assumed 12-month window that does not exist in this data.
            "span_days": ((datetime.fromisoformat(dates[-1][:10]).date()
                           - datetime.fromisoformat(dates[0][:10]).date()).days
                          if len(dates) > 1 else 0),
            "buys_per_year": (round(b["n_buys"] / max(
                0.25, ((datetime.fromisoformat(dates[-1][:10]).date()
                        - datetime.fromisoformat(dates[0][:10]).date()).days / 365.0)), 1)
                              if len(dates) > 1 else None),
            "portfolio_n": b["portfolio_n"],
            "portfolio_value": b["portfolio_value"],
            "box_price": _pb[0],
            "price_basis": _pb[1],
            "nominal_dropped": _pb[2],
            "box_value": _band([o.get("estimatedValue") for o in mine]),
            "box_beds": _band([o.get("bedrooms") for o in mine]),
            "box_baths": _band([o.get("bathrooms") for o in mine]),
            "box_sqft": _band([o.get("squareFeet") for o in mine]),
            "box_lot": _band([o.get("lotAcres") for o in mine]),
            "box_equity": _band([o.get("equityPercent") for o in mine]),
            "top_cities": cities.most_common(4),
            "top_zips": zips.most_common(4),
            "distress_appetite": distress.most_common(3),
            "phones": numbers,
            "best_tier": best,
            "phone_source": phone_src,
            "name_known": name_known,
            "reachable": best in GOOD,
            "saved_uuid": b["saved_uuid"],
        }
        # Ty allowed an agent-titled officer as a FALLBACK principal, so the
        # sheet has to say so: a caller who thinks they have the owner will
        # open with the wrong assumption on a lawyer's line.
        p["principal_is_agent"] = bool(
            pr.get("principal")
            and pr.get("principal_confidence") == "low")
        p["buyer_type"], p["buyer_type_rule"] = buyer_type(b)
        p["narrative"] = _narrative(p)
        profiles.append(p)

    profiles.sort(key=lambda x: (-x["n_buys"], x["name"]))
    _save(OUT / "buyer_profiles.json", profiles)
    gate(len(profiles) > 0, "no profiles built")
    types = Counter(p["buyer_type"] for p in profiles)
    reach = sum(1 for p in profiles if p["reachable"])
    mark_done("profiles", buyers=len(profiles), reachable=reach,
              by_type=dict(types), min_buys=args.profile_min_buys)
    print("profiles: %d buyers with %d+ purchases%s"
          % (len(profiles), args.profile_min_buys,
             (" and active in the last %d days" % args.recent_days)
             if args.require_active else ""))
    if skipped_inactive:
        print("  dropped as no longer active      : %d" % skipped_inactive)
    basis = Counter(p["price_basis"] for p in profiles)
    print("  price basis: " + str(dict(basis)))
    print("  reachable now (Dial First/Second): %d" % reach)
    for t, n in types.most_common():
        print("   %-18s %4d" % (t, n))
    xlsx = OUT / ("Knox_Blount_Buyer_Profiles_%s.xlsx"
                  % datetime.now().strftime("%Y%m%d"))
    try:
        written = write_profiles_xlsx(profiles, xlsx) or xlsx
        print("  workbook: %s" % written)
        if written != xlsx:
            print("  (the original is open in Excel; close it and rename this file)")
    except ImportError:
        print("  openpyxl not installed, JSON written but no workbook")
    print()
    for p in profiles[:5]:
        print("  " + p["narrative"][:150])


# ------------------------------------------------------- profiles: excel --

NAVY, BLUE, GREEN, GOLD, RED = "0A1130", "316AFF", "1B9E5A", "B8860B", "C00000"


def _fmt_band(band: dict, money: bool = False, dec: int = 0) -> str:
    if not band or not band.get("mid"):
        return ""
    if band["n"] < 3:
        v = band["mid"]
        return (_money(v) if money else ("%.*f" % (dec, v))) + " (1 sale)" \
            if band["n"] == 1 else \
            (_money(v) if money else ("%.*f" % (dec, v))) + " (thin)"
    lo, hi = band["lo"], band["hi"]
    if money:
        return "%s to %s" % (_money(lo), _money(hi))
    return "%.*f to %.*f" % (dec, lo, dec, hi)


def write_profiles_xlsx(profiles: list, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor=NAVY)
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(color=NAVY, bold=True, size=14)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()

    def sheet(name, headers, rows, widths, freeze="A3", title=""):
        ws = wb.create_sheet(name)
        ws["A1"] = title or name
        ws["A1"].font = title_font
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=i, value=h)
            c.fill, c.font, c.border = hdr_fill, hdr_font, border
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for r, row in enumerate(rows, 3):
            for i, v in enumerate(row, 1):
                c = ws.cell(row=r, column=i, value=v)
                c.border = border
                c.alignment = wrap
                if isinstance(v, (int, float)):
                    c.alignment = Alignment(vertical="top", horizontal="right")
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = freeze
        ws.auto_filter.ref = "A2:%s%d" % (get_column_letter(len(headers)),
                                          max(2, len(rows) + 2))
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        return ws

    # ---- 1. Overview ----------------------------------------------------
    ov = wb.active
    ov.title = "Overview"
    ov["A1"] = "Knox and Blount cash buyer profiles"
    ov["A1"].font = title_font
    types = Counter(p["buyer_type"] for p in profiles)
    lines = [
        ("Buyers profiled", len(profiles)),
        ("Minimum purchases to qualify", 2),
        ("Cohort", "active in the last 12 months AND 2+ lifetime purchases"),
        ("Bought within the last year",
         sum(1 for p in profiles if p.get("active_365d"))),
        ("Bought within the last 180 days",
         sum(1 for p in profiles if p["active_180d"])),
        ("Reachable today (Dial First or Second)",
         sum(1 for p in profiles if p["reachable"])),
        ("Traced but no good number",
         sum(1 for p in profiles if p["phones"] and not p["reachable"])),
        ("Not yet skip traced", sum(1 for p in profiles if not p["phones"])),
        ("", ""),
        ("How the buy box is built", ""),
        ("", "Inferred from 12 months of deed history. Nobody was asked."),
        ("", "Price shown as a p10 to p90 band, never raw min to max."),
        ("", "Sales under $10,000 or under 20% of value are treated as"),
        ("", "nominal transfers (quitclaims, deeds in lieu) and excluded"),
        ("", "from the band, though they still count as a purchase."),
        ("", "A buyer with fewer than 3 priced sales gets a single figure"),
        ("", "marked thin, not a range."),
        ("", ""),
        ("What is NOT here, and why", ""),
        ("", "Year built: this endpoint does not return it at all."),
        ("", "Property type: constant, the sweep filters single family."),
        ("", "Portfolio size is unreliable for large entities."),
        ("", "Exited flippers have no mailing address, so no contact."),
        ("", ""),
        ("Coverage of the target cohort", ""),
        ("   (excludes institutional and not-a-target)", ""),
        ("   reachable now", sum(1 for p in profiles if p["reachable"]
                                 and p["buyer_type"] not in ("institutional", "not a target"))),
        ("   traced, every number scored low",
         sum(1 for p in profiles if p["phones"] and not p["reachable"]
             and p["buyer_type"] not in ("institutional", "not a target"))),
        ("   no verified principal yet",
         sum(1 for p in profiles if not p["phones"] and p["is_entity"]
             and not p["principal"]
             and p["buyer_type"] not in ("institutional", "not a target"))),
        ("   principal is an agent, verify before calling",
         sum(1 for p in profiles if p.get("principal_is_agent"))),
        ("", ""),
        ("Buyer types", ""),
    ]
    for t, n in types.most_common():
        lines.append(("   " + t, n))
    for r, (k, v) in enumerate(lines, 3):
        ov.cell(row=r, column=1, value=k).font = Font(bold=bool(v == "" and k))
        ov.cell(row=r, column=2, value=v)
    ov.column_dimensions["A"].width = 42
    ov.column_dimensions["B"].width = 66
    for r in range(3, len(lines) + 3):
        ov.cell(row=r, column=2).alignment = wrap

    # ---- 2. Buyer Profiles ----------------------------------------------
    heads = ["Who", "Buyer type", "Buys", "Held", "Exited", "Cash %",
             "Price band", "Value band", "Beds", "Baths", "Sqft", "Lot acres",
             "Where they buy", "Counties", "First bought", "Last bought",
             "Buys per year", "Active 1y", "Hot 6mo", "Bought in window",
             "Price basis",
             "Best phone tier", "Phone source", "Name known?", "Phones",
             "Entity", "Principal",
             "Principal is an agent",
             "Principal source", "Mailing address", "Summary"]
    rows = []
    for p in profiles:
        rows.append([
            p["display_name"], p["buyer_type"], p["n_buys"], p["n_held"],
            p["n_exited"], p["cash_pct"],
            _fmt_band(p["box_price"], money=True),
            _fmt_band(p["box_value"], money=True),
            _fmt_band(p["box_beds"]), _fmt_band(p["box_baths"], dec=1),
            _fmt_band(p["box_sqft"]), _fmt_band(p["box_lot"], dec=2),
            ", ".join("%s (%d)" % (c, n) for c, n in p["top_cities"]),
            ", ".join(p["counties"]), p["first_buy"], p["last_buy"],
            p["buys_per_year"], "yes" if p["active_365d"] else "",
            "yes" if p.get("active_182d") else "", p.get("n_in_window") or 0,
            p["price_basis"],
            p["best_tier"], p.get("phone_source", ""),
            "" if p.get("name_known") else "NO NAME",
            ", ".join(n["number"] for n in p["phones"][:4]),
            p["entity_name"], p["principal"],
            "VERIFY" if p.get("principal_is_agent") else "",
            p["principal_source"],
            p["mail"], p["narrative"],
        ])
    sheet("Buyer Profiles", heads, rows,
          [26, 15, 7, 7, 8, 8, 20, 20, 12, 12, 16, 14, 30, 14, 12, 12, 12, 9,
           8, 10, 20, 14, 18, 12, 30, 28, 18, 22, 22, 34, 78],
          title="Every buyer with 2 or more purchases in the last 12 months")

    # ---- 3. Call List ---------------------------------------------------
    order = {"Dial First": 0, "Dial Second": 1}
    call = sorted([p for p in profiles if p["reachable"]
                   and p["buyer_type"] != "not a target"],
                  key=lambda p: (order.get(p["best_tier"], 9), -p["n_buys"]))
    ch = ["Who", "Agent?", "Name known?", "Buyer type", "Buys", "Price band",
          "Beds", "Sqft", "Where they buy", "Tier", "Phone 1", "Phone 2",
          "Phone 3", "Summary"]
    crows = [[p["display_name"],
              "VERIFY" if p.get("principal_is_agent") else "",
              "" if p.get("name_known") else "ASK WHO",
              p["buyer_type"], p["n_buys"],
              _fmt_band(p["box_price"], money=True), _fmt_band(p["box_beds"]),
              _fmt_band(p["box_sqft"]),
              ", ".join(c for c, _ in p["top_cities"][:3]), p["best_tier"]]
             + [(p["phones"][i]["number"] if len(p["phones"]) > i else "")
                for i in range(3)]
             + [p["narrative"]] for p in call]
    sheet("Call List", ch, crows,
          [26, 9, 12, 15, 7, 20, 12, 16, 28, 13, 14, 14, 14, 78],
          title="Reachable buyers, best dial tier first")

    # ---- 4. By Type -----------------------------------------------------
    th = ["Buyer type", "Buyers", "Reachable", "Median buys",
          "Median price (mid)", "Median sqft", "Typical rule that fired"]
    trows = []
    for t, n in types.most_common():
        grp = [p for p in profiles if p["buyer_type"] == t]
        mids = sorted(p["box_price"]["mid"] for p in grp if p["box_price"]["mid"])
        sq = sorted(p["box_sqft"]["mid"] for p in grp if p["box_sqft"]["mid"])
        bys = sorted(p["n_buys"] for p in grp)
        trows.append([t, n, sum(1 for p in grp if p["reachable"]),
                      bys[len(bys) // 2] if bys else "",
                      _money(mids[len(mids) // 2]) if mids else "",
                      int(sq[len(sq) // 2]) if sq else "",
                      grp[0]["buyer_type_rule"] if grp else ""])
    sheet("By Type", th, trows, [18, 10, 12, 13, 20, 14, 46],
          title="How the cohort splits, and the rule behind each label")

    # ---- 5. Sources and Caveats -----------------------------------------
    sh = ["Item", "Detail"]
    srows = [
        ["Source", "SiftMap /properties/search/ and /properties/detail/, "
                   "Knox (47093) and Blount (47009), single family only"],
        ["Window", "NOT a fixed window. The filter selects properties whose "
                   "LAST SALE was an investor transaction, with no date "
                   "bound, so observed purchases run from 2019 to 2026. "
                   "Only 41% of these buyers bought within the last year "
                   "and the median last purchase is 456 days old. Use the "
                   "Active and Last bought columns before calling"],
        ["Transaction types", "pending, wholesale, wholetail, rental (current "
                              "owner is the buyer) and flip (the buyer is the "
                              "last-sale SELLER, since the exit already happened)"],
        ["Dedupe", "Mailing-address clustering with a suite guard: an address "
                   "carrying STE/UNIT/APT needs a fuzzy name match too, "
                   "because unrelated firms share office suites"],
        ["Principals", "Free reverse-address unmask first, then Enformion "
                       "BusinessV2 for repeat buyers only. An agent-titled "
                       "officer whose name does not appear in the company name "
                       "is treated as low confidence and is NOT shown as the "
                       "buyer, because that is usually their attorney"],
        ["Phones", "SmartSkip on buyers with 2+ purchases, Trestle scored"],
        ["Dial tiers", "81-100 Dial First, 61-80 Dial Second, 41-60 Third, "
                       "21-40 Fourth, 0-20 Drop"],
        ["Price band", "p10 to p90 of real sales. Nominal transfers under "
                       "$10,000 or under 20% of value are excluded"],
        ["Not a target", "Homebuilders, government and nonprofits are kept for "
                         "market context and labelled so nobody calls them"],
        ["Caution", "The buy box is inferred, never asked. Treat it as where "
                    "this buyer has been active, not a commitment"],
    ]
    sheet("Sources and Caveats", sh, srows, [22, 104],
          title="How to read this, and what not to trust")

    # Excel holds an EXCLUSIVE lock on an open workbook, so saving over one the
    # reviewer is looking at raises PermissionError and loses the run. Write to
    # a pending name and swap; if the swap is also blocked, keep the pending
    # file and say so rather than pretending the save worked.
    path.parent.mkdir(parents=True, exist_ok=True)
    pending = path.with_name("_PENDING_" + path.name)
    wb.save(pending)
    try:
        pending.replace(path)
        return path
    except PermissionError:
        log.warning("%s is open in Excel; left the new copy at %s",
                    path.name, pending.name)
        return pending


# ----------------------------------------------------------------- phones --
# Join the SmartSkip return back onto the registry. The download is keyed by
# name and mailing address, not by our buyer_key, so the input CSV is the
# bridge: it carries both.

def phase_phones(args) -> None:
    require("skipinput")
    import csv
    sys.path.insert(0, os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "skills", "deep-prospecting-v5", "scripts"))
    import parse_smartskip as ps

    vert = OUT / "smartskip_vertical.csv"
    if not vert.exists():
        raise SystemExit("no " + str(vert) + "; download the batch first")

    # Build the bridge from the REGISTRY, not from smartskip_input.csv. That
    # file is regenerated on every skipinput run, so a download from an earlier
    # batch loses its join the moment the cohort changes.
    from enformion_ftm import clean_owner_name
    reg = _load(OUT / "registry.json", [])
    principals = _load(OUT / "principals.json", {})
    bridge = {}
    name_only: dict = {}
    for b in reg:
        person = b["name"]
        if b["is_entity"]:
            person = (principals.get(b["buyer_key"]) or {}).get("principal") or ""
        if not person or not b.get("mail"):
            continue
        first, last = clean_owner_name(person)
        if not first or not last:
            continue
        ent_tokens = {w for w in norm_name(b["name"]).split() if len(w) > 2}
        if b["is_entity"] and first.upper() in ent_tokens and last.upper() not in ent_tokens:
            first, last = last, first
        # Index BOTH candidate addresses. skipinput traces a BusinessV2
        # principal at the officer's OWN address, so a bridge built only from
        # the company mailing address silently fails to match those rows back.
        rec = principals.get(b["buyer_key"]) or {}
        for cand in (b.get("mail"), rec.get("principal_address")):
            if not cand:
                continue
            street, _c, _s, _z = split_mail(cand)
            if street:
                bridge[(first.strip().upper(), last.strip().upper(),
                        street.strip().upper())] = b["buyer_key"]
        # Name-only fallback, used when neither address lines up. Ambiguous
        # names are dropped rather than guessed.
        nk = (first.strip().upper(), last.strip().upper())
        name_only[nk] = None if nk in name_only else b["buyer_key"]

    # Merge every downloaded batch, not just the newest one.
    verts = sorted(f for f in OUT.glob("smartskip*.csv")
                   if f.name != "smartskip_input.csv")

    recs = []
    for v in verts:
        recs.extend(ps.parse(str(v)))
    log.info("parsed %s records across %s downloaded batches", len(recs), len(verts))
    out: dict = _load(OUT / "phones.json", {})
    matched = unmatched = 0
    for rec in recs:
        k = ((rec.get("first") or "").strip().upper(),
             (rec.get("last") or "").strip().upper(),
             (rec.get("mailing_address") or "").strip().upper())
        key = bridge.get(k) or name_only.get((k[0], k[1]))
        if not key:
            unmatched += 1
            continue
        matched += 1
        phones = []
        for p in (rec.get("subject_phones") or []):
            num = re.sub(r"\D", "", str(p.get("number") if isinstance(p, dict) else p))
            if len(num) == 11 and num.startswith("1"):
                num = num[1:]
            if len(num) != 10:
                continue
            phones.append({"number": num,
                           "type": (p.get("type") if isinstance(p, dict) else "") or ""})
        if not phones and not rec.get("relatives"):
            continue
        out[key] = {
            "person": ((rec.get("first") or "") + " " + (rec.get("last") or "")).strip(),
            "phones": phones,
            # SmartSkip is WRONG about death often enough that this is a flag to
            # check, never a fact to act on. It is carried so a human can look.
            "deceased_flag": bool(rec.get("deceased")),
            "relatives": len(rec.get("relatives") or []),
        }

    _save(OUT / "phones.json", out)
    uniq = {p["number"] for v in out.values() for p in v["phones"]}
    gate(matched > 0, "no SmartSkip record joined back to a buyer_key")
    mark_done("phones", joined=matched, unjoined=unmatched,
              buyers_with_phones=len(out), unique_numbers=len(uniq))
    print("phones: joined %d of %d records (%d could not be matched back)"
          % (matched, len(recs), unmatched))
    print("  buyers now carrying a phone : %d" % len(out))
    print("  unique numbers              : %d" % len(uniq))
    print("  flagged deceased (VERIFY)   : %d"
          % sum(1 for v in out.values() if v["deceased_flag"]))
    print("  Trestle cost to score them  : $%.2f" % (len(uniq) * 0.015))


# ------------------------------------------------------------- crm-phones --
# THE NUMBERS WERE ALREADY THERE. A comment in dispo_flip_buyers said DataSift's
# native skip trace "covers humans; the entity principal is the part it cannot
# do", and that was taken as established rather than measured. It is wrong.
# Measured 2026-08-28 across the 128 unreachable target buyers holding a CRM
# record: 107 of them already carry phones, 514 rows and 316 unique mobiles, on
# records whose owner is the LLC itself. Only 2 had ever been Trestle scored.
#
# So this phase buys nothing. It reads what the account already holds, which is
# the step that should have come before any vendor call. An assertion in a
# comment is not a measurement.


def phase_crm_phones(args) -> None:
    require("profiles")
    import time
    from sms_agent import crm
    from sms_agent.seed import SKIP_PHONE_STATUSES

    profiles = _load(OUT / "buyer_profiles.json", [])
    reg = {b["buyer_key"]: b for b in _load(OUT / "registry.json", [])}
    out = _load(OUT / "phones.json", {})

    targets = [p for p in profiles if reg.get(p["buyer_key"], {}).get("saved_uuid")]
    if args.limit:
        targets = targets[:args.limit]
    log.info("reading %s CRM records for numbers already on file", len(targets))

    stats = Counter()
    added_buyers = 0
    for i, p in enumerate(targets, 1):
        key = p["buyer_key"]
        rec = _get_record(crm, reg[key]["saved_uuid"])
        if not rec:
            stats["record unreadable after retries"] += 1
            continue
        owner = rec.get("owner") or {}
        found = []
        for q in (owner.get("phones") or rec.get("phones") or []):
            if not isinstance(q, dict) or not q.get("number"):
                continue
            num = re.sub(r"\D", "", str(q["number"]))
            if len(num) == 11 and num.startswith("1"):
                num = num[1:]
            if len(num) != 10:
                stats["not a 10 digit number"] += 1
                continue
            # A disposition somebody set by hand outranks anything this
            # harvest wants to do. Same set the seeder suppresses on.
            if (q.get("status") or "").upper() in SKIP_PHONE_STATUSES:
                stats["suppressed by its existing disposition"] += 1
                continue
            found.append({"number": num, "type": (q.get("type") or "").upper(),
                          "status": (q.get("status") or "UNKNOWN").upper(),
                          "is_connected": bool(q.get("is_connected")),
                          "source": "datasift-native"})
        if not found:
            stats["record has no usable number"] += 1
            continue

        entry = out.get(key)
        if entry is None:
            entry = {"person": "", "phones": [], "deceased_flag": False,
                     "relatives": 0}
            out[key] = entry
            added_buyers += 1
        # Merge, never overwrite. A number SmartSkip already returned keeps its
        # place and simply gains a second source.
        have = {q["number"]: q for q in entry["phones"]}
        for q in found:
            if q["number"] in have:
                src = have[q["number"]].get("source") or "smartskip"
                if "datasift-native" not in src:
                    have[q["number"]]["source"] = src + "+datasift-native"
                stats["already had this number"] += 1
            else:
                entry["phones"].append(q)
                stats["number added"] += 1
        # The CRM record's owner is the COMPANY, so a harvested number comes
        # with no person attached. The caller needs to know that.
        entry.setdefault("name_known", bool(entry.get("person")))
        if i % 40 == 0:
            _save(OUT / "phones.json", out)
            log.info("read %s/%s records", i, len(targets))

    _save(OUT / "phones.json", out)
    uniq = {q["number"] for v in out.values() for q in v["phones"]}
    gate(stats["number added"] > 0 or added_buyers == 0,
         "harvest read records but added no numbers; check crm auth")
    mark_done("crm_phones", records=len(targets), buyers_added=added_buyers,
              **{k.replace(" ", "_"): v for k, v in stats.items()})
    print("crm-phones: read %d records" % len(targets))
    for k, v in stats.most_common():
        print("  %-42s %d" % (k, v))
    print("  buyers newly carrying a phone            %d" % added_buyers)
    print("  buyers with phones overall               %d" % len(out))
    print("  unique numbers overall                   %d" % len(uniq))


# ------------------------------------------------------------------ score --
# Trestle scoring, and the phase without which the SMS side cannot send a
# single message. Measured live 2026-08-28: of 165 records in "Dispo - 02 Ready
# to Text", exactly ONE carried a dial tier, because `dispo_flip_buyers --phase
# trace` writes phones tagged "dispo" and never scores them. `seed.from_preset`
# gates on the literal tier names, so an unscored cohort is an empty cohort and
# it looks exactly like a preset that matched nobody.

def _phone_objects(rec: dict) -> list:
    owner = rec.get("owner") or {}
    phones = owner.get("phones") or rec.get("phones") or []
    return [p for p in phones if isinstance(p, dict) and p.get("number")]


def _is_dnc(p: dict) -> bool:
    return bool(p.get("doNotCall") or p.get("do_not_call"))


def _get_record(crm, uuid: str, tries: int = 5):
    """get_record with backoff.

    /api/internal throttles hard and `crm.get_record` swallows the 429, logs a
    warning and returns nothing. Left alone that silently drops a record's
    entire phone list from the scoring set while the run still reports success,
    which is the exact failure this codebase keeps rediscovering.
    """
    import time
    for attempt in range(tries):
        rec = crm.get_record(uuid)
        if rec:
            return rec
        time.sleep(1.5 * (attempt + 1))
    return None


def _tag_names(p: dict) -> list:
    out = []
    for t in (p.get("tags") or []):
        if isinstance(t, dict):
            t = t.get("title") or t.get("name") or t.get("tag")
        if t:
            out.append(str(t))
    return out


def _score_numbers(pv, numbers: list, cache: dict, cache_path: Path,
                   commit: bool) -> dict:
    """Trestle-score whatever is not already cached. Returns the cache."""
    unscored = [n for n in numbers if n not in cache]
    print("  unique numbers needing a score      %d" % len(unscored))
    print("  estimated Trestle cost              $%.2f" % (len(unscored) * 0.015))
    if not commit:
        print("\nDRY RUN. Re-run with --commit to score.")
        return cache
    key = os.getenv("TRESTLE_PAID_API_KEY") or os.getenv("TRESTLE_API_KEY")
    if not key:
        raise SystemExit("no TRESTLE_API_KEY / TRESTLE_PAID_API_KEY in the environment")
    for i, num in enumerate(unscored, 1):
        try:
            data = pv.call_trestle(num, key)
        except Exception as e:  # noqa: BLE001
            cache[num] = {"error": str(e)[:120]}
            continue
        score = data.get("activity_score") if isinstance(data, dict) else None
        cache[num] = {"score": score, "tier": pv.assign_tier(score, pv.DEFAULT_TIERS)}
        if i % 50 == 0:
            _save(cache_path, cache)
            log.info("scored %s/%s", i, len(unscored))
    _save(cache_path, cache)
    return cache


def phase_score(args) -> None:
    """Score every textable dispo phone and write its dial tier back."""
    import phone_validator as pv

    cache_path = OUT / "phone_scores.json"
    cache = _load(cache_path, {})

    # The registry source scores numbers that are not in the CRM yet, so the
    # mirror can write them with their tier already attached instead of
    # landing records the seeder will refuse to text.
    if args.source == "registry":
        phones = _load(OUT / "phones.json", {})
        gate(bool(phones), "no phones.json; run --phase phones first")
        nums = sorted({p["number"] for v in phones.values() for p in v["phones"]})
        print("registry: %d buyers, %d unique numbers" % (len(phones), len(nums)))
        cache = _score_numbers(pv, nums, cache, cache_path, args.commit)
        if not args.commit:
            return
        got = [n for n in nums if (cache.get(n) or {}).get("tier")]
        by_tier = Counter((cache[n] or {}).get("tier") for n in got)
        mark_done("score_registry", scored=len(got), by_tier=dict(by_tier))
        print("score: %d numbers scored" % len(got))
        print("  by tier: " + str(dict(by_tier)))
        reachable = sum(1 for v in phones.values()
                        if any((cache.get(p["number"]) or {}).get("tier")
                               in ("Dial First", "Dial Second") for p in v["phones"]))
        print("  buyers with a Dial First/Second number: %d of %d"
              % (reachable, len(phones)))
        return

    from sms_agent import crm

    tiers = pv.DEFAULT_TIERS
    tier_names = set(tiers)

    must, matched = crm.resolve_preset(args.preset)
    if not must:
        raise SystemExit("preset not found: " + args.preset)
    log.info("cohort preset %r", matched)

    records = list(crm.fetch_cohort(must, limit=args.limit or 0))
    log.info("%s records in the cohort", len(records))
    gate(len(records) > 0, "cohort preset matched no records")

    # THE DNC FLAG LIVES ONLY ON THE SEARCH ROW. Verified live 2026-08-28: a
    # number returned as doNotCall=true by /property/ search comes back from
    # /property/{uuid}/ with the field ABSENT, not false. So the do-not-call
    # set has to be built here, from the search rows, or the filter below can
    # never fire and we pay Trestle to score numbers the seeder will refuse to
    # text anyway. Coverage is partial by construction: the search returns one
    # representative phone per record, so DNC status is unknown for the rest.
    dnc_numbers = set()
    for rec in records:
        p = rec.get("phone") if isinstance(rec.get("phone"), dict) else {}
        if p.get("doNotCall"):
            n = pv.clean_phone(p.get("number") or "")
            if n:
                dnc_numbers.add(n)
    log.info("%s numbers flagged do-not-call on the search rows", len(dnc_numbers))

    # ---- collect the numbers worth scoring ------------------------------
    targets: dict = {}
    stats = Counter()
    unreadable = 0
    for rec in records:
        uuid = rec.get("uuid")
        full = _get_record(crm, uuid)
        if not full:
            unreadable += 1
            continue
        for p in _phone_objects(full):
            num = pv.clean_phone(p.get("number") or "")
            if not num:
                continue
            stats["phones seen"] += 1
            if (num in dnc_numbers or _is_dnc(p)) and not args.include_dnc:
                stats["skipped, on the do-not-call list"] += 1
                continue
            if (p.get("type") or "").upper() == "LANDLINE":
                stats["skipped, landline"] += 1
                continue
            if set(_tag_names(p)) & tier_names:
                stats["already tiered"] += 1
                continue
            targets.setdefault(num, []).append((uuid, p))
            stats["to score"] += 1

    unscored = [n for n in targets if n not in cache]
    print("cohort: %d records" % len(records))
    if unreadable:
        # Never let a throttled read look like a record with no phones.
        print("  UNREADABLE after retries              %d" % unreadable)
    for k, v in stats.most_common():
        print("  %-38s %d" % (k, v))
    print("  unique numbers needing a score      %d" % len(unscored))
    print("  estimated Trestle cost              $%.2f" % (len(unscored) * 0.015))
    gate(unreadable < max(5, len(records) // 10),
         "too many records unreadable (%d of %d); the API is throttling and the "
         "scoring set would be silently incomplete" % (unreadable, len(records)))
    if not args.commit:
        print("\nDRY RUN. Re-run with --commit to score and write tiers.")
        return
    if not os.getenv("TRESTLE_API_KEY") and not os.getenv("TRESTLE_PAID_API_KEY"):
        raise SystemExit("no TRESTLE_API_KEY / TRESTLE_PAID_API_KEY in the environment")

    key = os.getenv("TRESTLE_PAID_API_KEY") or os.getenv("TRESTLE_API_KEY")
    for i, num in enumerate(unscored, 1):
        try:
            data = pv.call_trestle(num, key)
        except Exception as e:  # noqa: BLE001
            cache[num] = {"error": str(e)[:120]}
            continue
        score = data.get("activity_score") if isinstance(data, dict) else None
        cache[num] = {"score": score, "tier": pv.assign_tier(score, tiers)}
        if i % 25 == 0:
            _save(cache_path, cache)
            log.info("scored %s/%s", i, len(unscored))
    _save(cache_path, cache)

    # ---- write the tier back onto each phone ----------------------------
    # upsert-phones upserts BY NUMBER and replaces the object, so the existing
    # type, tags, status and flags must be carried through or they are wiped.
    written = 0
    failed = 0
    by_tier = Counter()
    for num, holders in targets.items():
        info = cache.get(num) or {}
        tier = info.get("tier")
        if not tier or tier == "Unknown":
            continue
        by_tier[tier] += 1
        for uuid, p in holders:
            owner_uuid, phone_obj = crm.find_phone_object(uuid, num)
            if not owner_uuid or not phone_obj:
                failed += 1
                continue
            # Strip any OTHER tier before adding this one. A union would leave
            # a re-scored phone carrying two tiers at once, so it would match
            # "Dial First" and "Dial Third" presets simultaneously and get
            # texted from two lanes.
            tags = sorted((set(_tag_names(phone_obj)) - tier_names) | {tier})
            payload = {
                "number": num,
                "type": phone_obj.get("type") or "UNKNOWN",
                "tags": tags,
                "status": (phone_obj.get("status") or "UNKNOWN").upper(),
                "is_connected": phone_obj.get("is_connected", True),
                "verified": phone_obj.get("verified", False),
            }
            c = crm.client()
            try:
                c._request("/api/internal/owner/%s/upsert-phones/" % owner_uuid,
                           method="POST", body={"phones": [payload]})
                written += 1
            except Exception as e:  # noqa: BLE001
                log.warning("tier write failed for %s: %s", num, str(e)[:110])
                failed += 1

    mark_done("score", scored=len(cache), written=written, failed=failed,
              by_tier=dict(by_tier))
    print("score: %d numbers scored, %d tier tags written, %d failed"
          % (len(cache), written, failed))
    print("  by tier: " + str(dict(by_tier)))


# -------------------------------------------------------------- writeback --
# Put the verified numbers into the CRM, which is both what Ty asked for and the
# gate on everything downstream. A filter preset queries the CRM and
# seed.from_preset keeps a row only when its phone carries a dial-tier tag.
# Measured before this existed: 4 of 25 reachable buyers had any tier tag and
# 154 good numbers were SmartSkip-only and not in the CRM at all, so a preset
# would have returned almost nobody however well the numbers scored locally.


def phase_writeback(args) -> None:
    require("profiles")
    import time
    from sms_agent import crm
    from sms_agent.seed import SKIP_PHONE_STATUSES
    import phone_validator as pv

    profiles = _load(OUT / "buyer_profiles.json", [])
    reg = {b["buyer_key"]: b for b in _load(OUT / "registry.json", [])}
    ph = _load(OUT / "phones.json", {})
    scores = _load(OUT / "phone_scores.json", {})

    # NO tag-existence gate. Verified live: a phone tag AUTO-CREATES on write,
    # so posting "Drop" on a phone that never carried it simply works. Two dead
    # ends before that: dial_tier_uuids() learns names by joining sampled
    # records, so it can never report a tier no phone carries yet, and
    # /api/internal/tag/ is the PROPERTY tag namespace, which does not contain
    # the phone tiers at all. Gating on either one silently skipped every
    # Dial Third and Drop number.
    tier_names = set(pv.DEFAULT_TIERS) | {"Unknown"}
    targets = [p for p in profiles
               if p["phones"] and reg.get(p["buyer_key"], {}).get("saved_uuid")]
    if args.limit:
        targets = targets[:args.limit]

    stats = Counter()
    planned = []
    for p in targets:
        b = reg[p["buyer_key"]]
        rec = _get_record(crm, b["saved_uuid"])
        time.sleep(0.35)
        if not rec:
            stats["record unreadable"] += 1
            continue
        owner = rec.get("owner") or {}
        owner_uuid = owner.get("uuid") or owner.get("id")
        if not owner_uuid:
            stats["no owner uuid on the record"] += 1
            continue
        live = {}
        for q in (owner.get("phones") or []):
            if isinstance(q, dict) and q.get("number"):
                live[re.sub(r"\D", "", str(q["number"]))[-10:]] = q

        payload = []
        for q in p["phones"]:
            num = q["number"]
            tier = (scores.get(num) or {}).get("tier")
            if not tier or tier == "Unknown":
                stats["number never scored"] += 1
                continue
            cur = live.get(num)
            if cur and (cur.get("status") or "").upper() in SKIP_PHONE_STATUSES:
                # A disposition somebody set by hand outranks anything here.
                stats["left alone, already dispositioned"] += 1
                continue
            names = [(x.get("title") or x.get("name")) if isinstance(x, dict) else str(x)
                     for x in ((cur or {}).get("tags") or [])]
            names = [n for n in names if n]
            if tier in names:
                stats["already tagged correctly"] += 1
                continue
            # Strip any OTHER tier so a phone never carries two, then add this one.
            tags = sorted((set(names) - tier_names) | {tier})
            payload.append({
                "number": num,
                "type": (cur or {}).get("type") or q.get("type") or "UNKNOWN",
                "tags": tags,
                "status": ((cur or {}).get("status") or "UNKNOWN").upper(),
                "is_connected": (cur or {}).get("is_connected",
                                                q.get("is_connected", True)),
                "verified": (cur or {}).get("verified", False),
            })
            stats["number already on the record, tagging"] += 1 if cur else 0
            stats["NEW number to write"] += 0 if cur else 1
        if payload:
            planned.append((b["name"], owner_uuid, payload[:20]))

    total = sum(len(x[2]) for x in planned)
    print("writeback: %d records, %d phone rows to write" % (len(planned), total))
    for k, v in stats.most_common():
        print("  %-40s %d" % (k, v))
    if not args.commit:
        print("")
        print("DRY RUN. Re-run with --commit to write.")
        return

    c = crm.client()
    ok = fail = 0
    for i, (name, owner_uuid, payload) in enumerate(planned, 1):
        try:
            c._request("/api/internal/owner/%s/upsert-phones/" % owner_uuid,
                       method="POST", body={"phones": payload})
            ok += len(payload)
        except Exception as e:  # noqa: BLE001
            fail += len(payload)
            log.warning("writeback failed for %s: %s", name[:30], str(e)[:110])
        if i % 25 == 0:
            log.info("wrote %s/%s records", i, len(planned))
        time.sleep(0.3)

    gate(ok > 0, "writeback wrote nothing; check crm auth")
    mark_done("writeback", records=len(planned), rows_written=ok, rows_failed=fail)
    print("writeback: %d phone rows written, %d failed" % (ok, fail))


# ------------------------------------------------------------------- qa ----

def phase_qa(args) -> None:
    s = state()
    reg = _load(OUT / "registry.json", [])
    rep = _load(OUT / "merge_report.json", {"merges": [], "refused": []})
    print("phases run:", ", ".join(sorted(s)) or "none")
    if not reg:
        print("no registry yet")
        return
    ent = sum(1 for b in reg if b["is_entity"])
    with_mail = sum(1 for b in reg if b["mail"])
    both = sum(1 for b in reg if len(b["counties"]) > 1)
    tiers = Counter(b["tier"] for b in reg)
    print("buyers                 : " + str(len(reg)))
    print("  entities             : " + str(ent))
    print("  people               : " + str(len(reg) - ent))
    print("  with mailing address : " + str(with_mail))
    print("  buy in both counties : " + str(both))
    print("  tiers                : " + str(dict(tiers)))
    print("merges / refused       : " + str(len(rep["merges"])) + " / "
          + str(len(rep["refused"])))
    print("")
    print("top 15 by purchase count:")
    for b in reg[:15]:
        kind = "ENT" if b["is_entity"] else "per"
        print("  %3d  %-44s %s  %s" % (b["n_buys"], b["name"][:44], kind,
                                       ",".join(b["counties"])))


# ----------------------------------------------------------------- main ----

PHASES = {"sweep": phase_sweep, "recent": phase_recent,
          "hydrate": phase_hydrate,
          "aggregate": phase_aggregate, "unmask": phase_unmask,
          "principals": phase_principals, "skipinput": phase_skipinput,
          "phones": phase_phones, "crm-phones": phase_crm_phones,
          "profiles": phase_profiles, "writeback": phase_writeback,
          "score": phase_score,
          "qa": phase_qa}


def main() -> int:
    ap = argparse.ArgumentParser(description="Knox and Blount buyer registry")
    ap.add_argument("--phase", required=True, choices=sorted(PHASES))
    ap.add_argument("--commit", action="store_true",
                    help="allow billed calls / CRM writes (trace, mirror)")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--workers", type=int, default=4)
    ap.add_argument("--min-interval", type=float, default=0.3)
    ap.add_argument("--max-pages", type=int, default=200)
    ap.add_argument("--refresh", action="store_true",
                    help="re-run a segment already cached in rows.json")
    ap.add_argument("--preset", default="Dispo - 02 Ready to Text",
                    help="CRM filter preset naming the cohort to score")
    ap.add_argument("--active-only", action="store_true", default=True,
                    help="principals: only entities active in the window")
    ap.add_argument("--min-buys", type=int, default=2,
                    help="buy floor for the PAID principals pass")
    ap.add_argument("--trace-businessv2", action="store_true",
                    help="include BusinessV2 principals in SmartSkip (4%% hit rate)")
    ap.add_argument("--retrace", action="store_true",
                    help="re-buy numbers for buyers already traced")
    ap.add_argument("--include-low", action="store_true",
                    help="include low-confidence (agent-titled) principals")
    ap.add_argument("--trace-min-buys", type=int, default=2,
                    help="purchase floor for the PAID skip trace (0 = everyone)")
    ap.add_argument("--profile-min-buys", type=int, default=2,
                    help="purchase floor for the profiles phase")
    ap.add_argument("--recent-days", type=int, default=365,
                    help="the qualifying activity window in days")
    ap.add_argument("--require-active", dest="require_active",
                    action="store_true", default=True,
                    help="profiles: only buyers active in the window")
    ap.add_argument("--no-require-active", dest="require_active",
                    action="store_false",
                    help="profiles: ignore recency (the old behaviour)")
    ap.add_argument("--source", choices=("preset", "registry"), default="preset",
                    help="score phase: CRM preset cohort, or the registry's own phones")
    ap.add_argument("--include-dnc", action="store_true",
                    help="also score numbers flagged do-not-call")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
    OUT.mkdir(parents=True, exist_ok=True)
    PHASES[args.phase](args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
