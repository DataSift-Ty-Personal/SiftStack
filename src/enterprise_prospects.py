"""Enterprise outbound prospect pipeline: top property buyers nationwide.

Builds the DataSift sales team's enterprise outbound list from the nationwide
buyers dataset plus SiftMap recency data. Read-only against SiftMap: it NEVER
calls add-properties-by-query (which spends record allowance).

    python src/enterprise_prospects.py --phase audit
    python src/enterprise_prospects.py --phase classify
    python src/enterprise_prospects.py --phase probe
    python src/enterprise_prospects.py --phase sweep [--limit 3]
    python src/enterprise_prospects.py --phase rank
    python src/enterprise_prospects.py --phase verify [--resume]
    python src/enterprise_prospects.py --phase queue
    (research fan-out runs in-session, writes research/results_batch_NN.json)
    python src/enterprise_prospects.py --phase merge
    python src/enterprise_prospects.py --phase export
    python src/enterprise_prospects.py --phase qa

Contract notes carried in from the rest of the repo:
  * nationwide_buyers.csv has BuyerCity/BuyerState VALUES SWAPPED vs headers,
    and BuyerPurchases6MSum is a regional total repeated on every county row,
    so volume = max per name, never a sum across rows.
  * SiftMap /properties/search/ silently ignores unknown filter keys. Every
    filter is proven by count delta, never by HTTP 200.
  * a county address needs search=<bare county name> + type="county" on top of
    the rich object, and the viewport polygon is a list of {lon,lat} dicts.
  * zero data is a FAILURE. Every phase gates on row counts before saving.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import math
import os
import re
import sys
import threading
import time
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    import config  # noqa: F401  loads .env
    OUTPUT_ROOT = Path(getattr(config, "OUTPUT_DIR", "output"))
    PROJECT_ROOT = Path(getattr(config, "PROJECT_ROOT", "."))
except Exception:
    OUTPUT_ROOT = Path("output")
    PROJECT_ROOT = Path(".")

from siftmap_standalone import SiftMapClient, SiftMapError  # noqa: E402

log = logging.getLogger("enterprise_prospects")

OUT = OUTPUT_ROOT / "enterprise_prospects"
STATE_PATH = OUT / "state.json"

SEED_A = (PROJECT_ROOT / "Skills for REI" / "extracted" / "skills_expanded"
          / "buyer-prospector" / "buyer-prospector" / "data" / "nationwide_buyers.csv")
SEED_B = PROJECT_ROOT / "skills" / "buyer-prospector" / "data" / "nationwide_buyers.csv"

MAP_BASE = "https://map.reisift.io"

# Whole-world viewport (searches intersect the map viewport). Dicts, not pairs.
WORLD = [{"lon": -17.89461189115002, "lat": 72.08452694723852},
         {"lon": -17.89461189115002, "lat": -13.881763595427103},
         {"lon": -163.30518783395442, "lat": -13.881763595427103},
         {"lon": -163.30518783395442, "lat": 72.08452694723852}]

# ── taxonomy ─────────────────────────────────────────────────────────────
# Polarity is INVERTED from export_buyer_list.py: high volume is the signal
# here, so there is no volume auto-exclude. HIGH = brand-specific keyword,
# excluded outright. MED = generic word; near the cutline it goes to REVIEW
# for research instead of silent exclusion.

EXCLUDE_GROUPS: dict[str, dict[str, list[str]]] = {
    "GOVERNMENT": {
        "high": ["CITY OF", "COUNTY OF", "STATE OF", "TOWN OF", "VILLAGE OF",
                 "HOUSING AUTHORITY", "LAND BANK", "REDEVELOPMENT",
                 "HOUSING DEVELOPMENT", "DEPARTMENT OF", "UNITED STATES",
                 "SECRETARY OF HOUSING", "SECRETARY OF VETERANS",
                 "SCHOOL DISTRICT", "METROPOLITAN GOVERNMENT",
                 "PORT AUTHORITY", "PUBLIC HOUSING", "HUD "],
        "med": []},
    "GSE_BANK_SERVICER": {
        "high": ["FANNIE MAE", "FEDERAL NATIONAL MORTGAGE", "FREDDIE MAC",
                 "FEDERAL HOME LOAN", "PLANET HOME LENDING", "NEWREZ",
                 "VILLAGE CAPITAL", "PENNYMAC", "LAKEVIEW LOAN",
                 "CARRINGTON MORTGAGE", "MTGLQ", "WILMINGTON SAVINGS",
                 "WILMINGTON TRUST", "US BANK", "U S BANK", "WELLS FARGO",
                 "DEUTSCHE BANK", "BANK OF AMERICA", "JPMORGAN", "CITIBANK",
                 "HSBC", "PNC BANK", "TRUIST", "NATIONSTAR", "MR COOPER",
                 "FREEDOM MORTGAGE", "ROCKET MORTGAGE", "LOANDEPOT",
                 "GITSIT", "SELENE FINANCE", "SN SERVICING", "RUSHMORE",
                 "SHELLPOINT", "OCWEN", "PHH MORTGAGE", "SPECIALIZED LOAN",
                 "BANK OF NEW YORK", "GOLDMAN SACHS", "MORGAN STANLEY"],
        "med": ["BANK", "MORTGAGE", "LENDING", "LOAN", "SERVICING",
                "AS TRUSTEE", "CREDIT UNION", "SAVINGS"]},
    "HOMEBUILDER": {
        "high": ["LENNAR", "D R HORTON", "DR HORTON", "NVR ", "PULTE",
                 "MERITAGE", "KB HOME", "TOLL BROTHERS", "BEAZER",
                 "PERRY HOMES", "WEEKLEY", "TAYLOR MORRISON",
                 "CENTURY COMMUNITIES", "CENTURY COMPLETE", "LGI HOMES",
                 "CLAYTON PROPERTIES", "CMH HOMES", "STONE MARTIN",
                 "DREAM FINDERS", "US HOME", "HIGHLAND HOMES",
                 "ASHTON WOODS", "TRI POINTE", "M/I HOMES", "MI HOMES",
                 "SMITH DOUGLAS", "ADAMS HOMES", "HOLIDAY BUILDERS",
                 "MARONDA", "RAUSCH COLEMAN", "CENTEX", "RICHMOND AMERICAN",
                 "SHEA HOMES", "HOVNANIAN", "GEHAN HOMES", "HISTORY MAKER",
                 "BLOOMFIELD HOMES", "FIRST TEXAS HOMES", "ANTARES HOMES",
                 "IMPRESSION HOMES", "TROPHY SIGNATURE", "BRIGHTLAND",
                 "GREAT SOUTHERN HOMES", "MUNGO HOMES", "TRUE HOMES",
                 "EASTWOOD HOMES", "DAN RYAN", "RYAN HOMES",
                 "STANLEY MARTIN", "LANDSEA", "STARLIGHT HOMES",
                 "EXPRESS HOMES", "OPEN HOUSE TEXAS"],
        "med": ["BUILDERS", "HOMEBUILDER", "NEW HOMES", "CONSTRUCTION"]},
    "RELOCATION": {
        "high": ["RELOCATION", "NOMINEE", "NATIONAL RESIDENTIAL", "CARTUS",
                 "SIRVA", "GRAEBEL", "WEICHERT WORKFORCE"],
        "med": []},
    "TITLE_ESCROW": {
        "high": ["TITLE ", " TITLE", "ESCROW"],
        "med": []},
    "ESTATE_NONPROFIT": {
        "high": ["ESTATE OF", "CHURCH", "MINISTRIES", "HABITAT FOR HUMANITY",
                 "DIOCESE", "BAPTIST", "CATHOLIC", "SYNAGOGUE",
                 "COMMUNITY LAND TRUST", "UNIVERSITY"],
        "med": ["FOUNDATION", "TEMPLE", "MISSION", "SCHOOL"]},
    "LAND_DEV": {
        "high": ["COLONY RIDGE", "MINERALS"],
        "med": ["DEVELOPMENT", "DEVELOPERS", "LAND CO"]},
}

TIER2_KEYWORDS = ["OPENDOOR", "OFFERPAD", "ZILLOW", "HOMEVESTORS",
                  "UGLY HOUSES", "WE BUY", "SUNDAE", "MARKETPLACE HOMES",
                  "NEW WESTERN", "HOMEGO", "HOME BUYERS", "HOUSE BUYERS",
                  "HOUSEBUYERS", "CASH OFFER", "CASHOFFER"]

TIER3_KEYWORDS = ["INVITATION HOMES", "AMERICAN HOMES 4 RENT", "AMH ",
                  "PROGRESS RESIDENTIAL", "FIRSTKEY", "TRICON", "VINEBROOK",
                  "AMHERST", "MAIN STREET RENEWAL", "PRETIUM", "CERBERUS",
                  "BLACKSTONE", "STARWOOD", "ARMM ASSET", "STAR BORROWER"]
TIER3_PATTERNS = [re.compile(r"\bSFR\b"),
                  re.compile(r"BORROWER\b.*\b(LP|LLC)$"),
                  re.compile(r"\bOWNER\s+[IVXL0-9]+\s+(LLC|LP)$"),
                  re.compile(r"ACQUISITION TRUST"),
                  re.compile(r"\bJV[- ]?\d"),
                  re.compile(r"\bRCA?F\s*\d")]

ENTITY_RX = re.compile(
    r"\b(LLC|L L C|LLP|LP|INC|CORP|TRUST|PROPERTIES|HOLDINGS|HOMES|CAPITAL|"
    r"INVEST|VENTURES|GROUP|REALTY|SOLUTIONS|PARTNERS|ENTERPRISES|ASSET)\b")

# Broader organization test for the principal check: a reverse-address owner
# only counts as a human principal when it matches NONE of these.
ORG_RX = re.compile(
    r"\b(LLC|LLP|LP|INC|CORP|TRUST|PROPERTIES|HOLDINGS|HOMES|CAPITAL|INVEST|"
    r"VENTURES|GROUP|REALTY|SOLUTIONS|PARTNERS|ENTERPRISES|ASSET|BANK|COMPANY|"
    r"ASSOCIATION|FUND|MANAGEMENT|SERVICES|SVC|SVCS|POSTAL|CHURCH|MINISTRIES|"
    r"CITY|COUNTY|STATE|AUTHORITY|DISTRICT|FOUNDATION)\b")

UNIT_RX = re.compile(r"\b(STE|SUITE|UNIT|APT|BLDG|FL|FLOOR|PMB|RM|#)\b.*$",
                     re.IGNORECASE)


def strip_unit(addr: str) -> str:
    return UNIT_RX.sub("", addr or "").strip().rstrip(",")

INVESTOR_SIGNALS = ["PROPERTIES", "HOMES", "INVESTMENT", "HOLDINGS", "CAPITAL",
                    "REALTY", "GROUP", "VENTURES", "PROPERTY SOLUTIONS",
                    "ENTERPRISES", "ACQUISITIONS", "EQUITY", "ASSET"]

MIN_VOLUME = 10          # Tier 1 floor (max purchases per 6 months)
TOP_SELECT = 500         # over-select for research attrition (wave 1 measured
                         # 44% exclusion in the top 200, so 400 was too thin)
TOP_FINAL = 250
POOL_SIZE = 1000         # second-tier scored pool shipped unresearched

# DataSift brand (same constants as list_inventory_report.py)
NAVY, BLUE, GREEN, GOLD, RED = "0A1130", "316AFF", "1B9E5A", "B8860B", "C00000"


# ── shared helpers ───────────────────────────────────────────────────────

def norm_name(raw: str) -> str:
    """Canonical buyer-name join key used by every phase."""
    s = re.sub(r"[^A-Z0-9 ]", " ", (raw or "").upper())
    s = re.sub(r"\bL\s+L\s+C\b", "LLC", s)
    s = re.sub(r"\bL\s+L\s+P\b", "LLP", s)
    s = re.sub(r"\bL\s+P\b", "LP", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def load_json(p: Path):
    return json.loads(p.read_text(encoding="utf-8"))


def save_json(p: Path, obj) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=1, default=str), encoding="utf-8")


def load_state() -> dict:
    return load_json(STATE_PATH) if STATE_PATH.exists() else {}


def mark_done(phase: str, **stats) -> None:
    st = load_state()
    st[phase] = {"done": datetime.now().isoformat(timespec="seconds"), **stats}
    save_json(STATE_PATH, st)


def require(phase: str) -> dict:
    st = load_state()
    if phase not in st:
        raise SystemExit(f"upstream phase '{phase}' has not completed; run it first")
    return st[phase]


def gate(ok: bool, msg: str) -> None:
    if not ok:
        raise SystemExit(f"GATE FAILED: {msg}")
    log.info("gate ok: %s", msg)


# ── SiftMap search client (bulk sold query) ──────────────────────────────

class SearchClient:
    """POST /properties/search/ with Api-Key first, Bearer JWT fallback.

    Thread-safe throttle. Read-only: this class exposes search() only.
    """

    def __init__(self, min_interval: float = 0.25):
        self._key = os.getenv("REISIFT_API_KEY", "")
        self._mode = "apikey" if self._key else "bearer"
        self._api = None
        self._lock = threading.Lock()
        self._last = 0.0
        self._min_interval = min_interval
        self.calls = 0

    def _bearer(self):
        if self._api is None:
            from datasift_api_upload import Api
            self._api = Api()
            self._api._mint()
        return self._api

    def _headers(self) -> dict:
        h = {"accept": "application/json", "content-type": "application/json",
             "origin": "https://beta.reisift.io",
             "referer": "https://beta.reisift.io/"}
        if self._mode == "apikey":
            h["authorization"] = f"Api-Key {self._key}"
        else:
            h["Authorization"] = "Bearer " + self._bearer().token
        return h

    def _wait(self) -> None:
        with self._lock:
            gap = time.time() - self._last
            if gap < self._min_interval:
                time.sleep(self._min_interval - gap)
            self._last = time.time()
            self.calls += 1

    def search(self, body: dict, timeout: float = 240.0) -> dict:
        data = json.dumps(body).encode()
        for attempt in range(6):
            self._wait()
            req = urllib.request.Request(
                MAP_BASE + "/properties/search/", data=data, method="POST",
                headers=self._headers())
            try:
                with urllib.request.urlopen(req, timeout=timeout) as r:
                    return json.loads(r.read())
            except urllib.error.HTTPError as e:
                txt = ""
                try:
                    txt = e.read().decode("utf-8", "replace")[:300]
                except Exception:
                    pass
                if e.code == 401:
                    if self._mode == "apikey":
                        log.warning("Api-Key rejected on /properties/search/; "
                                    "falling back to Bearer JWT")
                        self._mode = "bearer"
                        continue
                    self._bearer()._mint()
                    if attempt < 5:
                        continue
                if e.code == 429 and attempt < 5:
                    wait = 5.0 * (attempt + 1)
                    ra = e.headers.get("Retry-After") if e.headers else None
                    if ra:
                        try:
                            wait = float(ra)
                        except ValueError:
                            pass
                    log.warning("429 on search, sleeping %.0fs", wait)
                    time.sleep(wait)
                    continue
                if e.code in (500, 502, 503, 504) and attempt < 5:
                    time.sleep(3.0 * (attempt + 1))
                    continue
                raise RuntimeError(f"HTTP {e.code} on /properties/search/: {txt}")
            except urllib.error.URLError as e:
                if attempt < 5:
                    time.sleep(3.0 * (attempt + 1))
                    continue
                raise RuntimeError(f"network error on search: {e}")
        raise RuntimeError("exhausted retries on /properties/search/")

    @property
    def mode(self) -> str:
        return self._mode


def county_address(fips: str, name: str, state: str) -> dict:
    title = f"{name} County, {state}"
    return {"state": state, "title": title, "value": title, "county": name,
            "searchType": "county",
            "counties": [{"fips": fips, "county_name": name}],
            "search": name, "type": "county"}


def search_body(fips: str, name: str, state: str, filters: dict,
                page: int = 1, page_size: int = 250) -> dict:
    # result_index is a ROW OFFSET, not a page number (verified live: offset
    # 251 has zero overlap with offset 1 on a 250-row page). Ordering is
    # stable, so pages are 1, 251, 501, ...
    return {"result_index": 1 + (page - 1) * page_size,
            "with_boundaries": False,
            "filters": filters,
            "addresses": [county_address(fips, name, state)],
            "polygon": WORLD}


def find_paths(obj, pred, path=()) -> list[tuple]:
    """All key paths in a nested dict/list whose LEAF KEY matches pred."""
    out = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if pred(str(k)) and not isinstance(v, (dict, list)):
                out.append(path + (k,))
            out.extend(find_paths(v, pred, path + (k,)))
    elif isinstance(obj, list) and obj:
        out.extend(find_paths(obj[0], pred, path + ("[]",)))
    return out


def get_path(obj, path):
    cur = obj
    for k in path:
        if k == "[]":
            if not isinstance(cur, list) or not cur:
                return None
            cur = cur[0]
        elif isinstance(cur, dict):
            cur = cur.get(k)
        else:
            return None
    return cur


# ── phase 0: audit ───────────────────────────────────────────────────────

def phase_audit(args) -> None:
    gate(SEED_A.exists() or SEED_B.exists(), "at least one seed CSV exists")
    texts = {}
    for tag, p in (("A", SEED_A), ("B", SEED_B)):
        if p.exists():
            texts[tag] = p.read_text(encoding="utf-8", errors="replace") \
                          .replace("\r\n", "\n").replace("\r", "\n")
    if len(texts) == 2 and texts["A"] != texts["B"]:
        la, lb = texts["A"].split("\n"), texts["B"].split("\n")
        diff = sum(1 for x, y in zip(la, lb) if x != y) + abs(len(la) - len(lb))
        gate(diff == 0, f"seed copies differ beyond line endings ({diff} lines); "
                        "reconcile before proceeding")
    text = texts.get("A") or texts.get("B")
    rows = list(csv.DictReader(text.splitlines()))
    gate(len(rows) > 80000, f"seed row count {len(rows)} looks truncated")

    # Prove the city/state swap empirically before applying it.
    two_letter_in_city_col = sum(
        1 for r in rows[:2000] if len((r["BuyerCity"] or "").strip()) == 2)
    swapped = two_letter_in_city_col / 2000 > 0.85
    log.info("city/state swap detected: %s (%.0f%% two-letter values in "
             "BuyerCity column)", swapped, two_letter_in_city_col / 20)

    buyers: dict[str, dict] = {}
    for r in rows:
        name = norm_name(r["BuyerFullName"])
        if not name:
            continue
        city = (r["BuyerState"] if swapped else r["BuyerCity"]).strip()
        state = (r["BuyerCity"] if swapped else r["BuyerState"]).strip().upper()
        try:
            purch = int(float(r["BuyerPurchases6MSum"] or 0))
        except ValueError:
            purch = 0
        b = buyers.setdefault(name, {
            "name": name, "raw_name": r["BuyerFullName"].strip(),
            "address": (r["BuyerAddress"] or "").strip(),
            "city": city, "state": state, "zip": (r["BuyerZIP"] or "").strip(),
            "max_purchases": 0, "counties": []})
        b["max_purchases"] = max(b["max_purchases"], purch)
        b["counties"].append({"fips": r["FIPS"].strip(),
                              "county": r["County Name"].strip(),
                              "state": r["County State"].strip(),
                              "purchases": purch})
        if not b["address"] and r["BuyerAddress"]:
            b["address"] = r["BuyerAddress"].strip()

    for b in buyers.values():
        b["n_counties"] = len({c["fips"] for c in b["counties"]})

    gate(len(buyers) > 50000, f"unique buyers {len(buyers)} below expectation")
    n10 = sum(1 for b in buyers.values() if b["max_purchases"] >= 10)
    gate(n10 > 3000, f"buyers at 10+ purchases {n10} below expectation")

    save_json(OUT / "seed_normalized.json", {
        "swapped_city_state": swapped, "rows_in": len(rows),
        "buyers": list(buyers.values())})
    mark_done("audit", rows_in=len(rows), unique_buyers=len(buyers),
              buyers_10plus=n10, swap_applied=swapped)
    print(f"audit: {len(rows)} rows -> {len(buyers)} unique buyers, "
          f"{n10} at 10+ purchases, swap={swapped}")


# ── phase 1: classify ────────────────────────────────────────────────────

def classify_one(name: str, volume: int) -> tuple[str, str, str]:
    """Returns (tier, reason, confidence). tier in EXCLUDE/REVIEW/1/2/3/below."""
    for kw in TIER2_KEYWORDS:
        if kw in name:
            return "2", f"tier2 keyword: {kw}", "high"
    for kw in TIER3_KEYWORDS:
        if kw in name:
            return "3", f"tier3 keyword: {kw}", "high"
    for rx in TIER3_PATTERNS:
        if rx.search(name):
            return "3", f"tier3 pattern: {rx.pattern}", "med"
    for group, kws in EXCLUDE_GROUPS.items():
        for kw in kws["high"]:
            if kw in name:
                return "EXCLUDE", f"{group}: {kw}", "high"
    for group, kws in EXCLUDE_GROUPS.items():
        for kw in kws["med"]:
            if kw in name:
                # generic word: exclude confidently only at low volume,
                # otherwise send to research review
                if volume < MIN_VOLUME:
                    return "EXCLUDE", f"{group} (generic): {kw}", "med"
                return "REVIEW", f"{group} generic keyword at volume: {kw}", "med"
    if volume < MIN_VOLUME:
        return "below", "under volume floor", "high"
    return "1", "core ICP by volume", "high"


def phase_classify(args) -> None:
    require("audit")
    seed = load_json(OUT / "seed_normalized.json")
    buyers = seed["buyers"]
    counts: Counter = Counter()
    excl_log = []
    for b in buyers:
        tier, reason, conf = classify_one(b["name"], b["max_purchases"])
        b["tier"], b["tier_reason"], b["tier_conf"] = tier, reason, conf
        counts[tier] += 1
        if tier == "EXCLUDE":
            excl_log.append({"name": b["raw_name"], "volume": b["max_purchases"],
                             "n_counties": b["n_counties"], "reason": reason,
                             "confidence": conf, "stage": "rules"})
    gate(sum(counts.values()) == len(buyers), "classification covered all buyers")
    gate(counts["1"] > 1000, f"tier 1 count {counts['1']} suspiciously low")
    gate(counts["EXCLUDE"] > 100, "exclusions suspiciously low; keyword lists dead?")

    save_json(OUT / "classified.json", {"buyers": buyers, "counts": dict(counts)})
    save_json(OUT / "exclusions_log.json",
              sorted(excl_log, key=lambda x: -x["volume"]))
    mark_done("classify", **{f"tier_{k}": v for k, v in counts.items()})
    print("classify:", dict(counts))


# ── phase 2: probe ───────────────────────────────────────────────────────

def phase_probe(args) -> None:
    sc = SearchClient(min_interval=args.min_interval)
    knox = ("47093", "Knox", "TN")
    report: dict = {"when": datetime.now().isoformat(timespec="seconds")}

    base_filters = {"value_min": 1}
    r = sc.search(search_body(*knox, base_filters))
    baseline = r.get("total_results", 0)
    report["auth_mode"] = sc.mode
    report["baseline_count"] = baseline
    gate(baseline > 10000, f"Knox baseline {baseline}; county query broken")

    date_min = (datetime.now() - timedelta(days=183)).strftime("%Y-%m-%d")
    dated = sc.search(search_body(*knox, dict(base_filters,
                                              extra_last_sale_date_min=date_min)))
    report["dated_count"] = dated.get("total_results", 0)
    gate(0 < report["dated_count"] < baseline,
         f"extra_last_sale_date_min did not filter "
         f"({report['dated_count']} vs baseline {baseline})")

    # extra_last_sale_price_min is silently IGNORED (verified live: a $10M
    # floor left the count unchanged). Record the fact; do not use the key.
    priced = sc.search(search_body(*knox, dict(
        base_filters, extra_last_sale_date_min=date_min,
        extra_last_sale_price_min=10_000_000)))
    report["price_filter_ignored"] = (
        priced.get("total_results") == dated.get("total_results"))

    rows = dated.get("data") or []
    report["page_size"] = len(rows)
    report["total_pages"] = dated.get("total_pages")
    report["row_keys"] = sorted(rows[0].keys()) if rows else []
    gate(bool(rows), "search returned zero data rows on a nonzero count")

    # Does the row carry the buyer (current owner) without hydration?
    owner_paths = find_paths(rows[0], lambda k: "owner" in k.lower())
    report["owner_paths"] = [list(p) for p in owner_paths]
    name_path = next((p for p in owner_paths
                      if "name" in p[-1].lower()
                      and isinstance(get_path(rows[0], p), str)
                      and get_path(rows[0], p).strip()), None)
    report["owner_name_path"] = list(name_path) if name_path else None
    # No owner name on the row, but corporateOwned + id are there, so the
    # sweep hydrates only corporate-owned sold rows (the target population).
    report["branch"] = "A" if name_path else "corp_hydrate"
    report["corp_owned_page1"] = sum(
        1 for x in rows if x.get("corporateOwned"))
    report["sample_row"] = rows[0]

    # pagination: offset semantics, stable ordering
    page2 = sc.search(search_body(*knox, dict(
        base_filters, extra_last_sale_date_min=date_min), page=2))
    rows2 = page2.get("data") or []
    ids1 = {x.get("id") for x in rows}
    ids2 = {x.get("id") for x in rows2}
    report["pagination_ok"] = bool(rows2) and not (ids1 & ids2)
    gate(report["pagination_ok"], "offset pagination returned overlapping rows")

    # duplicate-county-name disambiguation: two different Franklin counties
    fr_tn = sc.search(search_body("47067", "Franklin", "TN", base_filters))
    fr_oh = sc.search(search_body("39049", "Franklin", "OH", base_filters))
    report["dup_county_check"] = {
        "franklin_tn_47067": fr_tn.get("total_results"),
        "franklin_oh_39049": fr_oh.get("total_results")}
    gate(fr_tn.get("total_results") != fr_oh.get("total_results"),
         "same-named counties returned identical totals; fips not honored")

    save_json(OUT / "probe_report.json", report)
    mark_done("probe", branch=report["branch"], auth_mode=report["auth_mode"],
              page_size=report["page_size"], baseline=baseline,
              dated=report["dated_count"],
              corp_frac_page1=report["corp_owned_page1"] / max(len(rows), 1),
              pagination_ok=report["pagination_ok"])
    print(f"probe: auth={report['auth_mode']} branch={report['branch']} "
          f"page_size={report['page_size']} counts {baseline} -> "
          f"{report['dated_count']} corp_page1={report['corp_owned_page1']} "
          f"pagination_ok={report['pagination_ok']}")


# ── phase 3: sweep ───────────────────────────────────────────────────────

def pick_counties(buyers: list[dict], top_n: int) -> list[tuple[str, str, str]]:
    """Top counties by count of rule-surviving 10+ candidates present."""
    tally: Counter = Counter()
    meta: dict[str, tuple[str, str]] = {}
    for b in buyers:
        if b["tier"] in ("1", "2", "3", "REVIEW") and b["max_purchases"] >= MIN_VOLUME:
            for c in b["counties"]:
                tally[c["fips"]] += 1
                meta[c["fips"]] = (c["county"], c["state"])
    return [(f, meta[f][0], meta[f][1]) for f, _ in tally.most_common(top_n)]


def phase_sweep(args) -> None:
    """County recency sweep: paginate every property SOLD in the last 6
    months (offset pagination, 250 rows/page), keep the corporate-owned rows
    (search rows carry no owner name, but corporateOwned + the dataflik id are
    on the row and entity buyers are the target population), then hydrate a
    capped random sample per county via get_detail to read the deed buyer.
    """
    require("probe")
    require("classify")
    classified = load_json(OUT / "classified.json")["buyers"]
    sweep_dir = OUT / "sweep"
    sweep_dir.mkdir(parents=True, exist_ok=True)

    counties = pick_counties(classified, args.counties)
    if args.limit:
        counties = counties[:args.limit]
    date_min = (datetime.now() - timedelta(days=183)).strftime("%Y-%m-%d")
    filters = {"value_min": 1, "extra_last_sale_date_min": date_min}

    sc = SearchClient(min_interval=args.min_interval)
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import random
    rng = random.Random(39049)
    detail_calls = {"n": 0}

    def one_county(cty):
        fips, name, state = cty
        out_p = sweep_dir / f"county_{fips}.json"
        if args.resume and out_p.exists():
            return fips, "cached", len(load_json(out_p).get("buyers", {}))
        corp_rows, page, total = [], 1, 0
        while page <= 200:
            r = sc.search(search_body(fips, name, state, filters, page=page))
            total = r.get("total_results", 0)
            rows = r.get("data") or []
            if not rows:
                break
            corp_rows.extend(
                {"id": x.get("id"), "address": x.get("address")}
                for x in rows if x.get("corporateOwned") and x.get("id"))
            if page * 250 >= total:
                break
            page += 1

        sample = corp_rows
        if len(sample) > args.hydrate_cap:
            sample = rng.sample(corp_rows, args.hydrate_cap)
        smc = SiftMapClient(min_interval=0.15)
        agg: dict[str, dict] = {}
        hydrated = 0
        for row in sample:
            try:
                d = smc.get_detail(row["id"])
            except SiftMapError as e:
                if "401" in str(e):
                    raise
                continue
            hydrated += 1
            detail_calls["n"] += 1
            sh = (d.get("sale_history") or [{}])[0]
            oi = d.get("owner_info") or {}
            buyer = norm_name(str(sh.get("buyer_name")
                                  or oi.get("owner_name") or ""))
            if not buyer:
                continue
            a = agg.setdefault(buyer, {
                "n": 0, "cash_n": 0,
                "raw_name": sh.get("buyer_name") or oi.get("owner_name"),
                "mail": oi.get("owner_mail_address"),
                "portfolio_n": oi.get("total_properties"),
                "portfolio_value": oi.get("portfolio_value")})
            a["n"] += 1
            a["cash_n"] += 1 if sh.get("is_cash_sale") else 0
        save_json(out_p, {
            "fips": fips, "county": name, "state": state,
            "sold_total": total, "corp_rows": len(corp_rows),
            "hydrated": hydrated, "sampled": len(sample), "buyers": agg})
        return fips, "ok", len(agg)

    done = zeroes = 0
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(one_county, c): c for c in counties}
        for f in as_completed(futs):
            fips, status, n = f.result()
            done += 1
            zeroes += (n == 0)
            log.info("sweep %s: %s buyers (%s) [%d/%d]",
                     fips, n, status, done, len(counties))

    gate(zeroes <= max(1, len(counties) // 10),
         f"{zeroes} of {len(counties)} counties returned zero buyers")

    fresh: Counter = Counter()
    meta: dict[str, dict] = {}
    for p in sweep_dir.glob("county_*.json"):
        for owner, a in load_json(p).get("buyers", {}).items():
            fresh[owner] += a["n"]
            meta.setdefault(owner, a)
    save_json(OUT / "sweep_summary.json", {
        "branch": "corp_hydrate", "counties_swept": len(counties),
        "zero_counties": zeroes, "search_calls": sc.calls,
        "detail_calls": detail_calls["n"],
        "fresh_buyers": dict(fresh.most_common()),
        "buyer_meta": meta})
    mark_done("sweep", branch="corp_hydrate", counties=len(counties),
              fresh_buyers=len(fresh), search_calls=sc.calls,
              detail_calls=detail_calls["n"])
    print(f"sweep: {len(counties)} counties, {len(fresh)} fresh corporate "
          f"buyers, {sc.calls} search + {detail_calls['n']} detail calls")


# ── phase 4: rank ────────────────────────────────────────────────────────

def phase_rank(args) -> None:
    require("classify")
    classified = load_json(OUT / "classified.json")["buyers"]
    fresh = {}
    if (OUT / "sweep_summary.json").exists():
        fresh = load_json(OUT / "sweep_summary.json").get("fresh_buyers", {})

    cands = [b for b in classified
             if b["tier"] in ("1", "2", "3", "REVIEW")
             and b["max_purchases"] >= MIN_VOLUME]
    for b in cands:
        f = fresh.get(b["name"], 0)
        b["fresh_buys"] = f
        b["score"] = round(3.0 * math.log10(1 + b["max_purchases"])
                           + 1.5 * math.log10(1 + b["n_counties"])
                           + 2.0 * math.log10(1 + f), 4)
    cands.sort(key=lambda b: -b["score"])
    top = cands[:TOP_SELECT]
    pool = cands[TOP_SELECT:TOP_SELECT + POOL_SIZE]
    gate(len(top) >= TOP_FINAL, f"only {len(top)} candidates survived ranking")

    save_json(OUT / "candidates_top400.json", top)
    save_json(OUT / "pool_second_tier.json", pool)
    mark_done("rank", candidates=len(cands), selected=len(top), pool=len(pool))
    print(f"rank: {len(cands)} candidates -> top {len(top)} + pool {len(pool)}")


# ── phase 5: verify ──────────────────────────────────────────────────────

def _match_candidate(hits: list[dict], addr: str) -> dict | None:
    want = (addr or "").split()
    if not want:
        return None
    for h in hits:
        title = str(h.get("title") or "")
        if title.split() and title.split()[0] == want[0]:
            return h
    return None


def verify_target(smc: SiftMapClient, b: dict) -> dict:
    out = {"name": b["name"], "resolved": False, "principal": None,
           "principal_source": None, "portfolio_n": None,
           "portfolio_value": None, "recent_sales": None, "miss": None}
    street = strip_unit(b["address"])
    try:
        hits = smc.autocomplete(f"{street}, {b['city']}, {b['state']}")
        hit = _match_candidate(hits, street)
        if not hit and street != b["address"]:
            hits = smc.autocomplete(f"{b['address']}, {b['city']}, {b['state']}")
            hit = _match_candidate(hits, b["address"])
        if not hit:
            out["miss"] = "no autocomplete match on mailing address"
            return out
        d = smc.get_detail(hit["id"])
        oi = d.get("owner_info") or {}
        out["resolved"] = True
        out["siftmap_owner"] = oi.get("owner_name")
        out["portfolio_n"] = oi.get("total_properties")
        out["portfolio_value"] = oi.get("portfolio_value")
        sh = d.get("sale_history") or []
        out["recent_sales"] = [{"date": s.get("sale_date"),
                                "price": s.get("sale_price"),
                                "buyer": s.get("buyer_name")} for s in sh[:3]]
        # the Harper move: reverse the mailing address; a human owner at the
        # entity's mailing address is the principal
        owner = str(oi.get("owner_name") or "")
        if ENTITY_RX.search(b["name"]) and owner and not ORG_RX.search(
                norm_name(owner)):
            out["principal"] = owner
            out["principal_source"] = "siftmap-reverse-address"
        secondary = oi.get("secondary_owner_names") or []
        humans = [s for s in secondary if not ORG_RX.search(norm_name(str(s)))]
        if humans and not out["principal"]:
            out["principal"] = humans[0]
            out["principal_source"] = "siftmap-secondary-owner"
    except SiftMapError as e:
        if "401" in str(e):
            raise
        out["miss"] = str(e)[:160]
    return out


def phase_verify(args) -> None:
    require("rank")
    top = load_json(OUT / "candidates_top400.json")
    vdir = OUT / "verified"
    vdir.mkdir(parents=True, exist_ok=True)
    merged_p = OUT / "verified_merged.json"
    done: dict[str, dict] = {}
    if args.resume and merged_p.exists():
        done = {v["name"]: v for v in load_json(merged_p)}

    todo = [b for b in top if b["name"] not in done]
    if args.limit:
        todo = todo[:args.limit]
    from concurrent.futures import ThreadPoolExecutor, as_completed
    results = list(done.values())
    lock = threading.Lock()

    def work(b):
        smc = SiftMapClient(min_interval=max(args.min_interval, 0.15))
        return verify_target(smc, b)

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(work, b): b for b in todo}
        for i, f in enumerate(as_completed(futs), 1):
            r = f.result()
            with lock:
                results.append(r)
                if i % 25 == 0:
                    save_json(merged_p, results)
                    log.info("verify checkpoint %d/%d", i, len(todo))

    resolved = sum(1 for r in results if r["resolved"])
    total = len(results)
    # AUTH-or-COVERAGE guard (buyer_sweep.py precedent): zero resolutions of a
    # nonzero target list is an auth failure or a systemic break, never data.
    gate(resolved > 0, f"verified 0 of {total} targets: AUTH or COVERAGE "
                       "failure, not an empty market")
    gate(resolved + sum(1 for r in results if r["miss"]) == total,
         "every target must be resolved or carry an explicit miss reason")
    save_json(merged_p, results)
    mark_done("verify", targets=total, resolved=resolved,
              principals=sum(1 for r in results if r["principal"]))
    print(f"verify: {resolved}/{total} resolved, "
          f"{sum(1 for r in results if r['principal'])} principals unmasked")


# ── phase 6: queue ───────────────────────────────────────────────────────

def phase_queue(args) -> None:
    require("rank")
    top = load_json(OUT / "candidates_top400.json")
    verified = {}
    if (OUT / "verified_merged.json").exists():
        verified = {v["name"]: v for v in load_json(OUT / "verified_merged.json")}
    rdir = OUT / "research"
    rdir.mkdir(parents=True, exist_ok=True)

    batch, n_batches = [], 0
    for i, b in enumerate(top, 1):
        v = verified.get(b["name"], {})
        batch.append({
            "rank": i, "buyer_name": b["raw_name"], "norm_name": b["name"],
            "address": b["address"], "city": b["city"], "state": b["state"],
            "zip": b["zip"], "max_purchases": b["max_purchases"],
            "n_counties": b["n_counties"], "fresh_buys": b.get("fresh_buys", 0),
            "counties": [f"{c['county']}, {c['state']}"
                         for c in b["counties"][:8]],
            "rule_tier": b["tier"], "rule_reason": b["tier_reason"],
            "siftmap": {k: v.get(k) for k in
                        ("resolved", "siftmap_owner", "portfolio_n",
                         "portfolio_value", "principal", "principal_source")}})
        if len(batch) == 10:
            n_batches += 1
            save_json(rdir / f"queue_batch_{n_batches:02d}.json", batch)
            batch = []
    if batch:
        n_batches += 1
        save_json(rdir / f"queue_batch_{n_batches:02d}.json", batch)
    gate(n_batches >= 25, f"only {n_batches} research batches emitted")
    mark_done("queue", batches=n_batches, targets=len(top))
    print(f"queue: {n_batches} batches of 10 written to {rdir}")


# ── phase 7: merge ───────────────────────────────────────────────────────

RESEARCH_REQUIRED = ["buyer_name", "tier", "entity_type", "direct_to_seller",
                     "confidence"]


def phase_merge(args) -> None:
    require("verify")
    top = {b["name"]: b for b in load_json(OUT / "candidates_top400.json")}
    rdir = OUT / "research"
    results_files = sorted(rdir.glob("results_*.json"))
    gate(bool(results_files), "no research results files found")

    researched: dict[str, dict] = {}
    bad = []
    for rf in results_files:
        for rec in load_json(rf):
            missing_keys = [k for k in RESEARCH_REQUIRED if k not in rec]
            if missing_keys:
                bad.append({"file": rf.name, "name": rec.get("buyer_name"),
                            "missing": missing_keys})
                continue
            researched[norm_name(rec["buyer_name"])] = rec
    gate(not bad, f"{len(bad)} malformed research records; see stdout")
    coverage = sum(1 for n in top if n in researched) / max(len(top), 1)
    gate(coverage >= 0.95,
         f"research covers only {coverage:.0%} of the top {len(top)}")
    # fold the verify pass into each candidate before ranking
    if (OUT / "verified_merged.json").exists():
        vmap = {v["name"]: v for v in load_json(OUT / "verified_merged.json")}
        for n, b in top.items():
            v = vmap.get(n)
            if v:
                b["siftmap"] = {k: v.get(k) for k in
                                ("resolved", "siftmap_owner", "portfolio_n",
                                 "portfolio_value", "principal",
                                 "principal_source")}

    reclass = []
    final = []
    for name, b in top.items():
        r = researched.get(name)
        if not r:
            continue
        tier = str(r["tier"])
        if tier != b["tier"]:
            reclass.append({"name": b["raw_name"], "rule_tier": b["tier"],
                            "research_tier": tier,
                            "reason": r.get("tier_reason", "")})
        if tier == "EXCLUDE":
            continue
        final.append({**b, "research": r, "final_tier": tier})
    final.sort(key=lambda x: -x["score"])
    gate(len(final) >= TOP_FINAL,
         f"only {len(final)} researched non-excluded targets; need {TOP_FINAL}")
    final = final[:TOP_FINAL]
    save_json(OUT / "final_250.json", final)
    save_json(OUT / "reclassifications.json", reclass)
    mark_done("merge", researched=len(researched), final=len(final),
              reclassified=len(reclass))
    print(f"merge: {len(researched)} researched -> final {len(final)}, "
          f"{len(reclass)} reclassified")


# ── phase 8: export ──────────────────────────────────────────────────────

def phase_export(args) -> None:
    require("merge")
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    final = load_json(OUT / "final_250.json")
    pool = load_json(OUT / "pool_second_tier.json") \
        if (OUT / "pool_second_tier.json").exists() else []
    excl = load_json(OUT / "exclusions_log.json")
    st = load_state()

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor=NAVY)
    head_font = Font(color="FFFFFF", bold=True, size=10)
    tier_fill = {"1": PatternFill("solid", fgColor=GREEN),
                 "2": PatternFill("solid", fgColor=BLUE),
                 "3": PatternFill("solid", fgColor=GOLD)}

    def sheet(title, headers):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.fill, c.font = head_fill, head_font
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        return ws

    def fmt_row(b):
        r = b.get("research", {})
        g = r.get("gbp") or {}
        d2s = r.get("direct_to_seller") or {}
        pr = r.get("principals") or []
        p0 = pr[0] if pr else {}
        sift = dict(b.get("siftmap") or {})
        # a reverse-address owner that looks like an organization is noise
        if sift.get("principal") and ORG_RX.search(
                norm_name(str(sift["principal"]))):
            sift["principal"] = None
        return [
            b.get("_rank"), b["raw_name"], b["final_tier"],
            r.get("entity_type", ""),
            (d2s.get("verdict") or "unclear").title(),
            b["max_purchases"], b["n_counties"], b.get("fresh_buys", 0),
            ", ".join((r.get("footprint") or {}).get("states", [])[:6]),
            f"{b['city']}, {b['state']}",
            sift.get("portfolio_n") or b.get("portfolio_n"),
            sift.get("portfolio_value") or b.get("portfolio_value"),
            p0.get("name") or sift.get("principal") or "",
            p0.get("confidence", ""),
            r.get("website", ""),
            g.get("rating"), g.get("reviews_n"), g.get("phone", ""),
            f"{b['address']}, {b['city']}, {b['state']} {b['zip']}".strip(", "),
            r.get("confidence", ""),
            "; ".join((r.get("sources") or [])[:4]),
            "", "",  # Phone / Email: Enformion pass pending
            r.get("notes", "")]

    MASTER_HEADERS = [
        "Rank", "Company", "Tier", "Entity Type", "Direct to Seller",
        "6Mo Purchases", "County Count", "Fresh Buys (SiftMap)", "States",
        "HQ City, State", "Portfolio Props", "Portfolio Value",
        "Principal", "Principal Conf", "Website", "GBP Rating",
        "GBP Reviews", "GBP Phone", "Mailing Address", "Research Conf",
        "Sources", "Phone (Enformion pass pending)",
        "Email (Enformion pass pending)", "Notes"]

    # 1 Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    tiers = Counter(b["final_tier"] for b in final)
    ws.append(["DataSift Enterprise Prospect List"])
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws.append([f"Built {datetime.now():%Y-%m-%d}"])
    ws.append([])
    ws.append(["Tier", "Count", "Description"])
    for c in ws[4]:
        c.fill, c.font = head_fill, head_font
    ws.append(["Tier 1 Core ICP", tiers.get("1", 0),
               "Local and regional investors, 10-99 purchases per 6 months"])
    ws.append(["Tier 2 Institutional D2S", tiers.get("2", 0),
               "Institutions running direct-to-seller acquisition"])
    ws.append(["Tier 3 Wholesale Channel", tiers.get("3", 0),
               "Institutional buyers acquiring through wholesalers and MLS"])
    ws.append([])
    ws.append(["Pipeline", ""])
    for k in ("audit", "classify", "sweep", "verify", "merge"):
        if k in st:
            ws.append([k, json.dumps({x: y for x, y in st[k].items()
                                      if x != "done"})])
    ws.append([])
    ws.append(["Top 10"])
    ws.append(["Rank", "Company", "Tier", "6Mo Purchases", "Counties"])
    for c in ws[ws.max_row]:
        c.fill, c.font = head_fill, head_font
    for i, b in enumerate(final[:10], 1):
        ws.append([i, b["raw_name"], b["final_tier"], b["max_purchases"],
                   b["n_counties"]])

    # 2 Master + 3-5 tier views
    for i, b in enumerate(final, 1):
        b["_rank"] = i
    master = sheet("Top 250 Master", MASTER_HEADERS)
    for b in final:
        master.append(fmt_row(b))
        tf = tier_fill.get(b["final_tier"])
        if tf:
            master.cell(row=master.max_row, column=3).fill = tf
            master.cell(row=master.max_row, column=3).font = Font(
                color="FFFFFF", bold=True)
    for tier, title, extra in (("1", "Tier 1 Core ICP", None),
                               ("2", "Tier 2 Institutional D2S", None),
                               ("3", "Tier 3 Wholesale Channel",
                                "Dispo Angle")):
        hs = MASTER_HEADERS + ([extra] if extra else [])
        wt = sheet(title, hs)
        for b in [x for x in final if x["final_tier"] == tier]:
            row = fmt_row(b)
            if extra:
                row.append("Sell-to-them dispo relationship; enterprise pitch "
                           "only if research shows a direct-to-seller arm")
            wt.append(row)

    # 6 Second Tier Pool
    wp = sheet("Second Tier Pool", [
        "Company", "Rule Tier", "6Mo Purchases", "County Count",
        "Fresh Buys", "Score", "HQ City, State", "Mailing Address"])
    for b in pool:
        wp.append([b["raw_name"], b["tier"], b["max_purchases"],
                   b["n_counties"], b.get("fresh_buys", 0), b["score"],
                   f"{b['city']}, {b['state']}",
                   f"{b['address']}, {b['city']}, {b['state']} {b['zip']}"])

    # 7 Knox Depth Layer
    wk = sheet("Knox Depth Layer", [
        "Zip", "Buyer", "Principal", "Principal Source", "Buys", "Cash Buys",
        "Avg Price", "Portfolio Props", "Portfolio Value", "Mailing"])
    knox_rows = 0
    for zp in ("37803", "37914", "37917"):
        for p in OUTPUT_ROOT.glob(f"buyer_sweep_{zp}*.json"):
            for b in load_json(p).get("ranked", []):
                wk.append([zp, b.get("buyer"), b.get("principal"),
                           b.get("principal_source"), b.get("n_buys"),
                           b.get("cash_n"), b.get("avg_price"),
                           b.get("portfolio_n"), b.get("portfolio_value"),
                           b.get("owner_mail")])
                knox_rows += 1
            break

    # 8 Exclusions Log
    we = sheet("Exclusions Log", ["Company", "6Mo Purchases", "County Count",
                                  "Reason", "Confidence", "Stage"])
    groups = Counter(e["reason"].split(":")[0] for e in excl)
    for e in excl[:500]:
        we.append([e["name"], e["volume"], e["n_counties"], e["reason"],
                   e["confidence"], e["stage"]])
    we.append([])
    we.append(["Totals by group"])
    for g, n in groups.most_common():
        we.append([g, n])

    # 9 Methodology
    wm = sheet("Methodology and Data Notes", ["Note"])
    probe = load_json(OUT / "probe_report.json") \
        if (OUT / "probe_report.json").exists() else {}
    sweep = load_json(OUT / "sweep_summary.json") \
        if (OUT / "sweep_summary.json").exists() else {}
    for note in [
        "Seed: nationwide buyers dataset, 84,048 buyer-by-county rows across "
        "2,329 counties. Vintage is not stamped in the file; SiftMap recency "
        "weighting compensates and fresh-buy counts are labeled separately.",
        "The seed's BuyerCity/BuyerState values are swapped relative to their "
        "headers; the pipeline detects and corrects this empirically.",
        "Purchase volume is the MAX per buyer name across county rows, never "
        "a sum: the source repeats a regional total on every county row and "
        "summing would double count.",
        f"SiftMap auth mode used: {probe.get('auth_mode', 'n/a')}. Sold-date "
        "filters proven by count delta (baseline "
        f"{probe.get('baseline_count')} -> dated {probe.get('dated_count')} "
        f"-> priced {probe.get('priced_count')} in Knox County TN).",
        f"Recency layer: {sweep.get('counties_swept', 'n/a')} counties swept "
        "for properties sold in the last 6 months, corporate-owned rows "
        "hydrated to deed level for the buyer of record "
        f"({sweep.get('detail_calls', 'n/a')} deed reads). Individual-name "
        "buyers are covered by the seed dataset; the sweep intentionally "
        "hydrates entity purchases only.",
        "Google Business profile fields are derived from web search results, "
        "not the Places API; treat rating and review counts as indicative.",
        "Principals come from SiftMap reverse-mailing-address resolution "
        "where available. Phone and email columns are intentionally blank: "
        "the Enformion contact resolution pass is deferred.",
        "Tier 3 wholesale-channel institutions are kept on the list because "
        "research can show a direct-to-seller acquisition arm; the default "
        "relationship is dispo (sell to them), not an enterprise CRM pitch.",
        "Confidence letters follow the buyer-prospector research guide: "
        "H = SOS active plus verified officer, M = SOS inactive or deed-only, "
        "L = web-only or ambiguous name.",
    ]:
        wm.append([note])
        wm.cell(row=wm.max_row, column=1).alignment = Alignment(wrap_text=True)
    wm.column_dimensions["A"].width = 110

    # column widths
    for ws in wb.worksheets:
        if ws.title == "Methodology and Data Notes":
            continue
        for col in range(1, ws.max_column + 1):
            width = max((len(str(ws.cell(row=r, column=col).value or ""))
                         for r in range(1, min(ws.max_row, 60) + 1)),
                        default=8)
            ws.column_dimensions[get_column_letter(col)].width = min(
                max(width + 2, 8), 52)

    # dash audit: no em or en dashes anywhere
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    assert "—" not in c.value and "–" not in c.value, \
                        f"dash found in {ws.title}!{c.coordinate}"

    out_p = OUT / f"DataSift_Enterprise_Prospects_{datetime.now():%Y%m%d}.xlsx"
    wb.save(out_p)

    # read-back gate
    rb = load_workbook(out_p, read_only=True)
    n_master = rb["Top 250 Master"].max_row - 1
    gate(n_master == len(final),
         f"master sheet rows {n_master} != final {len(final)}")
    mark_done("export", workbook=str(out_p), master_rows=n_master,
              knox_rows=knox_rows, pool_rows=len(pool))
    print(f"export: {out_p} ({n_master} master rows, {knox_rows} Knox rows)")


# ── phase 9: qa ──────────────────────────────────────────────────────────

def phase_qa(args) -> None:
    st = load_state()
    gates = []

    def check(name, ok, detail=""):
        gates.append({"gate": name, "ok": bool(ok), "detail": str(detail)})

    check("audit done", "audit" in st, st.get("audit", {}))
    check("classify done", "classify" in st, st.get("classify", {}))
    check("probe done", "probe" in st, st.get("probe", {}))
    check("sweep done", "sweep" in st, st.get("sweep", {}))
    check("verify done", "verify" in st, st.get("verify", {}))
    if "verify" in st:
        v = st["verify"]
        check("verify resolution rate >= 50%",
              v.get("resolved", 0) >= v.get("targets", 1) * 0.5,
              f"{v.get('resolved')}/{v.get('targets')}")
    check("merge done", "merge" in st, st.get("merge", {}))
    if "merge" in st:
        check("final == 250", st["merge"].get("final") == TOP_FINAL,
              st["merge"].get("final"))
    check("export done", "export" in st, st.get("export", {}))
    if "export" in st:
        check("workbook exists", Path(st["export"]["workbook"]).exists(),
              st["export"]["workbook"])

    save_json(OUT / "qa_gates.json", gates)
    width = max(len(g["gate"]) for g in gates)
    bad = 0
    for g in gates:
        mark = "PASS" if g["ok"] else "FAIL"
        bad += (not g["ok"])
        print(f"  {g['gate']:<{width}}  {mark}  {g['detail']}")
    if bad:
        raise SystemExit(f"{bad} QA gate(s) failed")
    print("qa: all gates pass")


# ── main ─────────────────────────────────────────────────────────────────

PHASES = {"audit": phase_audit, "classify": phase_classify,
          "probe": phase_probe, "sweep": phase_sweep, "rank": phase_rank,
          "verify": phase_verify, "queue": phase_queue, "merge": phase_merge,
          "export": phase_export, "qa": phase_qa}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--phase", required=True, choices=sorted(PHASES))
    ap.add_argument("--resume", action="store_true")
    ap.add_argument("--limit", type=int, default=0,
                    help="cap counties (sweep) or targets (verify) for smoke tests")
    ap.add_argument("--workers", type=int, default=4)
    ap.add_argument("--min-interval", type=float, default=0.25,
                    help="seconds between SiftMap search calls per client")
    ap.add_argument("--counties", type=int, default=100,
                    help="how many top counties the sweep covers")
    ap.add_argument("--hydrate-cap", type=int, default=400,
                    help="max corporate-owned sold rows hydrated per county")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
    OUT.mkdir(parents=True, exist_ok=True)
    PHASES[args.phase](args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
