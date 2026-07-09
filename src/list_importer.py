"""Import emailed county distress lists into NoticeData records.

Three sources, all manually dropped into a folder after arriving by email:

  - `sarpy_nod`     — Sarpy County "Notice of Default" report (digital text PDF).
                      Owner = Grantor. No street address (legal desc only).
  - `douglas_export`— Douglas County Register-of-Deeds export (spreadsheet, .xlsx/.csv).
                      ONE 28-column template; branch on RPTTYPE:
                        NDEF → foreclosure (has ADDRESS + PARCEL)
                        DSL  → tax_sale   (owner name + cert # only, no address)
                      Owner = SNDPARTY (NOT FRSTPARTY — that's the trustee / "STATE").

See docs/tickets/TICKET-03-emailed-list-importer.md for the full mapping.

Output is a list[NoticeData] ready for the standard enrichment pipeline. Rows without a
street address (Sarpy, Douglas DSL) carry their legal description / owner name in raw_text
so a downstream name/parcel-based address lookup can resolve them.
"""

import csv
import logging
import re
from datetime import datetime
from pathlib import Path

import config
from notice_parser import NoticeData

logger = logging.getLogger(__name__)

# RPTTYPE / CHARACTER → notice_type for the Douglas export template.
_RPTTYPE_TO_NOTICE_TYPE = {
    "NDEF": "foreclosure",
    "DSL": "tax_sale",
}

# Party-name role suffixes to strip from the tail of a name (trustee markers etc.).
_ROLE_SUFFIX_RE = re.compile(r"\s+(?:TR|TRUSTEE|TRUSTEES|TTEE|ET AL|ETAL)\.?$", re.IGNORECASE)

# A "Grantee" value that means "no real counterparty" — ignore it.
_GENERIC_GRANTEE = "WHOM IT MAY CONCERN"


# ── Shared helpers ─────────────────────────────────────────────────────


def _clean(value) -> str:
    """Coerce a cell to a trimmed string ('' for None)."""
    if value is None:
        return ""
    return str(value).strip()


def _normalize_name(raw: str, surname_first: bool = False) -> str:
    """Normalize a county-record party name to 'First [Middle] Last'.

    Handles the two conventions seen in the emailed lists:
      - Douglas: 'LAST,FIRST MID' (explicit comma)          → 'First Mid Last'
      - Sarpy:   'LAST FIRST MID' (surname first, no comma)  → 'First Mid Last'
    Business entities (LLC/INC/etc.) are left as-is (just cleaned + title-cased lightly).
    Trailing role suffixes (TR, TRUSTEE, ET AL) are stripped.
    """
    name = _clean(raw)
    if not name:
        return ""

    name = _ROLE_SUFFIX_RE.sub("", name).strip()

    # Entities: keep word order, don't reorder as a person name.
    if config.BUSINESS_RE.search(name):
        return _titlecase(name)

    if "," in name:
        last, _, rest = name.partition(",")
        name = f"{rest.strip()} {last.strip()}".strip()
    elif surname_first:
        tokens = name.split()
        if len(tokens) > 1:
            name = " ".join(tokens[1:] + tokens[:1])

    return _titlecase(name)


def _titlecase(name: str) -> str:
    """Title-case a name while preserving all-caps entity acronyms (LLC, INC).

    Capitalizes each hyphen-separated part so 'BERRY-FISHER' → 'Berry-Fisher'.
    """
    out = []
    for word in name.split():
        if config.BUSINESS_RE.fullmatch(word) or (word.isupper() and len(word) <= 3):
            out.append(word)  # keep LLC / INC / LP / SR-like acronyms as-is
        else:
            out.append("-".join(part.capitalize() for part in word.split("-")))
    return " ".join(out)


def _join_owners(primary: str, secondary: str, surname_first: bool = False) -> str:
    """Combine a primary + optional co-owner into one owner_name string."""
    p = _normalize_name(primary, surname_first=surname_first)
    s = _normalize_name(secondary, surname_first=surname_first)
    if p and s:
        return f"{p} & {s}"
    return p or s


def _parse_date(raw: str) -> str:
    """Parse a county date/time string to ISO 'YYYY-MM-DD' ('' if unparseable)."""
    raw = _clean(raw)
    if not raw:
        return ""
    # Try a handful of formats seen across the exports/report.
    for fmt in (
        "%m/%d/%Y %I:%M:%S %p",  # 04/21/2026 09:46:21 AM  (Sarpy PDF)
        "%m/%d/%Y %H:%M:%S",     # 06/15/2026 14:57:47     (Douglas sheet)
        "%m/%d/%Y",              # 04/21/2026
        "%m/%d/%y",              # 04/20/26                (report header)
    ):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Last resort: pull a bare M/D/Y out of the string.
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", raw)
    if m:
        mo, day, yr = m.groups()
        yr = ("20" + yr) if len(yr) == 2 else yr
        try:
            return datetime(int(yr), int(mo), int(day)).strftime("%Y-%m-%d")
        except ValueError:
            pass
    logger.debug("Unparseable date: %r", raw)
    return ""


def _split_address(raw: str) -> tuple[str, str]:
    """Split 'STREET, CITY' → (street, city). No comma → (whole, '')."""
    raw = _clean(raw)
    if not raw:
        return ("", "")
    if "," in raw:
        street, _, city = raw.rpartition(",")
        return (street.strip(), city.strip().title())
    return (raw, "")


# ── Spreadsheet reader (Douglas export) ────────────────────────────────


def _read_table_rows(path: Path) -> list[dict]:
    """Read an .xlsx or .csv spreadsheet into a list of row dicts.

    Keys are uppercased, stripped header names for robust lookup regardless of
    minor header-casing/spacing changes between exports.
    """
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover - dependency guard
            raise RuntimeError(
                "openpyxl is required to read .xlsx lists — add it to requirements.txt"
            ) from e
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        keys = [str(h).strip().upper() if h is not None else "" for h in header]
        rows = [dict(zip(keys, r)) for r in rows_iter]
        wb.close()
        return rows

    # CSV (Google Sheets export, or a saved .csv)
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return []
        keys = [h.strip().upper() for h in header]
        return [dict(zip(keys, r)) for r in reader]


def _parse_douglas_export(path: Path) -> list[NoticeData]:
    """Parse the Douglas Register-of-Deeds export (.xlsx/.csv), branching on RPTTYPE."""
    rows = _read_table_rows(path)
    if not rows:
        logger.warning("No rows in %s", path.name)
        return []

    notices: list[NoticeData] = []
    skipped = 0
    for row in rows:
        rpttype = _clean(row.get("RPTTYPE")).upper()
        notice_type = _RPTTYPE_TO_NOTICE_TYPE.get(rpttype)
        if not notice_type:
            skipped += 1
            continue

        owner = _join_owners(row.get("SNDPARTY", ""), row.get("OTHRSND", ""))
        if not owner:
            skipped += 1
            continue

        street, city = _split_address(row.get("ADDRESS", ""))
        instr = _clean(row.get("INSTR NUMB")) or _clean(row.get("REFNUMBER"))

        # Retain everything useful for a downstream name/parcel address lookup.
        legal_bits = [
            f"legal={_clean(row.get('ADDNNAME'))} "
            f"Lot:{_clean(row.get('LOT'))} Block:{_clean(row.get('BLOCK'))}".strip(),
            f"trustee/first_party={_clean(row.get('FRSTPARTY'))}",
            f"ref={_clean(row.get('REFNUMBER'))}",
        ]
        if notice_type == "tax_sale":
            legal_bits.append(f"cert={_clean(row.get('UCCNUMB'))}")
        raw_text = " | ".join(b for b in legal_bits if b and not b.endswith("="))

        notices.append(
            NoticeData(
                address=street,
                city=city or "Omaha",
                state="NE",
                zip=_clean(row.get("ZIP")),
                owner_name=owner,
                notice_type=notice_type,
                county="Douglas",
                parcel_id=_clean(row.get("PARCEL")),
                date_added=_parse_date(row.get("FILEDATE TIME", "")),
                source_url=f"email-list://{path.name}#instr={instr}",
                raw_text=raw_text,
            )
        )

    logger.info(
        "Douglas export %s: %d records (%d skipped) — %s",
        path.name, len(notices), skipped,
        ", ".join(sorted({n.notice_type for n in notices})) or "none",
    )
    return notices


# ── PDF reader (Sarpy Notice of Default report) ────────────────────────

# One data row: NOTICE OF DEFAULT <filed date> <instr#> <grantor> WHOM IT MAY CONCERN <legal>
_SARPY_ROW_RE = re.compile(
    r"NOTICE OF DEFAULT\s+"
    r"(\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}\s*[AP]M)\s+"  # 1: filed date
    r"(\d{6,})\s+"                                               # 2: instrument #
    r"(.+?)\s+" + re.escape(_GENERIC_GRANTEE) + r"\s+"          # 3: grantor
    r"(.+?)"                                                     # 4: legal
    r"(?=\s+NOTICE OF DEFAULT\b|\s+Total Document Count\b|$)",
    re.IGNORECASE,
)


def _extract_pdf_text(path: Path) -> str:
    """Extract text from a digital (non-scanned) PDF via pypdfium2 — no OCR."""
    import pypdfium2 as pdfium

    doc = pdfium.PdfDocument(str(path))
    try:
        chunks = []
        for i in range(len(doc)):
            page = doc[i]
            textpage = page.get_textpage()
            chunks.append(textpage.get_text_range())
            textpage.close()
        return "\n".join(chunks)
    finally:
        doc.close()


def _parse_sarpy_nod(path: Path) -> list[NoticeData]:
    """Parse the Sarpy County 'Notice of Default' PDF report into NoticeData."""
    text = _extract_pdf_text(path)
    flat = re.sub(r"\s+", " ", text).strip()

    notices: list[NoticeData] = []
    for m in _SARPY_ROW_RE.finditer(flat):
        filed, instr, grantor, legal = m.groups()
        owner = _normalize_name(grantor, surname_first=True)
        if not owner:
            continue
        notices.append(
            NoticeData(
                address="",  # no street address — resolved downstream from legal/owner
                city="",
                state="NE",
                owner_name=owner,
                notice_type="foreclosure",
                county="Sarpy",
                date_added=_parse_date(filed),
                source_url=f"email-list://{path.name}#instr={instr.strip()}",
                raw_text=f"legal={legal.strip()} | instrument={instr.strip()}",
            )
        )

    if not notices:
        logger.warning(
            "Sarpy NOD %s: 0 rows matched — check the report layout / anchor phrase",
            path.name,
        )
    else:
        logger.info("Sarpy NOD %s: %d records", path.name, len(notices))
    return notices


# ── Dispatch ───────────────────────────────────────────────────────────

_PARSERS = {
    "douglas_export": _parse_douglas_export,
    "sarpy_nod": _parse_sarpy_nod,
}


def detect_source(path: Path) -> str | None:
    """Best-effort auto-detection of a dropped file's source profile."""
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm", ".csv"):
        try:
            rows = _read_table_rows(path)
        except Exception:  # noqa: BLE001 - detection is best-effort
            return None
        if rows and "RPTTYPE" in rows[0]:
            return "douglas_export"
        return None
    if ext == ".pdf":
        try:
            head = _extract_pdf_text(path)[:2000].upper()
        except Exception:  # noqa: BLE001
            return None
        if "NOTICE OF DEFAULT" in head:
            return "sarpy_nod"  # extend when other PDF reports appear
    return None


def import_list(path: str | Path, source: str | None = None) -> list[NoticeData]:
    """Import one emailed list file into NoticeData records.

    Args:
        path: Path to the dropped file (.xlsx/.csv/.pdf).
        source: One of _PARSERS keys, or None to auto-detect from content.

    Returns:
        List of NoticeData (un-deduplicated, un-enriched).
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"List file not found: {path}")

    if source is None:
        source = detect_source(path)
        if source is None:
            raise ValueError(
                f"Could not auto-detect source for {path.name}; "
                f"pass --source ({', '.join(_PARSERS)})"
            )
        logger.info("Auto-detected source for %s: %s", path.name, source)

    parser = _PARSERS.get(source)
    if parser is None:
        raise ValueError(f"Unknown source {source!r}; expected one of {list(_PARSERS)}")

    return parser(path)
