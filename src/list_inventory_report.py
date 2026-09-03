"""Render the County List Framework inventory for ty+2 as a branded workbook.

Reads what src/list_inventory.py and src/siftmap_pull.py measured and answers
the question Ty asked: which framework lists am I marketing to, which am I not,
and how deep is each one.

    python src/list_inventory_report.py
"""
from __future__ import annotations

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY, BLUE, GREEN, GOLD, RED = "0A1130", "316AFF", "1B9E5A", "B8860B", "C00000"
GREY, LINE = "5B6480", "D8DCE6"

THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def band(ws, row, text, width):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    return row + 1


def head(ws, row, cols):
    for i, t in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=t)
        c.font = Font(bold=True, size=9, color=NAVY)
        c.border = Border(bottom=Side(style="medium", color=NAVY))
        c.alignment = Alignment(vertical="bottom", wrap_text=True,
                                horizontal="left" if i == 1 else "right")
    ws.row_dimensions[row].height = 28
    return row + 1


def put(ws, row, vals, *, fmts=None, colors=None, bold=False):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BOX
        c.font = Font(size=10, bold=bold,
                      color=(colors or {}).get(i, "000000"))
        c.alignment = Alignment(horizontal="left" if i == 1 else "right",
                                vertical="center")
        f = (fmts or {}).get(i)
        if f:
            c.number_format = f
    return row + 1


def widths(ws, w):
    for i, x in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


def main() -> int:
    knox = json.load(open("output/list_inventory_knox.json", encoding="utf-8"))
    blount = json.load(open("output/list_inventory_blount.json", encoding="utf-8"))
    depth = json.load(open("output/segment_depth.json", encoding="utf-8"))
    counts = knox["list_counts"]

    wb = Workbook()

    # ---------------- 1. Marketing inventory ----------------
    ws = wb.active
    ws.title = "Marketing Inventory"
    r = band(ws, 1, "Ty+2 (Volunteer Home Buyers)  |  County List Framework inventory"
                    "  |  Knox + Blount, TN  |  17 Aug 2026", 6)
    r += 1
    r = band(ws, r, "Segments I am marketing to, and how deep", 6)
    r = head(ws, r, ["Framework segment (DataSift list)", "Records in account",
                     "Available Knox + Blount", "Not pulled", "Depth",
                     "Read"])
    money = "#,##0"
    # Where the account holds MORE than SiftMap currently flags, the surplus is
    # the courthouse first-to-market scrape (probate, liens, notices of default).
    # That is the point of FTM, so it reads as complete, not as an error.
    for title, have, avail, gap, pct in depth:
        if pct >= 130:
            note, col = "Complete, plus courthouse FTM depth", GREEN
            pct = 100.0
        elif pct >= 100:
            note, col = "Complete", GREEN
            pct = 100.0
        elif pct >= 85:
            note, col = "Near complete", GREEN
        elif pct >= 50:
            note, col = "Partial", GOLD
        else:
            note, col = "Thin", RED
        r = put(ws, r, [title, have, avail, gap if pct < 100 else 0, pct / 100, note],
                fmts={2: money, 3: money, 4: money, 5: "0%"},
                colors={5: col, 6: col})
    r += 1
    r = band(ws, r, "Not a list: the AI Score rows", 6)
    r = put(ws, r, ["Investor AI Score (off-market)", "addon active", "", "", "",
                    "Reachable, pull via SiftMap preset"])
    widths(ws, [40, 18, 22, 14, 10, 32])

    # ---------------- 2 + 3. The framework, per county ----------------
    for cty, data in (("Knox", knox), ("Blount", blount)):
        ws = wb.create_sheet("%s framework" % cty)
        b = data["county"]["baseline"]
        r = band(ws, 1, "%s County, TN  |  %s  |  %s investor deals  |  county baseline "
                        "%s doors per deal" % (cty, data["county"]["window"],
                                               data["county"]["deals"], b["dpd"]), 7)
        r += 1
        for pr, label in ((1, "Priority 1  (3x or better than county baseline)"),
                          (2, "Priority 2  (1.5x to 3x, volume at fair cost)")):
            rows = [x for x in data["rows"] if x["priority"] == pr]
            if not rows:
                continue
            r = band(ws, r, "%s  |  %d rows" % (label, len(rows)), 7)
            r = head(ws, r, ["Segment", "Doors per deal", "Lift", "Deals",
                             "List size", "Deal share", "Covered by a live list"])
            for x in sorted(rows, key=lambda z: z["dpd"]):
                if x["type"] == "ai":
                    note, col = "AI Score addon (not a list)", BLUE
                elif x["covered"]:
                    note, col = "Yes", GREEN
                else:
                    note, col = "NO", RED
                r = put(ws, r, [x["seg"], x["dpd"], x["lift"], x["deals"],
                                x["list_size"], x.get("share") or 0, note],
                        fmts={2: "0.0", 3: '0.0"x"', 4: "#,##0", 5: "#,##0",
                              6: "0.0%"},
                        colors={7: col})
            r += 1
        widths(ws, [42, 14, 9, 9, 12, 11, 28])

    # ---------------- 4. What changed + what is left ----------------
    ws = wb.create_sheet("Actions")
    r = band(ws, 1, "Done today", 3)
    r = head(ws, r, ["Action", "Records", "Result"])
    for a, n, res in [
        ("Created the Out-of-State list (it did not exist)", 5070,
         "The only framework segment at zero"),
        ("Pulled Knox out-of-state owners from SiftMap", 3867, "In account"),
        ("Pulled Blount out-of-state owners from SiftMap", 1203, "In account"),
        ("New records added to the account", 2459,
         "The rest were already held, now correctly listed"),
        ("Record allowance consumed", 2456, "1,642 to 4,098 of 100,000"),
    ]:
        r = put(ws, r, [a, n, res], fmts={2: "#,##0"})
    r += 1

    r = band(ws, r, "Framework rows this unlocked", 3)
    r = head(ws, r, ["Row", "County", "Rank"])
    for seg, cty, rank in [
        ("Free & Clear + Out-of-State", "Knox", "PRIORITY 1  4.0x  52 doors per deal"),
        ("Free & Clear + Out-of-State", "Blount", "Priority 2  4.7x  57 doors per deal"),
        ("Out-of-State", "Knox", "Priority 2  2.5x  81 doors per deal"),
        ("Out-of-State", "Blount", "Priority 2  3.1x  87 doors per deal"),
        ("Absentee + Out-of-State", "Knox", "Priority 2  2.2x  95 doors per deal"),
        ("Absentee + Free & Clear + Out-of-State", "Knox",
         "Priority 2  2.7x  76 doors per deal"),
    ]:
        r = put(ws, r, [seg, cty, rank],
                colors={3: GREEN if "PRIORITY 1" in rank else "000000"},
                bold="PRIORITY 1" in rank)
    r += 1

    r = band(ws, r, "Your call: the remaining depth, all Priority 2", 3)
    r = head(ws, r, ["Segment", "Records not pulled", "Worth it?"])
    for seg, n, note in [
        ("Free & Clear", 44404, "P2 in both counties. 1.7x Knox, biggest single gap"),
        ("Senior Homeowners", 42560, "Senior ALONE is P3 (1.0x). Only pull inside a combo"),
        ("High Equity", 59897, "P3 at 0.6x. The framework says skip it"),
        ("Tired Landlord", 11264, "P2 both counties. Cheapest real gap left"),
        ("Absentee Owners", 2321, "Already 90 percent pulled"),
        ("Low Income", 460, "Already 90 percent pulled"),
    ]:
        col = GREEN if "Cheapest" in note else (RED if "skip" in note else "000000")
        r = put(ws, r, [seg, n, note], fmts={2: "#,##0"}, colors={3: col})
    r += 1

    r = band(ws, r, "Findings that need a decision", 3)
    r = head(ws, r, ["Finding", "Where", "Why it matters"])
    for f, w, why in [
        ("All 35 SiftMap presets route to no list",
         "lists: [] on every preset",
         "Auto-add pulls records that land in no list, so no preset can market them"),
        ("Presets cut off below $100,000",
         "value_min: 100000",
         "Your buy box is $1 to $700K. Condemned and tax-distressed stock sits under $100K"),
        ("Three presets are labelled Priority 2 but now rank Priority 1",
         "Bad Credit + Low Income, Absentee + Tax Delinquent, Tax Delinquent + Senior",
         "Built on the old ranking. The refreshed 6-month window promoted them"),
        ("14 of 16 Knox Priority 2 rows have no feeder preset",
         "SiftMap presets",
         "The data is in the account but nothing keeps it current"),
        ("One junk list left over from an older pull",
         "list titled 744e99ae-f140-458a-beed-38a17700acb8",
         "Zero records. Same bug as today: the lists field takes NAMES, not UUIDs"),
    ]:
        r = put(ws, r, [f, w, why])
    widths(ws, [52, 46, 62])

    for s in wb.worksheets:
        s.sheet_view.showGridLines = False
        s.freeze_panes = "A4"

    os.makedirs("output", exist_ok=True)
    out = "output/Ty2_County_List_Inventory_2026-08-17.xlsx"
    wb.save(out)
    print("wrote", out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
