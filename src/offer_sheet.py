#!/usr/bin/env python3
"""One-page wholesale offer sheet, laid out like the Fortune Builders analyzer.

This is the default deliverable after a walkthrough (Ty, 2026-08-27). It answers
exactly one question, OFFER TO SELLER, and it answers it on ONE tab.

    "I'm really just looking for a very simple offer price estimation, and let's
     not do a bunch of different options. As a default, let's just look to
     wholesale this and build out an offer price structure in that way. Similar
     to how the Fortune Builders rehab estimator breaks down profit and ROI."

Shape is lifted from `output/Copy of The Repair Estimator.xlsx`, tab "Deal
Analyzer for Flips": paired left and right blocks under banded headers, a
percent column beside every dollar column, inputs INLINE on the same page, and
a net profit and ROI snapshot. The three mortgage tranches are dropped because
they are a lender-package concern, not an offer concern.

Every result is a live Excel formula off a workbook defined name, so changing a
blue cell moves the offer, the buyer's profit and every ROI together.

Usage:
    python src/offer_sheet.py --pack output/1342_grainger_pack.json \
                              --walk walk_1342_grainger.json \
                              --out "1342_Grainger_Ave_Offer.xlsx"
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from lender_package import (BOX, CALC_FILL, GREEN, GREY, INPUT_FILL, MONEY, MULT,
                            NAVY, NUM, PCT0, PCT1, RED, TOTAL_FILL, _band, _kv,
                            _para, _polish, _row, _title)
import post_walkthrough as pw
from post_walkthrough import MarketListing, build_rehab_matrix, rng

logger = logging.getLogger(__name__)

# Defaults are the operator's, not the model's. Every one is a blue cell.
DEFAULT_RULE = 0.70          # the classic maximum-allowable-offer haircut
DEFAULT_FEE = 15_000         # Ty runs a flat $10-15K assignment fee
DEFAULT_HOLD = 6             # months, for the buy-it-ourselves block only
DEFAULT_INSURANCE = 225      # vacant dwelling, monthly
DEFAULT_UTILITIES = 200      # monthly while held
DEFAULT_ESCROW_BUY = 900
DEFAULT_TITLE_PCT = 0.0077
DEFAULT_ESCROW_SELL = 900
DEFAULT_REALTOR_PCT = 0.054
DEFAULT_TRANSFER_PCT = 0.0037   # TN transfer tax plus recording
MAX_COMPS = 5
MAX_GATES = 4


def _num(v, default=0.0) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _ba(v) -> str:
    """1.0 renders as 1, 1.5 stays 1.5."""
    f = _num(v)
    return str(int(f)) if f == int(f) else str(f)


def _revive(rows: list) -> list:
    out = []
    for c in rows or []:
        c = dict(c)
        c.pop("raw", None)
        try:
            out.append(MarketListing(**c))
        except TypeError:
            continue
    return out


def _hydrate(pack: dict, walk: dict, args) -> None:
    """Give the reused post_walkthrough builders exactly what they read.

    They want live MarketListing objects on pack["sold"], a recomputed rehab,
    and an exits["inputs"] band. A saved pack carries the comps SERIALIZED as
    pack["sold_comps"], so without this the Comps tab renders empty and looks
    like a thin market rather than a wiring bug.
    """
    pack["sold"] = _revive(pack.get("sold_comps"))
    pack["active"] = _revive(pack.get("active_comps"))
    pack["walk"] = walk
    pack["rehab"] = build_rehab_matrix(pack["subject"], walk)   # never trust the saved one
    pack.setdefault("buckets", {})
    pack.setdefault("deeds", {})
    pack.setdefault("buyers", [])
    pack.setdefault("max_buyers", 25)
    pack.setdefault("arv_comp_ids", list((pack.get("arv") or {}).get("comp_ids") or []))

    target = walk.get("target_config") or {}
    finished = dict(pack["subject"])
    finished["beds"] = int(target.get("beds") or finished.get("beds") or 3)
    finished["baths"] = float(target.get("baths") or finished.get("baths") or 1.0)
    pack["finished"] = finished

    # sheet_buyers ranks buyers against THIS deal's band. The band is now the
    # offer band, which is a truer target than the old exit engine's spread.
    arv = _num(args.arv) or _num(((pack.get("arv") or {}).get("base") or {}).get("arv"))
    rehab_total = _num(next(iter((pack["rehab"].get("totals") or {"reno": 0}).values())))
    as_is = _num(args.as_is) or _num(pack.get("as_is"))
    offer = round(_num(args.rule) * arv - rehab_total - _num(args.fee))
    assign = offer + _num(args.fee)
    pack["exits"] = {"inputs": {
        "shell": bool(walk.get("unfinanceable")),
        "contract": rng(offer, confident=True), "contract_point": offer,
        "assign": rng(assign, confident=True), "assign_point": assign,
        "as_is": rng(as_is, confident=True), "as_is_point": as_is,
    }}


def sheet_offer(wb: Workbook, pack: dict, walk: dict, args) -> None:
    subject = pack["subject"]
    lead = pack.get("lead") or {}
    arv_block = pack.get("arv") or {}
    base = arv_block.get("base") or {}
    tight = arv_block.get("tight") or {}

    # THE TRAP: pack["rehab"]["totals"] can be stale (it is serialized at the
    # original run and survives later walk edits). Always recompute.
    rehab = build_rehab_matrix(subject, walk)
    rehab_total = _num((rehab.get("totals") or {}).get(
        next(iter(rehab.get("totals") or {"reno": 0}))))
    scen = (rehab.get("scenarios") or [{}])[0]
    scen_label = scen.get("label") or "Comp-Match Reno"

    # An explicit CLI figure WINS over the pack. A saved pack is a snapshot and
    # goes stale the moment the walk or the ARV is revised, which has already
    # bitten this deal twice (the rehab totals and the ARV both).
    arv = _num(args.arv) or _num(base.get("arv"))
    as_is = _num(args.as_is) or _num(pack.get("as_is"))
    sqft = int(_num(subject.get("sqft"), 0))

    # Annual property tax: Zillow publishes a rate, the county Values tab 500s.
    # This is a seed for a blue cell, never presented as a fact.
    tax_rate = _num(args.tax_rate, 0.0039)
    tax_annual = round(_num(lead.get("estimated_value")) * tax_rate) or 1100

    ws = wb.create_sheet("Offer")
    names: dict[str, str] = {}
    wide_rows: list[int] = []

    def name(nm: str, col: int, row: int) -> None:
        names[nm] = f"'Offer'!${get_column_letter(col)}${row}"

    addr = subject.get("full_address") or subject.get("address") or ""
    _title(ws, "Wholesale offer analysis", addr,
           "Blue cells are yours to change. Grey cells calculate off them. "
           "Change one blue cell and the offer, the buyer's profit and every "
           "return move with it.")

    # ── property header ────────────────────────────────────────────────
    r = 5
    r = _band(ws, r, "Property", col=1, span=7)
    _kv(ws, r, "Address", addr, span=2)
    _row(ws, r, 5, "Beds / baths",
         val=f"{subject.get('beds','')} / {_ba(subject.get('baths'))}", val_fmt=None)
    r += 1
    _kv(ws, r, "Square footage", sqft, fmt=NUM)
    _row(ws, r, 5, "Year built", val=subject.get("year_built", ""), val_fmt=NUM)
    r += 1
    _kv(ws, r, "County / parcel",
        f"{subject.get('county','')} / {str(subject.get('parcel_id','')).upper()}")
    _row(ws, r, 5, "Prepared", val=args.prepared or date.today().isoformat(),
         val_fmt=None)
    r += 2

    # ── values and pricing | holding costs ─────────────────────────────
    top = r
    r = _band(ws, r, "Property values and pricing", col=1, span=3)
    r = _row(ws, r, 1, "After repair value (ARV)", val=arv, fill=INPUT_FILL)
    name("ARV", 3, r - 1)
    r = _row(ws, r, 1, "Current as-is value", val=as_is, fill=INPUT_FILL)
    name("AsIs", 3, r - 1)
    r = _row(ws, r, 1, f"Estimated repair costs: {scen_label}",
             val=rehab_total, fill=INPUT_FILL)
    name("Rehab", 3, r - 1)
    r = _row(ws, r, 1, "Rule of thumb (percent of ARV)", val=_num(args.rule),
             val_fmt=PCT0, fill=INPUT_FILL)
    name("RulePct", 3, r - 1)
    r = _row(ws, r, 1, "Hold time, months (if we keep it)", val=args.hold,
             val_fmt=NUM, fill=INPUT_FILL)
    name("HoldMonths", 3, r - 1)
    r = _row(ws, r, 1, "Repair cost per square foot",
             val=f"=IFERROR(Rehab/{max(sqft,1)},0)", val_fmt=MONEY, fill=CALC_FILL)
    left_end = r

    r = top
    r = _band(ws, r, "Holding costs (monthly)", col=5, span=3,
              headers=("Annually", "Monthly"))
    r = _row(ws, r, 5, "Property taxes", pct=tax_annual, val=f"=ROUND({tax_annual}/12,0)",
             pct_fmt=MONEY, fill=CALC_FILL, pct_fill=INPUT_FILL)
    name("TaxAnnual", 6, r - 1)
    r = _row(ws, r, 5, "Insurance (vacant dwelling)", val=args.insurance,
             fill=INPUT_FILL)
    name("Insurance", 7, r - 1)
    r = _row(ws, r, 5, "Utilities", val=args.utilities, fill=INPUT_FILL)
    name("Utilities", 7, r - 1)
    r = _row(ws, r, 5, "Total monthly holding:",
             val="=ROUND(TaxAnnual/12,0)+Insurance+Utilities",
             bold=True, fill=TOTAL_FILL)
    name("HoldMonthly", 7, r - 1)
    r = _row(ws, r, 5, "Total over the hold",
             val="=HoldMonthly*HoldMonths", fill=CALC_FILL)
    name("HoldTotal", 7, r - 1)
    r = max(r, left_end) + 1

    # ── buying | selling transaction costs ─────────────────────────────
    top = r
    r = _band(ws, r, "Buying costs", col=1, span=3, headers=("Pct of purch", "Total"))
    r = _row(ws, r, 1, "Escrow / attorney fees", val=args.escrow_buy, fill=INPUT_FILL)
    name("EscrowBuy", 3, r - 1)
    r = _row(ws, r, 1, "Title insurance / search", pct=_num(args.title_pct),
             val="=ROUND(OfferToSeller*TitlePct,0)", fill=CALC_FILL,
             pct_fill=INPUT_FILL)
    name("TitlePct", 2, r - 1)
    r = _row(ws, r, 1, "Total buying costs:", val="=EscrowBuy+ROUND(OfferToSeller*TitlePct,0)",
             bold=True, fill=TOTAL_FILL)
    name("BuyCosts", 3, r - 1)
    left_end = r

    r = top
    r = _band(ws, r, "Selling costs", col=5, span=3, headers=("Pct of ARV", "Total"))
    r = _row(ws, r, 5, "Escrow / attorney, recording", val=args.escrow_sell,
             fill=INPUT_FILL)
    name("EscrowSell", 7, r - 1)
    r = _row(ws, r, 5, "Realtor fee", pct=_num(args.realtor_pct),
             val="=ROUND(ARV*RealtorPct,0)", fill=CALC_FILL, pct_fill=INPUT_FILL)
    name("RealtorPct", 6, r - 1)
    r = _row(ws, r, 5, "Transfer tax and conveyance", pct=_num(args.transfer_pct),
             val="=ROUND(ARV*TransferPct,0)", fill=CALC_FILL, pct_fill=INPUT_FILL)
    name("TransferPct", 6, r - 1)
    r = _row(ws, r, 5, "Total selling costs:",
             val="=EscrowSell+ROUND(ARV*RealtorPct,0)+ROUND(ARV*TransferPct,0)",
             bold=True, fill=TOTAL_FILL)
    name("SellCosts", 7, r - 1)
    r = max(r, left_end) + 1

    # ── THE OFFER ──────────────────────────────────────────────────────
    r = _band(ws, r, "THE OFFER", col=1, span=7)
    r = _row(ws, r, 1, "Buyer's maximum  (rule of thumb x ARV, less repairs)",
             val="=ROUND(RulePct*ARV-Rehab,0)", fill=CALC_FILL)
    name("BuyerMax", 3, r - 1)
    r = _row(ws, r, 1, "Our assignment fee", val=args.fee, fill=INPUT_FILL)
    name("Fee", 3, r - 1)

    c = ws.cell(row=r, column=1, value="OFFER TO SELLER")
    c.font = Font(bold=True, size=13, color=NAVY)
    v = ws.cell(row=r, column=3, value="=BuyerMax-Fee")
    v.number_format = MONEY
    v.font = Font(bold=True, size=13, color=GREEN)
    v.fill = TOTAL_FILL
    v.border = BOX
    name("OfferToSeller", 3, r)
    ws.cell(row=r, column=5,
            value="This is the number. Everything above it is an input, "
                  "everything below it is a consequence.").font = Font(size=9, color=GREY)
    ws.row_dimensions[r].height = 22
    r += 1
    r = _row(ws, r, 1, "What the buyer pays us at closing",
             val="=OfferToSeller+Fee", fill=CALC_FILL)
    name("BuyerPays", 3, r - 1)
    r = _row(ws, r, 1, "Offer as a percent of ARV",
             val="=IFERROR(OfferToSeller/ARV,0)", val_fmt=PCT1, fill=CALC_FILL)
    r = _row(ws, r, 1, "Offer against the as-is comp band",
             val="=IFERROR(OfferToSeller/AsIs,0)", val_fmt=PCT1, fill=CALC_FILL,
             note="Under 100% means we are buying below what as-is houses trade for here")
    r += 1

    # ── if we wholesale it (the default) ───────────────────────────────
    top = r
    r = _band(ws, r, "If we wholesale it   (the default)", col=1, span=3)
    r = _row(ws, r, 1, "Our fee", val="=Fee", bold=True, fill=TOTAL_FILL)
    r = _row(ws, r, 1, "Capital we put up", val=0, fill=CALC_FILL)
    r = _row(ws, r, 1, "Rehab risk we carry", val="none", val_fmt=None, fill=CALC_FILL)
    left_end = r

    r = top
    r = _band(ws, r, "What the buyer gets at that price", col=5, span=3)
    r = _row(ws, r, 5, "Buyer's net profit",
             val="=ARV-BuyerPays-Rehab-HoldTotal-SellCosts-EscrowBuy",
             bold=True, fill=TOTAL_FILL)
    name("BuyerProfit", 7, r - 1)
    r = _row(ws, r, 5, "Buyer's return on total cost",
             val="=IFERROR(BuyerProfit/(BuyerPays+Rehab+HoldTotal+SellCosts+EscrowBuy),0)",
             val_fmt=PCT1, fill=CALC_FILL)
    r = _row(ws, r, 5, "Buyer's all-in against ARV",
             val="=IFERROR((BuyerPays+Rehab)/ARV,0)", val_fmt=PCT1, fill=CALC_FILL)
    r = max(r, left_end) + 1

    # ── if we buy and rehab it ourselves ───────────────────────────────
    r = _band(ws, r, "If we buy and rehab it ourselves", col=1, span=7)
    top = r
    r = _row(ws, r, 1, "Estimated NET PROFIT",
             val="=ARV-OfferToSeller-Rehab-HoldTotal-BuyCosts-SellCosts",
             bold=True, fill=TOTAL_FILL)
    name("NetProfit", 3, r - 1)
    r = _row(ws, r, 1, "Committed capital",
             val="=OfferToSeller+Rehab+HoldTotal+BuyCosts", fill=CALC_FILL)
    name("Capital", 3, r - 1)
    r = _row(ws, r, 1, "Purchase and repair per square foot",
             val=f"=IFERROR((OfferToSeller+Rehab)/{max(sqft,1)},0)", fill=CALC_FILL)
    left_end = r

    r = top
    r = _row(ws, r, 5, "Return on total cost (ROI)",
             val="=IFERROR(NetProfit/Capital,0)", val_fmt=PCT1, bold=True,
             fill=TOTAL_FILL)
    r = _row(ws, r, 5, "Purchase plus rehab ROI",
             val="=IFERROR(NetProfit/(OfferToSeller+Rehab),0)", val_fmt=PCT1,
             fill=CALC_FILL)
    r = _row(ws, r, 5, "Annualized cash on cash",
             val="=IFERROR(NetProfit/Capital*12/MAX(HoldMonths,1),0)",
             val_fmt=PCT1, fill=CALC_FILL)
    r = max(r, left_end) + 1

    # ── ARV basis ──────────────────────────────────────────────────────
    comp_ids = set(arv_block.get("comp_ids") or [])
    comps = [c for c in _revive(pack.get("sold_comps"))
             if str(c.zpid) in comp_ids and c.sqft]
    comps.sort(key=lambda c: abs((c.sqft or 0) - sqft))
    r = _band(ws, r, "What the ARV stands on", col=1, span=7)
    hdr = ["Comp", "Bed/bath", "Sqft", "Sold", "$/sf", "Date", ""]
    for j, h in enumerate(hdr, 1):
        cell = ws.cell(row=r, column=j, value=h)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = _band_fill()
        cell.border = BOX
    r += 1
    for c in comps[:MAX_COMPS]:
        for j, v in enumerate([c.address, f"{c.beds}/{c.baths}", c.sqft,
                               c.price, round(c.ppsf), c.sold_date], 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = Font(size=9.5)
            cell.border = BOX
            cell.fill = CALC_FILL
            if j == 3:
                cell.number_format = NUM
            if j in (4, 5):
                cell.number_format = MONEY
        r += 1
    basis = tight.get("basis") or ""
    if basis:
        wide_rows.append(r)
        r = _para(ws, r, basis, size=8.5, color=GREY, span=7)
    r += 1

    # ── before you offer ───────────────────────────────────────────────
    gates = [g for g in (walk.get("gates") or [])][:MAX_GATES]
    if gates:
        r = _band(ws, r, "Before you offer", col=1, span=7)
        for g in gates:
            wide_rows.append(r)
            r = _para(ws, r, g, size=9, color=RED, span=7)

    for nm, ref in names.items():
        wb.defined_names[nm] = DefinedName(nm, attr_text=ref)

    _polish(ws)
    # _autofit sizes off the longest thing in each column, and the full-width
    # paragraphs (the ARV basis, the gates) live in column A, so it blows A, D
    # and E out to 54 and the page scrolls sideways. Fixed widths, then row
    # heights recomputed against the width the text actually gets.
    _finalize(ws, wide_rows)
    ws.freeze_panes = "A5"


def sheet_house(wb: Workbook, pack: dict, walk: dict) -> None:
    """The house and the seller: what the walk found, and who we are buying from.

    Condenses the old Overview and Repair Logic tabs into one. Overview carried
    the Sift lead, Repair Logic carried the flags and gates, and reading a deal
    meant bouncing between them.
    """
    subject, lead = pack["subject"], (pack.get("lead") or {})
    ws = wb.create_sheet("The House")
    wide: list[int] = []

    _title(ws, "The house and the seller",
           subject.get("full_address") or subject.get("address", ""),
           f"From the walk on {walk.get('walk_date', '')}. "
           f"{walk.get('media', '')}".strip())

    r = 5
    r = _band(ws, r, "Who we are buying from", col=1, span=7)
    for label, val in (
        ("Owner of record", (lead.get("owner_name") or "")
         + (" [DECEASED on record]" if lead.get("owner_deceased") else "")),
        ("Mailing", lead.get("mailing", "")),
        ("Sift record", lead.get("url", "")),
        ("Status / SIFTline", f"{lead.get('status','')} | "
         f"{', '.join(lead.get('cards') or []) or 'not on a board'}"),
        ("Tags", ", ".join(str(t) for t in (lead.get("tags") or [])) or "(none)"),
        ("Last contact", f"{lead.get('last_contact','')} "
         f"({lead.get('last_contact_type','')})".strip()),
    ):
        if val:
            r = _kv(ws, r, label, val, span=6, fill=CALC_FILL)
    for key in ("owner", "motivation", "acquired"):
        if walk.get(key):
            wide.append(r)
            ws.cell(row=r, column=1, value=key.capitalize()).font = Font(bold=True, size=9)
            r = _para(ws, r, walk[key], size=9, col=2, span=6)
    r += 1

    r = _band(ws, r, "What the walk found", col=1, span=7)
    for key, label in (("layout", "Layout"), ("distress", "Condition"),
                       ("note", "Note"), ("comp_finish_basis", "Finish set by the comps")):
        if walk.get(key):
            ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=9)
            wide.append(r)
            r = _para(ws, r, walk[key], size=8.5, col=2, span=6)
    r += 1

    open_items = walk.get("still_open") or []
    if open_items:
        r = _band(ws, r, "Still open", col=1, span=7)
        for i in open_items:
            ws.cell(row=r, column=1, value=i.get("item", "")).font = Font(bold=True, size=9)
            wide.append(r)
            r = _para(ws, r, i.get("detail", ""), size=8.5, col=2, span=6)
        r += 1

    flags = [f for f in (walk.get("flags") or []) if _num(f.get("cost"))]
    if flags:
        r = _band(ws, r, "Priced line items from the walk", col=1, span=7,
                  headers=("", "Cost"))
        for f in flags:
            ph = bool(f.get("placeholder"))
            c = ws.cell(row=r, column=1, value=f.get("item", ""))
            c.font = Font(size=9, color=RED if ph else "000000")
            v = ws.cell(row=r, column=3, value=_num(f.get("cost")))
            v.number_format = MONEY
            v.fill = CALC_FILL
            v.border = BOX
            v.font = Font(size=9, color=RED if ph else "000000")
            ws.cell(row=r, column=5, value=f.get("note", "")[:150]).font = Font(
                size=8, color=GREY)
            r += 1
        t = ws.cell(row=r, column=1, value="Total priced from the walk")
        t.font = Font(bold=True, size=10)
        v = ws.cell(row=r, column=3, value=sum(_num(f.get("cost")) for f in flags))
        v.number_format = MONEY
        v.font = Font(bold=True, size=10)
        v.fill = TOTAL_FILL
        v.border = BOX
        r += 1
        wide.append(r)
        r = _para(ws, r, "Figures in red are PLACEHOLDERS pending a signed bid. They "
                         "are realistic assumed costs so the total stays a true "
                         "number, never a blank.", size=8, color=RED, span=7)
        r += 1

    gates = walk.get("gates") or []
    if gates:
        r = _band(ws, r, "Before you offer", col=1, span=7)
        for g in gates:
            wide.append(r)
            r = _para(ws, r, g, size=9, color=RED, span=7)

    _polish(ws)
    _finalize(ws, wide, first_col=26.0)


def build_workbook(pack: dict, walk: dict, args, out: str) -> str:
    """Offer page in front, backup tabs behind it, each with one job."""
    wb = Workbook()
    wb.remove(wb.active)
    sheet_offer(wb, pack, walk, args)
    if not args.offer_only:
        sheet_house(wb, pack, walk)
        for fn, title in ((pw.sheet_repair_numbers, "Repair Detail"),
                          (pw.sheet_comps, "Comps"),
                          (pw.sheet_buyers, "Buyers")):
            try:
                before = set(wb.sheetnames)
                fn(wb, pack)
                new = [s for s in wb.sheetnames if s not in before]
                if new:
                    s = wb[new[0]]
                    s.title = title
                    # The reused grids bring their own widths (a 17-column comp
                    # table needs them). Give them the same print setup and a
                    # frozen header so they read like the rest of the book.
                    s.sheet_view.zoomScale = 100
                    s.page_setup.orientation = "landscape"
                    s.page_setup.fitToWidth = 1
                    s.page_setup.fitToHeight = 0
                    s.sheet_properties.pageSetUpPr.fitToPage = True
                    s.freeze_panes = "A3"
            except Exception as exc:      # a thin section is not a failed run
                logger.warning("%s sheet skipped: %s", title, exc)
    return _save(wb, out)


def _save(wb: Workbook, out: str) -> str:
    """Excel holds an exclusive lock on an open workbook, so write then swap."""
    dest = Path(out)
    tmp = dest.with_name(f"_PENDING_{dest.name}")
    wb.save(tmp)
    try:
        tmp.replace(dest)
        return str(dest)
    except PermissionError:
        logger.warning("%s is open in Excel and could not be replaced. "
                       "The new workbook is at %s: close Excel and rename it.",
                       dest.name, tmp.name)
        return str(tmp)


# Label | pct | value  x2, with a narrow gutter between the two halves.
COL_WIDTHS = [44.0, 12.0, 15.0, 3.0, 34.0, 12.0, 15.0]


def _finalize(ws, wide_rows: list[int], first_col: float | None = None) -> None:
    """Lock column widths and re-height only the rows that wrap full width."""
    widths = list(COL_WIDTHS)
    if first_col:
        widths[0] = first_col
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    span = sum(widths)
    for row in wide_rows:
        cell = ws.cell(row=row, column=1)
        text = str(cell.value or "")
        if not text:                      # paragraph rendered in column B
            text = str(ws.cell(row=row, column=2).value or "")
            cell = ws.cell(row=row, column=2)
        size = cell.font.size or 10
        # ~1 width unit per character at 11pt; scale for the smaller face.
        per_line = max(int(span * (11.0 / size)), 20)
        lines = max(1, -(-len(text) // per_line))
        ws.row_dimensions[row].height = max(14.0, lines * (size + 3.2) + 4)


def _band_fill():
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor="44506B")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--pack", required=True, help="post_walkthrough --save-pack JSON")
    ap.add_argument("--walk", help="walkthrough JSON (defaults to the one in the pack)")
    ap.add_argument("--out", help="output .xlsx")
    ap.add_argument("--rule", type=float, default=DEFAULT_RULE)
    ap.add_argument("--fee", type=float, default=DEFAULT_FEE)
    ap.add_argument("--hold", type=int, default=DEFAULT_HOLD)
    ap.add_argument("--arv", type=float, default=0)
    ap.add_argument("--as-is", dest="as_is", type=float, default=0)
    ap.add_argument("--tax-rate", type=float, default=0.0039)
    ap.add_argument("--insurance", type=float, default=DEFAULT_INSURANCE)
    ap.add_argument("--utilities", type=float, default=DEFAULT_UTILITIES)
    ap.add_argument("--escrow-buy", type=float, default=DEFAULT_ESCROW_BUY)
    ap.add_argument("--title-pct", type=float, default=DEFAULT_TITLE_PCT)
    ap.add_argument("--escrow-sell", type=float, default=DEFAULT_ESCROW_SELL)
    ap.add_argument("--realtor-pct", type=float, default=DEFAULT_REALTOR_PCT)
    ap.add_argument("--transfer-pct", type=float, default=DEFAULT_TRANSFER_PCT)
    ap.add_argument("--prepared", default="")
    ap.add_argument("--offer-only", action="store_true",
                    help="Render just the Offer page, no backup tabs")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()
    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO,
                        format="%(levelname)s %(message)s")

    pack = json.loads(Path(args.pack).read_text(encoding="utf8"))
    walk = (json.loads(Path(args.walk).read_text(encoding="utf8"))
            if args.walk else (pack.get("walk") or {}))
    _hydrate(pack, walk, args)
    subject = pack["subject"]
    safe = "".join(ch for ch in (subject.get("address") or "Offer")
                   if ch.isalnum() or ch in " -")[:44].strip().replace(" ", "_")
    out = args.out or f"{safe}_Offer.xlsx"
    out = build_workbook(pack, walk, args, out)

    rehab = build_rehab_matrix(subject, walk)
    rt = _num(next(iter((rehab.get("totals") or {"reno": 0}).values())))
    arv = _num(args.arv) or _num((pack.get("arv", {}).get("base") or {}).get("arv"))
    buyer_max = round(args.rule * arv - rt)
    print(f"ARV ${arv:,.0f} | rehab ${rt:,.0f} | rule {args.rule:.0%} | fee ${args.fee:,.0f}")
    print(f"Buyer's maximum ${buyer_max:,.0f}  ->  OFFER TO SELLER "
          f"${buyer_max - args.fee:,.0f}")
    print(f"workbook: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
