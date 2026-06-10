"""Probate runner XLSX → NoticeData intake.

Rick's county runner produces a weekly XLSX per NJ county (Essex,
Middlesex, Somerset, Union) carrying decedent + executor + property data
in a fixed 22-column layout. This module reads the file, normalizes the
quirks (Excel zip-code float, docket-number float, missing rows), and
emits NoticeData records that match what the rest of the pipeline expects
from a scraper. From there the normal enrichment_pipeline path handles
Smarty / obituary / heir verification / DM address waterfall etc.

The mapper deliberately does *not* invent new NoticeData fields — every
runner column lands on an existing field (or in raw_text as a structured
prefix for dedup_tracker / debugging). Adding fields to NoticeData would
ripple through every CSV consumer + the SIFT_COLUMNS schema, which the
"manual intake" feature isn't allowed to touch.
"""
from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from notice_parser import NoticeData

logger = logging.getLogger(__name__)

# Canonical county labels — match what the rest of the pipeline + Slack
# summaries expect. The CLI takes lower-case from the operator and we
# title-case it here so dedup keys land in a single bucket.
_COUNTY_CANONICAL = {
    "essex": "Essex",
    "middlesex": "Middlesex",
    "somerset": "Somerset",
    "union": "Union",
}

_CLASSIFICATION_MAP = {
    "P": "probate",
    "S": "probate_same_address",   # decedent's address == subject property
    "N": "probate_no_property",    # no real property associated
}

# Sentinel text the runner uses when a county had no filings that week —
# operators get a blank/placeholder file rather than nothing so the
# weekly batch hand-off is consistent.
_NO_FILES_SENTINELS = (
    "no files available",
    "no files",
    "no data available",
    "no records",
)


# ── Excel cell normalization ──────────────────────────────────────────


def _norm_str(v) -> str:
    """Coerce an XLSX cell to a clean str, stripping NBSP / trailing ws."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.replace("\xa0", " ").strip()
    if isinstance(v, float):
        # Excel default-types numeric cells as float; trim trailing .0
        # for integer values so docket "295526.0" doesn't show up.
        if v.is_integer():
            return str(int(v))
        return str(v).strip()
    return str(v).strip()


def _norm_zip(v) -> str:
    """Excel mangles leading-zero zips into floats (07009 → 7009.0).

    Zero-pad to 5 digits when the cell parses as an integer; pass
    through 9-digit ZIP+4 strings; leave anything weird as the raw
    str so downstream Smarty can flag it.
    """
    s = _norm_str(v)
    if not s:
        return ""
    # Handle "07009-1234" form — preserve as-is
    if "-" in s and len(s) >= 10:
        return s
    # Pure numeric → zero-pad
    if s.isdigit():
        return s.zfill(5)
    # "7009" with leading-zero stripped earlier
    if s.replace(".", "").isdigit():
        return s.split(".")[0].zfill(5)
    return s


def _norm_docket(v) -> str:
    """Normalize a docket cell. Middlesex stores numeric → "295526.0";
    Essex stores "2026-1086"; Somerset stores "26-00765"."""
    s = _norm_str(v)
    if not s:
        return ""
    # Float-form integer ("295526.0") → strip trailing .0
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".", 1)[0]
    return s


def _norm_date(v) -> str:
    """Coerce a runner date cell to YYYY-MM-DD. Accepts datetime
    objects (the openpyxl default for date-typed cells) and a handful
    of free-text formats."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = _norm_str(v)
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # last-resort: keep what's there


# ── Header → column index ─────────────────────────────────────────────

# Map runner header text (lowercase, whitespace-collapsed) to the
# semantic field name we'll use internally.
_HEADER_ALIASES = {
    "deceased full name": "decedent_full",
    "notes": "notes",
    "property address": "prop_addr",
    "property city": "prop_city",
    "property state": "prop_state",
    "property zip": "prop_zip",
    "probate /same /none": "classification",
    "probate/same/none": "classification",  # spacing variant
    "attorney on file": "attorney",
    "representative first name": "rep_first",
    "representative middle name": "rep_middle",
    "representative last name": "rep_last",
    "mailing address": "mail_addr",
    "mailing city": "mail_city",
    "mailing state": "mail_state",
    "mailing zip": "mail_zip",
    "relationship": "relationship",
    "money": "money",
    "phone 1": "phone_1",
    "phone 1 tag": "phone_1_tag",
    "owner deceased": "dod",
    "docket number": "docket",
    # Union County's runner XLSX labels this column "Case Number" instead
    # of "Docket Number". Without this alias the docket lands nowhere, so
    # dedup_tracker can't build a probate_runner ID and every Union record
    # logs "no dedup ID ... not tracked". Map both (+ "#" shorthands).
    "case number": "docket",
    "docket #": "docket",
    "case #": "docket",
    "file date": "file_date",
}


def _index_headers(header_row: tuple) -> dict[str, int]:
    """Build {semantic_name: column_index} from the XLSX header row."""
    out: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = re.sub(r"\s+", " ", str(cell)).strip().lower()
        sem = _HEADER_ALIASES.get(key)
        if sem:
            out[sem] = idx
    return out


# ── Row → NoticeData ──────────────────────────────────────────────────


def _split_decedent_name(full: str) -> tuple[str, str]:
    """'JOHN A SMITH' → ('JOHN A', 'SMITH'). First space splits first vs last."""
    full = full.strip()
    if not full:
        return "", ""
    parts = full.split(None, 1)
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def _row_to_notice(row: tuple, idx: dict[str, int], county_canonical: str) -> NoticeData | None:
    """Convert one XLSX data row to a NoticeData record.

    Returns None for rows that should be silently dropped (blank rows,
    sentinel "no files" rows). Missing property address is OK — probate
    enrichment fills it via the Knox-style tax-API tier in the pipeline,
    and the address waterfall handles mailing-only records.
    """
    def cell(name: str):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    decedent_full = _norm_str(cell("decedent_full"))
    # Sentinel: a row with the placeholder text in any of the first few
    # columns — emit a marker so the orchestrator can post a "no files"
    # Slack note instead of treating it as data.
    if any(s in decedent_full.lower() for s in _NO_FILES_SENTINELS):
        return None

    rep_first = _norm_str(cell("rep_first"))
    rep_middle = _norm_str(cell("rep_middle"))
    rep_last = _norm_str(cell("rep_last"))

    if not decedent_full and not (rep_first or rep_last):
        # Wholly empty row — pad/trailing row in XLSX.
        return None

    # Build executor full name (the PR == decision maker).
    exec_parts = [p for p in (rep_first, rep_middle, rep_last) if p]
    executor_full = " ".join(exec_parts)

    # Classification flag drives downstream filtering.
    classification_raw = _norm_str(cell("classification")).upper()
    classification = _CLASSIFICATION_MAP.get(classification_raw[:1], "probate")

    docket = _norm_docket(cell("docket"))
    attorney = _norm_str(cell("attorney"))
    phone_1 = _norm_str(cell("phone_1"))
    phone_1_tag = _norm_str(cell("phone_1_tag"))
    relationship = _norm_str(cell("relationship")).lower()
    notes = _norm_str(cell("notes"))

    # raw_text packs the structured runner fields the dedup tracker and
    # downstream debugging want. Format matches the prefix style used by
    # the rest of the scrapers (`Docket:`, `Attorney:`, …) so existing
    # regex extractors keep working.
    raw_bits: list[str] = [f"Source: probate_runner"]
    if docket:
        raw_bits.append(f"Docket: {docket}")
    raw_bits.append(f"County: {county_canonical}")
    if classification_raw:
        raw_bits.append(f"Classification: {classification_raw}")
    if attorney:
        raw_bits.append(f"Attorney: {attorney}")
    if phone_1:
        raw_bits.append(f"Phone1: {phone_1}")
    if phone_1_tag:
        raw_bits.append(f"Phone1Tag: {phone_1_tag}")
    if notes:
        raw_bits.append(f"Notes: {notes}")
    raw_text = " | ".join(raw_bits)

    # Synthetic source_url — kept consistent across counties so the
    # dedup tracker can pull (county, docket) without hand-holding.
    source_url = f"probate_runner://{county_canonical.lower()}/{docket or 'unknown'}"

    n = NoticeData(
        date_added=_norm_date(cell("file_date")) or datetime.now().strftime("%Y-%m-%d"),
        address=_norm_str(cell("prop_addr")),
        city=_norm_str(cell("prop_city")),
        state=_norm_str(cell("prop_state")) or "NJ",
        zip=_norm_zip(cell("prop_zip")),
        owner_name=executor_full or decedent_full,
        notice_type="probate",
        county=county_canonical,
        source_url=source_url,
        raw_text=raw_text,
        decedent_name=decedent_full,
        date_of_death=_norm_date(cell("dod")),
        owner_street=_norm_str(cell("mail_addr")),
        owner_city=_norm_str(cell("mail_city")),
        owner_state=_norm_str(cell("mail_state")) or "NJ",
        owner_zip=_norm_zip(cell("mail_zip")),
    )

    # Court-named executor → high-confidence decision maker. Mirrors the
    # pattern in nj_middlesex_probate._row_to_notice so the downstream
    # obituary preset routes through the same code path and never
    # overrides the runner's executor with a wrong obit match.
    if executor_full:
        n.decision_maker_name = executor_full
        n.decision_maker_relationship = relationship
        n.decision_maker_source = "court_record"
        n.decision_maker_status = "verified_living"
        n.dm_confidence = "high"
        # Mailing address is the DM's residential mailing address — copy
        # so the heir-verification waterfall doesn't immediately try to
        # re-discover it.
        n.decision_maker_street = _norm_str(cell("mail_addr"))
        n.decision_maker_city = _norm_str(cell("mail_city"))
        n.decision_maker_state = _norm_str(cell("mail_state")) or "NJ"
        n.decision_maker_zip = _norm_zip(cell("mail_zip"))

    # Decedent confirmed deceased — court filing is authoritative, no need
    # to re-confirm via obituary search. The obit step still fires to pull
    # survivors / additional heirs, but it won't override owner_deceased.
    n.owner_deceased = "yes"

    # If the runner flagged "S" (subject == decedent address) or "N" (no
    # property), stash that on the classification slot in raw_text only —
    # we keep notice_type=probate so the existing pipeline doesn't choke
    # on an unknown type. Downstream consumers that care can split on
    # "Classification: S" / "Classification: N".
    if classification != "probate":
        logger.debug("Row classified %s (%s)", classification_raw, decedent_full)

    return n


# ── Public entry ──────────────────────────────────────────────────────


def parse_runner_workbook(
    source: Path | bytes | str,
    county: str,
    *,
    sheet_name: str | None = None,
) -> tuple[list[NoticeData], dict]:
    """Read a runner XLSX and return (notices, stats).

    `source` is either a filesystem path or in-memory bytes (used by the
    Modal cloud function — Modal passes the file as bytes from the local
    entrypoint). `county` is one of essex / middlesex / somerset / union
    (case-insensitive). Returns the parsed notices plus a stats dict for
    Slack reporting.

    Stats keys:
      - rows_read: data rows seen in the sheet
      - rows_parsed: rows that became NoticeData
      - rows_skipped_blank: rows dropped because they were blank
      - sentinel_hit: True if a "no files available" row was detected
    """
    county_lc = county.strip().lower()
    if county_lc not in _COUNTY_CANONICAL:
        raise ValueError(
            f"unknown county {county!r}; expected one of {sorted(_COUNTY_CANONICAL)}"
        )
    county_canonical = _COUNTY_CANONICAL[county_lc]

    if isinstance(source, (bytes, bytearray)):
        import io
        wb = load_workbook(io.BytesIO(source), data_only=True, read_only=True)
    else:
        wb = load_workbook(Path(source), data_only=True, read_only=True)

    ws = wb[sheet_name] if sheet_name else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return [], {
            "rows_read": 0, "rows_parsed": 0,
            "rows_skipped_blank": 0, "sentinel_hit": True,
        }
    idx = _index_headers(header_row)
    if not idx:
        # The "no files available" placeholder XLSX usually has either a
        # banner row instead of headers OR a header row that doesn't
        # match. Surface as sentinel so the orchestrator skips silently.
        logger.warning("No recognized headers in %s — treating as sentinel", county)
        wb.close()
        return [], {
            "rows_read": 0, "rows_parsed": 0,
            "rows_skipped_blank": 0, "sentinel_hit": True,
        }

    notices: list[NoticeData] = []
    rows_read = 0
    rows_skipped_blank = 0
    sentinel_hit = False
    for row in rows_iter:
        if row is None:
            continue
        rows_read += 1
        # Quick all-empty check before the more expensive parsing.
        if all(cell in (None, "") for cell in row):
            rows_skipped_blank += 1
            continue
        try:
            n = _row_to_notice(row, idx, county_canonical)
        except Exception as e:
            logger.warning("Row parse failed (%s): %s", county_canonical, e)
            n = None
        if n is None:
            # Differentiate: sentinel rows have recognizable text in the
            # decedent column, blank rows don't.
            joined = " ".join(str(c) for c in row if c is not None).lower()
            if any(s in joined for s in _NO_FILES_SENTINELS):
                sentinel_hit = True
            else:
                rows_skipped_blank += 1
            continue
        notices.append(n)

    wb.close()
    logger.info(
        "Probate runner %s: read %d rows, parsed %d, skipped %d blank, sentinel=%s",
        county_canonical, rows_read, len(notices), rows_skipped_blank, sentinel_hit,
    )
    return notices, {
        "rows_read": rows_read,
        "rows_parsed": len(notices),
        "rows_skipped_blank": rows_skipped_blank,
        "sentinel_hit": sentinel_hit,
    }


if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Usage: python -m probate_intake <county> <path/to/file.xlsx>")
        sys.exit(1)
    notices, stats = parse_runner_workbook(Path(sys.argv[2]), sys.argv[1])
    print(f"Stats: {stats}")
    for n in notices[:3]:
        print(f"  decedent={n.decedent_name!r}  exec={n.decision_maker_name!r}  "
              f"addr={n.address!r} {n.city!r} {n.zip!r}  docket={n.source_url}")
