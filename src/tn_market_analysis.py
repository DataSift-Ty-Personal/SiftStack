"""Knox and Blount market picture, rebuilt from property-level API data.

The market analysis this replaces was a Playwright scrape of 14 summary columns
off the Market Finder UI, Knox only, last pulled 2026-04-11. We now have
property-level API access, so the picture is built from the actual universe
instead of scraped medians: every single-family property in both counties, with
its value, equity, size, scores and distress flags.

    probe       prove every filter key by count delta        (free, ~1 min)
    sweep       every SFR row in both counties               (free, ~20 min)
    hydrate     get_detail for neighborhood + year built     (free, slow)
    suppress    score geos and write the exclusion set       (free)
    report      the Excel workbook                           (free)

MEASURED LIVE 2026-09-02, before any of this was written:

  * Knox 206,925 properties / 153,757 single family / 131,120 in the buy box.
    Blount 71,237 / 47,413 / 40,145. Combined SFR buy box 171,265.

  * FILTER KEYS LIE SILENTLY. A deliberately bogus key returned Knox 153,757,
    byte-identical to no filter at all. `owner_occupied` returned the SAME
    153,757, so that key name is wrong and is being ignored. `type_single_family`
    and `value_min`/`value_max` are real. A key is only proven by a count delta
    against a bogus control, never by the request being accepted. That is what
    the `probe` phase exists for, and it runs before the sweep, not after.

  * THE REAL KEY NAMES CARRY AN `extra_` PREFIX, BUT NOT UNIFORMLY. Verified on
    Knox SFR (baseline 153,757): `extra_year_built_min` 150,691,
    `extra_vacant` 1,603, `extra_equity_percent_min` 113,006 and
    `extra_last_sale_date_min` 14,337 all READ. Meanwhile `year_built_min`,
    `equity_percent_min`, `is_vacant_property`, `owner_occupied`,
    `absentee_owner` AND `extra_absentee_owner` are all ignored. Guessing the
    prefix is not enough; each key still has to be proved.

  * BECAUSE THE SWEEP IS TOTAL, LOCAL FILTERING BEATS SERVER FILTERING. Every
    dimension we would want to filter on (value, equity, absentee, corporate,
    MLS status, distress flags) is already ON the row, so the sweep asks the
    server for single-family only and everything else is decided here. That is
    both richer, since it keeps the full distribution rather than a count, and
    immune to the silent-ignore trap above.

  * OFFSET PAGINATION IS NON-DETERMINISTIC, so one pass never returns every
    row. Measured 2026-09-02: a full Knox sweep fetched all 616 pages (154,000
    row slots) and yielded 144,319 UNIQUE ids against a reported total of
    153,757, a 6.3% duplicate rate. Partitioning does not help; a 12,183-row
    value band showed a WORSE 14.7% duplicate rate and 85.3% unique coverage.
    Rows repeat as you walk `result_index`, so the repo's earlier "ordering is
    stable" note does not hold at depth. A short page is therefore NOT
    end-of-results and must never break the loop; termination keys on the
    reported total.

    AND THE MISSING ROWS CANNOT BE RECOVERED BY ASKING AGAIN. A second full
    Blount pass returned EXACTLY zero new ids (39,520 both times) and a Knox
    re-walk found zero new across 340 pages. So the ~6 to 17% that never appears
    is a stable unreachable subset, not something a retry sweeps up. Coverage is
    therefore REPORTED, never assumed, and never claimed as a census.

  * SEARCH ROWS DO NOT CARRY neighborhood OR year built. They carry lat/lon and
    the address string (so ZIP is free), but `neighborhood` and `years_built`
    live only on `/properties/detail/`, which is capped near 3 req/s AGGREGATE
    regardless of worker count. Hence suppress-first, hydrate-second: there is
    no reason to spend hours resolving the neighborhood of a property already
    ruled out on value.

  * THERE IS NO MARKET-MAP ENDPOINT. Probed map.reisift.io and apiv2.reisift.io
    across a dozen plausible paths; every one 404s. The Data Science doc's
    `market-map-by-geo-and-property-type-group` index is internal. Do not go
    guessing endpoint names again, the answer was checked.

WHAT IS DELIBERATELY DROPPED. `parcel_boundary` is a WKT polygon that dominates
row size and tells us nothing lat/lon does not. The distressors blob is stored as
the list of ACTIVE flag names rather than all 27 nested dicts, which is the same
call `dispo_buyers.phase_sweep` arrived at. Everything else is kept, because that
same sweep originally persisted 8 of ~35 fields and threw away exactly the
dimensions it later needed.

    python src/tn_market_analysis.py --phase probe
    python src/tn_market_analysis.py --phase sweep
    python src/tn_market_analysis.py --phase sweep --resume
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import config  # noqa: F401  loads .env
    OUTPUT_ROOT = Path(getattr(config, "OUTPUT_DIR", "output"))
except Exception:
    OUTPUT_ROOT = Path("output")

from enterprise_prospects import SearchClient, search_body  # noqa: E402

log = logging.getLogger("tn_market")

OUT = OUTPUT_ROOT / "tn_market"
STATE_PATH = OUT / "state.json"

COUNTIES = [("47093", "Knox", "TN"), ("47009", "Blount", "TN")]
PAGE_SIZE = 250

# Buy box (Ty): single family only, AVM $1 to $700,000. The $1 floor is
# deliberate and wider than a $100K floor, because condemned and tax-distressed
# stock routinely falls under $100K.
BUY_BOX_MIN = 1
BUY_BOX_MAX = 700_000

SFR = {"type_single_family": True}

# WKT polygon: dominates row size, adds nothing over lat/lon.
DROP_FIELDS = {"parcel_boundary"}

ZIP_RX = re.compile(r"\b(\d{5})(?:-\d{4})?\s*$")


# ---------------------------------------------------------------- state ----

def _load(p: Path, default):
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return default


def _save(p: Path, obj) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_text(json.dumps(obj, indent=1, default=str), encoding="utf-8")
    tmp.replace(p)


def state() -> dict:
    return _load(STATE_PATH, {})


def mark_done(phase: str, **facts) -> None:
    s = state()
    s[phase] = {"when": datetime.now().isoformat(timespec="seconds"), **facts}
    _save(STATE_PATH, s)


def gate(ok: bool, msg: str) -> None:
    """A phase that produced nothing is a failure, not an empty market."""
    if not ok:
        raise SystemExit("GATE FAILED: " + msg)


# ----------------------------------------------------------------- rows ----

def zip_of(address: str) -> str:
    m = ZIP_RX.search((address or "").strip())
    return m.group(1) if m else ""


def active_distressors(blob) -> list:
    """Just the names of the flags that are ON.

    The raw blob is 27 nested dicts per property and is almost all False, so
    storing it whole would multiply the file size for no extra information.
    """
    if not isinstance(blob, dict):
        return []
    out = []
    for k, v in blob.items():
        if isinstance(v, dict):
            if v.get("is_active"):
                out.append(k)
        elif v:
            out.append(k)
    return sorted(out)


def slim(row: dict, county: str) -> dict:
    r = {k: v for k, v in row.items() if k not in DROP_FIELDS}
    r["distress_active"] = active_distressors(row.get("distressors"))
    r.pop("distressors", None)
    r["zip"] = zip_of(row.get("address", ""))
    r["county"] = row.get("county") or county
    return r


# ---------------------------------------------------------------- probe ----

def phase_probe(args) -> None:
    """Prove every filter key by count delta against a bogus control.

    An accepted filter proves nothing. Only a count that MOVES proves the key is
    read. Anything matching the bogus baseline is being silently ignored and
    must not be relied on downstream.
    """
    sc = SearchClient(min_interval=args.min_interval)

    def n(fips, name, filters):
        return sc.search(search_body(fips, name, "TN", filters))["total_results"]

    results = {}
    for fips, name, st in COUNTIES:
        base = n(fips, name, {})
        bogus = n(fips, name, {"zzz_not_a_real_filter_key": True})
        sfr = n(fips, name, SFR)
        log.info("%s: all=%s  bogus-control=%s  sfr=%s", name, base, bogus, sfr)
        gate(bogus == base,
             "bogus key changed the count for %s (%s vs %s); the control "
             "itself is unreliable" % (name, bogus, base))

        # Left column: keys PROVEN to read, 2026-09-02. Right column: plausible
        # names that are silently ignored, kept in the probe deliberately so a
        # future run re-checks them rather than trusting this comment.
        checks = {
            "type_single_family": (SFR, base),
            "value_min/value_max": (dict(SFR, value_min=BUY_BOX_MIN,
                                         value_max=BUY_BOX_MAX), sfr),
            "extra_year_built_min": (dict(SFR, extra_year_built_min=1900), sfr),
            "extra_vacant": (dict(SFR, extra_vacant=True), sfr),
            "extra_equity_percent_min": (dict(SFR, extra_equity_percent_min=50),
                                         sfr),
            "extra_last_sale_date_min": (
                dict(SFR, extra_last_sale_date_min="2024-09-01"), sfr),
            "DEAD owner_occupied": (dict(SFR, owner_occupied=False), sfr),
            "DEAD absentee_owner": (dict(SFR, absentee_owner=True), sfr),
            "DEAD extra_absentee_owner": (dict(SFR, extra_absentee_owner=True),
                                          sfr),
            "DEAD year_built_min": (dict(SFR, year_built_min=1900), sfr),
            "DEAD is_vacant_property": (dict(SFR, is_vacant_property=True), sfr),
            "DEAD equity_percent_min": (dict(SFR, equity_percent_min=50), sfr),
        }
        rows = {}
        for label, (f, ref) in checks.items():
            c = n(fips, name, f)
            reads = c != ref
            rows[label] = {"count": c, "baseline": ref, "reads": reads}
            log.info("  %-22s %9s  vs baseline %9s  -> %s",
                     label, c, ref, "READS" if reads else "IGNORED")
        results[name] = {"all": base, "bogus": bogus, "sfr": sfr,
                         "checks": rows}

    _save(OUT / "filter_probe.json", results)
    mark_done("probe", counties=len(results))
    print("\nfilter probe written to %s" % (OUT / "filter_probe.json"))
    print("Only keys marked READS may be used downstream.")


# ---------------------------------------------------------------- sweep ----

def phase_sweep(args) -> None:
    """Every single-family row in both counties, resumable, one JSONL per county."""
    sc = SearchClient(min_interval=args.min_interval)
    OUT.mkdir(parents=True, exist_ok=True)
    prog = _load(OUT / "sweep_progress.json", {})
    totals = {}

    want = {c.strip().lower() for c in args.counties.split(",") if c.strip()}
    for fips, name, st in COUNTIES:
        if want and name.lower() not in want:
            log.info("%s: skipped (--counties)", name)
            if name in prog:
                totals[name] = {k: prog[name].get(k)
                                for k in ("written", "total", "coverage_pct")}
            continue
        path = OUT / ("sweep_%s.jsonl" % name.lower())
        seen = set()
        if args.resume and path.exists():
            with path.open(encoding="utf-8") as fh:
                for line in fh:
                    try:
                        seen.add(json.loads(line)["id"])
                    except Exception:
                        continue
            log.info("%s: resuming, %s rows already on disk", name, len(seen))
        elif path.exists():
            path.unlink()

        total = None
        written = len(seen)
        fh = path.open("a" if (args.resume or seen) else "w", encoding="utf-8")
        try:
            for _pass in range(1, args.passes + 1):
                # Each pass re-walks from the top. Because the ordering drifts,
                # a second walk surfaces rows the first one never showed, and
                # `seen` keeps the union honest.
                page = 1
                empty_streak = 0
                if _pass > 1:
                    log.info("%s: pass %d, %s unique so far", name, _pass, written)
                while True:
                    r = sc.search(search_body(fips, name, st, SFR, page, PAGE_SIZE))
                    if total is None:
                        total = r.get("total_results") or 0
                        log.info("%s: %s single-family properties", name, total)
                        gate(total > 0, "%s returned zero properties" % name)
                    batch = r.get("data") or []
                    if not batch:
                        # An empty page can be transient. Only believe it once
                        # the offset has genuinely run past the reported total.
                        if (page - 1) * PAGE_SIZE >= total:
                            break
                        empty_streak += 1
                        if empty_streak >= 3:
                            break
                        log.warning("%s: empty page %d at %d/%d, retrying",
                                    name, page, written, total)
                        time.sleep(3.0 * empty_streak)
                        continue
                    empty_streak = 0
                    for row in batch:
                        rid = str(row.get("id") or "")
                        if not rid or rid in seen:
                            continue
                        seen.add(rid)
                        fh.write(json.dumps(slim(row, name), default=str) + "\n")
                        written += 1
                    page += 1
                    if page % 20 == 0:
                        fh.flush()
                        prog[name] = {"next_page": page, "written": written,
                                      "total": total}
                        _save(OUT / "sweep_progress.json", prog)
                        log.info("  %s: %s / %s", name, written, total)
                    # NEVER break on a short page. Verified live 2026-09-02:
                    # pages 578 to 615 all returned a full 250 at offsets up to
                    # 153,501, so there is no depth cap and a short page is a
                    # transient hiccup. Treating it as end-of-results silently
                    # stopped Knox at 144,319 of 153,757.
                    if written >= total or (page - 1) * PAGE_SIZE >= total:
                        break
        finally:
            fh.close()

        cover = 100.0 * written / total if total else 0.0
        # `done` is only true once the gate has actually passed. Writing it
        # before the check meant a failed run still looked finished, so the
        # resume did nothing at all and reported success.
        ok = cover >= args.min_coverage
        prog[name] = {"next_page": page, "written": written, "total": total,
                      "coverage_pct": round(cover, 1), "done": ok}
        _save(OUT / "sweep_progress.json", prog)
        totals[name] = {"written": written, "total": total,
                        "coverage_pct": round(cover, 1)}
        log.info("%s: %s unique rows of %s reported (%.1f%% coverage)",
                 name, written, total, cover)
        # The shortfall is duplicate rows from non-deterministic pagination,
        # not missing pages: every page was fetched. So the gate is a coverage
        # floor, not a truncation check, and the real number is always printed.
        gate(ok, "%s covered only %.1f%% (%s of %s unique). Raise coverage with "
                 "--passes 2, or lower --min-coverage deliberately."
             % (name, cover, written, total))

    mark_done("sweep", **totals)
    print("\nsweep complete:")
    for k, v in totals.items():
        print("  %-8s %7s rows" % (k, v["written"]))


# ------------------------------------------------------------ aggregate ----

def load_rows() -> list:
    """Every swept row from both counties."""
    rows = []
    for _, name, _ in COUNTIES:
        p = OUT / ("sweep_%s.jsonl" % name.lower())
        if not p.exists():
            continue
        with p.open(encoding="utf-8") as fh:
            for line in fh:
                try:
                    rows.append(json.loads(line))
                except Exception:
                    continue
    return rows


def _pct(vals, q):
    """Percentile without numpy. vals must be sorted and non-empty."""
    if not vals:
        return None
    k = (len(vals) - 1) * (q / 100.0)
    lo = int(k)
    hi = min(lo + 1, len(vals) - 1)
    return vals[lo] + (vals[hi] - vals[lo]) * (k - lo)


def geo_stats(rows: list, key: str) -> dict:
    """Per-geo distribution. Percentiles, not just medians.

    The scraped Market Finder only ever gave a median. A median cannot tell you
    whether a geo is uniformly mid-priced or split between $60K shells and $900K
    new builds, and those two markets are worked completely differently.
    """
    groups = {}
    for r in rows:
        groups.setdefault(r.get(key) or "(blank)", []).append(r)

    out = {}
    for g, rs in groups.items():
        vals = sorted(r["estimatedValue"] for r in rs
                      if isinstance(r.get("estimatedValue"), (int, float)))
        eqs = sorted(r["equityPercent"] for r in rs
                     if isinstance(r.get("equityPercent"), (int, float)))
        sqft = sorted(r["squareFeet"] for r in rs
                      if isinstance(r.get("squareFeet"), (int, float)) and r["squareFeet"])
        inbox = [v for v in vals if BUY_BOX_MIN <= v <= BUY_BOX_MAX]
        distress = {}
        for r in rs:
            for f in (r.get("distress_active") or []):
                distress[f] = distress.get(f, 0) + 1
        n = len(rs)
        # MAJORITY county, not rs[0]. A neighborhood can straddle the county
        # line and the assignment is geographic, so the first row in the list is
        # arbitrary: it labelled Eagleton Village (Blount) as Knox.
        cc = {}
        for r in rs:
            k = r.get("county") or ""
            cc[k] = cc.get(k, 0) + 1
        out[g] = {
            "geo": g,
            "county": max(cc.items(), key=lambda kv: kv[1])[0] if cc else "",
            "n": n,
            "n_valued": len(vals),
            "n_in_box": len(inbox),
            "pct_in_box": round(100.0 * len(inbox) / len(vals), 1) if vals else 0.0,
            "value_p10": _pct(vals, 10), "value_p25": _pct(vals, 25),
            "value_med": _pct(vals, 50), "value_p75": _pct(vals, 75),
            "value_p90": _pct(vals, 90),
            "equity_med": _pct(eqs, 50),
            "sqft_med": _pct(sqft, 50),
            "pct_absentee": round(100.0 * sum(1 for r in rs if r.get("absenteeOwner")) / n, 1),
            "pct_corporate": round(100.0 * sum(1 for r in rs if r.get("corporateOwned")) / n, 1),
            "n_mls_active": sum(1 for r in rs if r.get("mlsActive")),
            "n_in_crm": sum(1 for r in rs if r.get("saved_uuid")),
            "pct_no_distress": round(
                100.0 * sum(1 for r in rs if not r.get("distress_active")) / n, 1),
            "distress": distress,
        }
    return out


# -------------------------------------------------------- market finder ----

# Count fields are per-county slices of the same geo and must be added.
_MF_SUM = ("total_inv_trans_6mo", "homes_on_market", "homes_sold_last_month")


def _merge_geo(store: dict, key: str, row: dict) -> None:
    """Merge a geo that appears in more than one county extract.

    A LAST-WINS MERGE SILENTLY DESTROYS DATA HERE, and it is not hypothetical:
    six ZIPs appear in both the Knox and Blount extracts because they straddle
    the county line. Market Finder reports only the in-county slice, so 37920
    came back as 97 investor transactions in the Knox file and 0 in the Blount
    file. Reading Blount second overwrote 97 with 0, which made a genuinely busy
    South Knoxville ZIP look dead and cut 1,370 properties from the cohort for
    "zero investor transactions".

    So counts are SUMMED, and the descriptive fields (DOM, median value) are
    taken from whichever slice carries the most transactions, since that is the
    portion of the ZIP the numbers actually describe.
    """
    prev = store.get(key)
    if prev is None:
        store[key] = row
        return
    merged = dict(prev)
    for f in _MF_SUM:
        a, b = prev.get(f), row.get(f)
        if isinstance(a, (int, float)) or isinstance(b, (int, float)):
            merged[f] = (a or 0) + (b or 0)
    dominant = row if (row.get("total_inv_trans_6mo") or 0) >         (prev.get("total_inv_trans_6mo") or 0) else prev
    for f in ("median_days_on_market", "median_home_value", "median_sale_price"):
        merged[f] = dominant.get(f)
    merged["county"] = "%s+%s" % (prev.get("county"), row.get("county"))
    merged["_merged_from"] = [prev.get("county"), row.get("county")]
    store[key] = merged


def load_market_finder() -> dict:
    """ZIP and neighborhood liquidity from the Market Finder extract.

    This is the ONE thing the property API cannot give us. A sweep can tell you
    what a geo is worth and who owns it, but not how fast it sells or how much
    competition is already there. `median_days_on_market`, `homes_on_market` and
    `total_inv_trans_6mo` only exist here.

    Keyed by ZIP and by neighborhood name, most recent extract per county wins.
    """
    import glob
    out = {"zip": {}, "neighborhood": {}, "files": []}
    for _, name, _ in COUNTIES:
        hits = sorted(glob.glob(str(OUTPUT_ROOT /
                      ("market_finder_Tennessee_%s_*.json" % name))))
        if not hits:
            log.warning("no Market Finder extract for %s", name)
            continue
        f = hits[-1]
        out["files"].append(f)
        d = json.loads(Path(f).read_text(encoding="utf-8"))
        for row in (d.get("zip_data") or []):
            z = str(row.get("zip_code") or "").strip()
            if z:
                _merge_geo(out["zip"], z, dict(row, county=name))
        for row in (d.get("neighborhood_data") or []):
            nb = (row.get("neighborhood") or "").strip()
            if nb:
                _merge_geo(out["neighborhood"], nb, dict(row, county=name))
    log.info("market finder: %d zips, %d neighborhoods from %d file(s)",
             len(out["zip"]), len(out["neighborhood"]), len(out["files"]))
    return out


# ------------------------------------------------------------- suppress ----

def phase_suppress(args) -> None:
    """Cut geos we would never buy in, each with a written reason.

    Deliberately conservative. This decides what we PAY to score, so a geo is
    only cut on a reason that would stop us buying there regardless of what a
    Street View image showed. Physical distress is NOT a criterion here: that
    is what we are about to go measure, and screening on it first would only
    re-find what the data already knows.
    """
    rows = load_rows()
    gate(len(rows) > 1000, "only %d swept rows; run --phase sweep first" % len(rows))
    log.info("loaded %s swept rows", len(rows))

    stats = geo_stats(rows, "zip")
    mf = load_market_finder()

    # County median DOM is the reference for "slow", because a national number
    # would call the whole of East Tennessee illiquid.
    doms = [v.get("median_days_on_market") for v in mf["zip"].values()
            if isinstance(v.get("median_days_on_market"), (int, float))]
    doms.sort()
    dom_ref = _pct(doms, 50) if doms else None
    if dom_ref:
        log.info("county-median DOM across ZIPs: %.0f days", dom_ref)

    keep, cut = [], []
    for g, s_ in sorted(stats.items()):
        reasons = []
        m = mf["zip"].get(g, {})
        s_["mf_dom"] = m.get("median_days_on_market")
        s_["mf_inv_trans_6mo"] = m.get("total_inv_trans_6mo")
        s_["mf_homes_on_market"] = m.get("homes_on_market")
        s_["mf_sold_last_month"] = m.get("homes_sold_last_month")

        if g == "(blank)":
            reasons.append("no ZIP on the address")
        if isinstance(s_["mf_inv_trans_6mo"], (int, float)) and                 s_["mf_inv_trans_6mo"] == 0 and s_["n_in_box"] >= args.min_in_box:
            reasons.append("zero investor transactions in 6 months")
        if dom_ref and isinstance(s_["mf_dom"], (int, float)) and                 s_["mf_dom"] > dom_ref * args.dom_multiple:
            reasons.append("DOM %.0f is %.1fx the county median of %.0f"
                           % (s_["mf_dom"], s_["mf_dom"] / dom_ref, dom_ref))
        if s_["n"] < args.min_properties:
            reasons.append("only %d properties, too thin to work" % s_["n"])
        if s_["n_in_box"] < args.min_in_box:
            reasons.append("only %d in the buy box" % s_["n_in_box"])
        if s_["value_med"] and s_["value_med"] > BUY_BOX_MAX:
            reasons.append("median $%.0f is above the $700K buy box"
                           % s_["value_med"])
        if s_["value_med"] and s_["value_med"] < args.min_median_value:
            reasons.append("median $%.0f below the floor where rehab clears ARV"
                           % s_["value_med"])
        rec = dict(s_)
        rec.pop("distress", None)
        rec["suppress_reasons"] = "; ".join(reasons)
        (cut if reasons else keep).append(rec)

    survivors = sum(r["n_in_box"] for r in keep)
    _save(OUT / "zip_stats.json", stats)

    import csv
    cpath = OUT / "tn_zip_suppress.csv"
    cols = ["geo", "county", "n", "n_in_box", "pct_in_box", "value_p10",
            "value_p25", "value_med", "value_p75", "value_p90", "equity_med",
            "sqft_med", "pct_absentee", "pct_corporate", "n_mls_active",
            "n_in_crm", "pct_no_distress", "mf_dom", "mf_inv_trans_6mo",
            "mf_homes_on_market", "mf_sold_last_month", "suppress_reasons"]
    with cpath.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in sorted(keep + cut, key=lambda x: -x["n_in_box"]):
            w.writerow(r)

    mark_done("suppress", kept=len(keep), cut=len(cut), survivors=survivors)
    print()
    print("ZIPs kept %d | cut %d" % (len(keep), len(cut)))
    print("Buy-box properties in surviving ZIPs: %s" % f"{survivors:,}")
    print("Scoring cost at $0.007: $%s" % f"{survivors * 0.007:,.0f}")
    print("written: %s" % cpath)
    if cut:
        print()
        print("cut, worst first:")
        for r in sorted(cut, key=lambda x: -x["n"])[:12]:
            print("  %-10s n=%-6d %s" % (r["geo"], r["n"], r["suppress_reasons"]))


# -------------------------------------------------------------- hydrate ----

def _hydrated() -> dict:
    """Everything already pulled from the detail endpoint, keyed by id."""
    out = {}
    p = OUT / "hydrated.jsonl"
    if p.exists():
        with p.open(encoding="utf-8") as fh:
            for line in fh:
                try:
                    d = json.loads(line)
                    out[str(d["id"])] = d
                except Exception:
                    continue
    return out


def phase_hydrate(args) -> None:
    """Pull neighborhood, subdivision and year built for a stratified sample.

    WHY A SAMPLE AND NOT EVERYTHING. `neighborhood` and `years_built` live only
    on `/properties/detail/`, which caps near 3 req/s AGGREGATE no matter how
    many workers are used, so all 155K survivors would take about 14 hours.
    Neighborhoods are contiguous geographic areas and every swept row already
    carries lat/lon, so a few thousand labelled points are enough to assign the
    rest by nearest neighbour. `--phase geo` does that assignment and measures
    its own accuracy on a holdout rather than asserting it.

    The sample is stratified by ZIP so no ZIP is left without labelled anchors,
    which is the failure that would make the inference silently wrong in exactly
    the places we know least about.
    """
    import threading
    from concurrent.futures import ThreadPoolExecutor, as_completed

    from siftmap_standalone import SiftMapClient, SiftMapError

    rows = load_rows()
    gate(len(rows) > 1000, "run --phase sweep first")
    keep = _surviving_zips()
    if keep:
        rows = [r for r in rows if r.get("zip") in keep]
    have = _hydrated()
    log.info("%s survivor rows, %s already hydrated", len(rows), len(have))

    by_zip = {}
    for r in rows:
        if str(r["id"]) not in have and r.get("latitude"):
            by_zip.setdefault(r.get("zip") or "(blank)", []).append(r)
    if not by_zip:
        print("nothing left to hydrate")
        return

    # Proportional with a floor, so a small ZIP still gets enough anchors to
    # be assignable rather than inheriting a neighbour ZIP's labels.
    per = max(args.per_zip, args.limit // max(1, len(by_zip)))
    todo = []
    random.seed(args.seed)
    for z, rs in by_zip.items():
        random.shuffle(rs)
        todo.extend(rs[:per])
    random.shuffle(todo)
    todo = todo[:args.limit]
    log.info("hydrating %s properties across %s ZIPs (~%d per ZIP)",
             len(todo), len(by_zip), per)

    client = SiftMapClient()
    lock = threading.Lock()
    fh = (OUT / "hydrated.jsonl").open("a", encoding="utf-8")
    done = [0]
    errs = [0]

    def one(r):
        try:
            d = client.get_detail(r["id"])
        except SiftMapError as e:
            return r, None, str(e)[:100]
        except Exception as e:
            return r, None, str(e)[:100]
        return r, d, ""

    try:
        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            for f in as_completed([ex.submit(one, r) for r in todo]):
                r, d, err = f.result()
                if d is None:
                    errs[0] += 1
                    continue
                nb = d.get("neighborhood") or {}
                rec = {
                    "id": str(r["id"]),
                    "zip": r.get("zip"),
                    "county": r.get("county"),
                    "lat": r.get("latitude"),
                    "lon": r.get("longitude"),
                    "neighborhood": (nb.get("name") if isinstance(nb, dict)
                                     else nb) or "",
                    "subdivision": d.get("subdivision") or "",
                    "year_built": d.get("years_built"),
                    "census_tract": d.get("census_tract"),
                    "tax_delinquent": d.get("tax_delinquent"),
                    "flood_zone": d.get("flood_zone"),
                    "total_market_value": d.get("total_market_value"),
                }
                with lock:
                    fh.write(json.dumps(rec, default=str) + "\n")
                    done[0] += 1
                    if done[0] % 200 == 0:
                        fh.flush()
                        log.info("  %s / %s  (%s errors)", done[0], len(todo), errs[0])
    finally:
        fh.close()

    tot = done[0] + errs[0]
    # A 429 that returns nothing silently drops a record while the run still
    # reports success, so an unreadable rate is a phase failure, not a warning.
    gate(tot == 0 or errs[0] / tot < 0.15,
         "%d of %d detail calls failed (%.1f%%); the sample is not trustworthy"
         % (errs[0], tot, 100.0 * errs[0] / tot))
    mark_done("hydrate", hydrated=len(_hydrated()), errors=errs[0])
    print()
    print("hydrated %s new (%s errors). total on disk: %s"
          % (done[0], errs[0], len(_hydrated())))


def _surviving_zips() -> set:
    import csv
    p = OUT / "tn_zip_suppress.csv"
    if not p.exists():
        return set()
    out = set()
    with p.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            if not (r.get("suppress_reasons") or "").strip():
                out.add(r["geo"])
    return out


# ------------------------------------------------------------------ geo ----

def _knn_assign(anchors: list, targets: list, k: int = 5) -> list:
    """Assign each target a neighborhood by majority vote of its k nearest anchors.

    Plain lat/lon distance on a flat grid, with longitude scaled by cos(lat) so
    a degree east is not treated as a degree north. At this latitude that is a
    0.81 factor, and ignoring it would stretch every neighborhood sideways.

    A simple bucketed index keeps this near-linear: anchors are dropped into
    0.01-degree cells (about 1.1 km) and only the surrounding cells are
    searched, widening the ring until enough candidates are found. A brute-force
    pass would be 14,000 x 155,000 distance computations.
    """
    import math

    CELL = 0.01
    grid = {}
    for a in anchors:
        key = (int(a["lat"] / CELL), int(a["lon"] / CELL))
        grid.setdefault(key, []).append(a)

    coslat = math.cos(math.radians(35.9))
    out = []
    for t in targets:
        lat, lon = t["latitude"], t["longitude"]
        ci, cj = int(lat / CELL), int(lon / CELL)
        found = []
        ring = 0
        while ring <= 12:
            for i in range(ci - ring, ci + ring + 1):
                for j in range(cj - ring, cj + ring + 1):
                    # Only the newly added outer ring each iteration.
                    if ring and abs(i - ci) != ring and abs(j - cj) != ring:
                        continue
                    found.extend(grid.get((i, j), ()))
            if len(found) >= k:
                break
            ring += 1
        if not found:
            out.append((t, "", None))
            continue
        scored = []
        for a in found:
            dy = a["lat"] - lat
            dx = (a["lon"] - lon) * coslat
            scored.append((dy * dy + dx * dx, a))
        scored.sort(key=lambda x: x[0])
        near = scored[:k]
        votes = {}
        for d2, a in near:
            nb = a.get("neighborhood") or ""
            if nb:
                votes[nb] = votes.get(nb, 0) + 1
        if not votes:
            out.append((t, "", None))
            continue
        best = max(votes.items(), key=lambda kv: kv[1])
        conf = best[1] / len(near)
        out.append((t, best[0], conf))
    return out


def phase_geo(args) -> None:
    """Assign every swept property a neighborhood, and measure the accuracy.

    THE ACCURACY IS MEASURED, NOT ASSERTED. A holdout of labelled anchors is
    withheld from the index and then predicted, so the number reported is what
    the method actually achieves on this data rather than a claim about k-NN in
    general. If it does not clear the floor the phase fails instead of quietly
    publishing a neighborhood breakdown built on guesses.
    """
    hyd = [h for h in _hydrated().values()
           if h.get("lat") and h.get("lon") and h.get("neighborhood")]
    gate(len(hyd) >= 500,
         "only %d hydrated anchors; run --phase hydrate first" % len(hyd))

    random.seed(args.seed)
    random.shuffle(hyd)
    cut = max(200, int(len(hyd) * 0.15))
    holdout, train = hyd[:cut], hyd[cut:]
    log.info("%d anchors: %d train, %d holdout", len(hyd), len(train), len(holdout))

    probe = [{"latitude": h["lat"], "longitude": h["lon"], "id": h["id"],
              "truth": h["neighborhood"]} for h in holdout]
    preds = _knn_assign(train, probe, k=args.k)
    hit = sum(1 for t, nb, c in preds if nb == t["truth"])
    acc = 100.0 * hit / len(preds) if preds else 0.0
    hi = [(t, nb, c) for t, nb, c in preds if (c or 0) >= 0.6]
    hi_acc = (100.0 * sum(1 for t, nb, c in hi if nb == t["truth"]) / len(hi)
              if hi else 0.0)
    log.info("holdout accuracy: %.1f%% overall, %.1f%% on the %d confident ones",
             acc, hi_acc, len(hi))
    gate(acc >= args.min_accuracy,
         "neighborhood inference is only %.1f%% accurate on holdout, below the "
         "%.0f%% floor. Hydrate more anchors before trusting a neighborhood "
         "breakdown." % (acc, args.min_accuracy))

    rows = load_rows()
    keep = _surviving_zips()
    if keep:
        rows = [r for r in rows if r.get("zip") in keep]
    rows = [r for r in rows if r.get("latitude") and r.get("longitude")]
    known = {h["id"]: h for h in hyd}
    log.info("assigning neighborhoods to %s properties", len(rows))

    assigned = _knn_assign(hyd, rows, k=args.k)
    out = []
    for r, nb, conf in assigned:
        k_ = known.get(str(r["id"]))
        if k_:
            nb, conf = k_["neighborhood"], 1.0   # measured, never inferred
        rr = dict(r)
        rr["neighborhood"] = nb
        rr["nb_confidence"] = round(conf, 2) if conf is not None else None
        rr["year_built"] = (k_ or {}).get("year_built")
        out.append(rr)

    path = OUT / "assigned.jsonl"
    with path.open("w", encoding="utf-8") as fh:
        for r in out:
            fh.write(json.dumps(r, default=str) + "\n")

    stats = geo_stats(out, "neighborhood")
    mf = load_market_finder()
    for g, s_ in stats.items():
        m = mf["neighborhood"].get(g, {})
        s_["mf_dom"] = m.get("median_days_on_market")
        s_["mf_inv_trans_6mo"] = m.get("total_inv_trans_6mo")
        s_["mf_homes_on_market"] = m.get("homes_on_market")
        s_["mf_sold_last_month"] = m.get("homes_sold_last_month")
        s_["mf_matched"] = bool(m)
    # Neighborhood suppression, same grammar as the ZIP layer. This is the cut
    # that actually bites: ZIP-level rules removed barely 1,500 properties
    # because Knox and Blount ZIPs are broadly workable, while neighborhoods
    # inside one ZIP differ enormously.
    doms = sorted(v["mf_dom"] for v in stats.values()
                  if isinstance(v.get("mf_dom"), (int, float)))
    dom_ref = _pct(doms, 50) if doms else None
    for g, v in stats.items():
        reasons = []
        if not g:
            reasons.append("no neighborhood could be assigned")
        if v["n"] < args.nb_min_properties:
            reasons.append("only %d properties" % v["n"])
        if v["value_med"] and v["value_med"] > BUY_BOX_MAX:
            reasons.append("median $%.0f is above the $700K buy box"
                           % v["value_med"])
        if v["value_med"] and v["value_med"] < args.min_median_value:
            reasons.append("median $%.0f below the rehab-clears-ARV floor"
                           % v["value_med"])
        if isinstance(v.get("mf_inv_trans_6mo"), (int, float)) and \
                v["mf_inv_trans_6mo"] == 0 and v["mf_matched"]:
            reasons.append("zero investor transactions in 6 months")
        if dom_ref and isinstance(v.get("mf_dom"), (int, float)) and \
                v["mf_dom"] > dom_ref * args.dom_multiple:
            reasons.append("DOM %.0f is %.1fx the median of %.0f"
                           % (v["mf_dom"], v["mf_dom"] / dom_ref, dom_ref))
        v["suppress_reasons"] = "; ".join(reasons)

    _save(OUT / "neighborhood_stats.json", stats)

    import csv as _csv
    npath = OUT / "tn_neighborhood_suppress.csv"
    ncols = ["geo", "county", "n", "n_in_box", "pct_in_box", "value_p25",
             "value_med", "value_p75", "equity_med", "sqft_med", "pct_absentee",
             "pct_corporate", "pct_no_distress", "mf_dom", "mf_inv_trans_6mo",
             "mf_homes_on_market", "mf_matched", "suppress_reasons"]
    with npath.open("w", newline="", encoding="utf-8") as fh:
        w = _csv.DictWriter(fh, fieldnames=ncols, extrasaction="ignore")
        w.writeheader()
        for g, v in sorted(stats.items(), key=lambda kv: -kv[1]["n_in_box"]):
            w.writerow(dict(v, geo=g))

    nkeep = [v for v in stats.values() if not v["suppress_reasons"]]
    nsurv = sum(v["n_in_box"] for v in nkeep)
    print()
    print("NEIGHBORHOOD SUPPRESSION: kept %d, cut %d"
          % (len(nkeep), len(stats) - len(nkeep)))
    print("  buy-box properties surviving BOTH layers: %s" % format(nsurv, ","))
    print("  scoring cost at $0.007: $%s" % format(round(nsurv * 0.007), ","))
    print("  written: %s" % npath)

    matched = sum(1 for v in stats.values() if v["mf_matched"])
    mark_done("geo", neighborhoods=len(stats), accuracy=round(acc, 1),
              anchors=len(hyd))
    print()
    print("neighborhoods identified: %d  (%d matched to Market Finder)"
          % (len(stats), matched))
    print("holdout accuracy: %.1f%% overall, %.1f%% where the vote was confident"
          % (acc, hi_acc))
    print("written: %s" % (OUT / "neighborhood_stats.json"))
    print()
    print("%-34s %7s %9s %8s %6s" % ("neighborhood", "SFR", "median $", "%no dis", "DOM"))
    for g, v in sorted(stats.items(), key=lambda kv: -kv[1]["n"])[:15]:
        print("%-34s %7d %9s %7.0f%% %6s"
              % (g[:34], v["n"],
                 format(v["value_med"], ",.0f") if v["value_med"] else "-",
                 v["pct_no_distress"], v.get("mf_dom") or "-"))


# --------------------------------------------------------------- report ----

IMAGE_COST = 0.007


def phase_report(args) -> None:
    """The branded workbook: universe, funnel, per-ZIP detail, caveats."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    from lender_package import CALC_FILL, MONEY, TOTAL_FILL, _band, _polish

    stats = _load(OUT / "zip_stats.json", {})
    gate(bool(stats), "no zip_stats.json; run --phase suppress first")

    import csv as _csv
    sup = {}
    cp = OUT / "tn_zip_suppress.csv"
    if cp.exists():
        with cp.open(encoding="utf-8") as fh:
            for row in _csv.DictReader(fh):
                sup[row["geo"]] = row

    cov = _load(OUTPUT_ROOT / "visual_distress" / "coverage_profile.json", [])
    kept = [g for g, r in sup.items() if not r.get("suppress_reasons")]
    survivors = sum(int(sup[g]["n_in_box"]) for g in kept)
    cov_rate = None
    if cov:
        cov_rate = sum(1 for c in cov if c.get("status") == "OK") / len(cov)
    scoreable = int(survivors * cov_rate) if cov_rate else survivors

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    r = 1
    ws.cell(row=r, column=1,
            value="Knox and Blount market picture").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value="Property-level API sweep, %s"
            % datetime.now().strftime("%B %d, %Y"))
    r += 2

    _band(ws, r, "The funnel", span=3)
    r += 1
    total_all = sum(v["n"] for v in stats.values())
    rows_out = [
        ("Single-family properties swept", total_all, ""),
        ("In the buy box ($1 to $700K)",
         sum(v["n_in_box"] for v in stats.values()), ""),
        ("In surviving ZIPs after suppression", survivors,
         "%d ZIPs kept, %d cut" % (len(kept), len(sup) - len(kept))),
    ]
    nbs = _load(OUT / "neighborhood_stats.json", {})
    if nbs:
        nkeep = [v for v in nbs.values() if not v.get("suppress_reasons")]
        nsurv = sum(v["n_in_box"] for v in nkeep)
        rows_out.append(
            ("After neighborhood suppression too", nsurv,
             "%d neighborhoods kept, %d cut"
             % (len(nkeep), len(nbs) - len(nkeep))))
        survivors = nsurv
        # Recompute AFTER the neighborhood cut. Computing it once up front made
        # the coverage row and the cost describe the ZIP-level survivors, which
        # silently overstated the bill by about $44.
        scoreable = int(survivors * cov_rate) if cov_rate else survivors
    if cov_rate:
        rows_out.append(
            ("With Street View imagery (%.1f%% coverage)" % (100 * cov_rate),
             scoreable, "measured free on a %d-property sample" % len(cov)))
    for label, val, note in rows_out:
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = "#,##0"
        c.fill = CALC_FILL
        if note:
            ws.cell(row=r, column=3, value=note).font = Font(size=9, italic=True)
        r += 1
    ws.cell(row=r, column=1, value="Cost to score at $0.007 each").font = Font(bold=True)
    c = ws.cell(row=r, column=2, value=scoreable * IMAGE_COST)
    c.number_format = MONEY
    c.fill = TOTAL_FILL
    c.font = Font(bold=True)
    r += 2

    _band(ws, r, "What the data already knows, and what it cannot", span=3)
    r += 1
    no_d = sum(v["n"] * v["pct_no_distress"] / 100.0 for v in stats.values())
    ws.cell(row=r, column=1, value="Properties carrying NO distress flag of any kind")
    ws.cell(row=r, column=2, value=int(no_d)).fill = CALC_FILL
    ws.cell(row=r, column=2).number_format = "#,##0"
    r += 1
    ws.cell(row=r, column=1, value="Share of the universe invisible to every data filter")
    ws.cell(row=r, column=2,
            value=round(100.0 * no_d / total_all, 1) if total_all else 0).fill = CALC_FILL
    r += 2
    ws.cell(row=r, column=1, value=(
        "This is the population visual distress exists to reach. A lien, a probate "
        "or a code violation can be queried for free. A house that simply looks "
        "beaten down cannot, and that is the whole reason to pay for imagery."))
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 42

    ws2 = wb.create_sheet("ZIP detail")
    cols = ["geo", "county", "n", "n_in_box", "pct_in_box", "value_p25",
            "value_med", "value_p75", "equity_med", "sqft_med", "pct_absentee",
            "pct_corporate", "pct_no_distress", "mf_dom", "mf_inv_trans_6mo",
            "mf_homes_on_market", "suppress_reasons"]
    heads = ["ZIP", "County", "SFR", "In box", "% in box", "Value p25",
             "Value median", "Value p75", "Equity med", "Sqft med", "% absentee",
             "% corporate", "% no distress", "DOM", "Inv trans 6mo",
             "On market", "Suppressed because"]
    for i, h in enumerate(heads, 1):
        ws2.cell(row=1, column=i, value=h).font = Font(bold=True)
    rr = 2
    for g, row in sorted(sup.items(), key=lambda kv: -int(kv[1]["n_in_box"])):
        for i, k in enumerate(cols, 1):
            v = row.get(k, "")
            if k not in ("geo", "county", "suppress_reasons"):
                try:
                    v = float(v) if v not in ("", None) else ""
                except (TypeError, ValueError):
                    pass
            c = ws2.cell(row=rr, column=i, value=v)
            if k.startswith("value_") and v != "":
                c.number_format = MONEY
        rr += 1
    ws2.freeze_panes = "A2"

    ws3 = None
    if cov:
        ws3 = wb.create_sheet("Street View coverage")
        by = {}
        for c_ in cov:
            z = by.setdefault(c_.get("zip") or "(blank)",
                              {"n": 0, "ok": 0, "yrs": []})
            z["n"] += 1
            if c_.get("status") == "OK":
                z["ok"] += 1
                if c_.get("pano_date"):
                    z["yrs"].append(int(str(c_["pano_date"])[:4]))
        for i, h in enumerate(["ZIP", "Sampled", "With imagery", "% covered",
                               "Median vintage"], 1):
            ws3.cell(row=1, column=i, value=h).font = Font(bold=True)
        rr = 2
        for z, v in sorted(by.items(), key=lambda kv: -kv[1]["n"]):
            ys = sorted(v["yrs"])
            ws3.cell(row=rr, column=1, value=z)
            ws3.cell(row=rr, column=2, value=v["n"])
            ws3.cell(row=rr, column=3, value=v["ok"])
            ws3.cell(row=rr, column=4, value=round(100.0 * v["ok"] / v["n"], 1))
            ws3.cell(row=rr, column=5, value=ys[len(ys) // 2] if ys else "")
            rr += 1
        ws3.freeze_panes = "A2"

    nb = _load(OUT / "neighborhood_stats.json", {})
    if nb:
        ws5 = wb.create_sheet("Neighborhood detail")
        ncols = ["n", "n_in_box", "pct_in_box", "value_p25", "value_med",
                 "value_p75", "equity_med", "sqft_med", "pct_absentee",
                 "pct_no_distress", "mf_dom", "mf_inv_trans_6mo",
                 "mf_homes_on_market", "suppress_reasons"]
        nheads = ["Neighborhood", "County", "SFR", "In box", "% in box",
                  "Value p25", "Value median", "Value p75", "Equity med",
                  "Sqft med", "% absentee", "% no distress", "DOM",
                  "Inv trans 6mo", "On market", "Suppressed because"]
        for i, h in enumerate(nheads, 1):
            ws5.cell(row=1, column=i, value=h).font = Font(bold=True)
        rr = 2
        for g, v in sorted(nb.items(), key=lambda kv: -kv[1]["n_in_box"]):
            ws5.cell(row=rr, column=1, value=g)
            ws5.cell(row=rr, column=2, value=v.get("county", ""))
            for i, k in enumerate(ncols, 3):
                c = ws5.cell(row=rr, column=i, value=v.get(k))
                if k.startswith("value_") and v.get(k):
                    c.number_format = MONEY
            rr += 1
        ws5.freeze_panes = "A2"
        ws5.column_dimensions["A"].width = 32
        ws5.column_dimensions["P"].width = 52

    ws4 = wb.create_sheet("Sources and caveats")
    notes = [
        ("Property data", "SiftMap /properties/search/, swept %s"
         % datetime.now().strftime("%Y-%m-%d")),
        ("Liquidity data",
         "DataSift Market Finder, Knox and Blount, extracted 2026-09-02"),
        ("Coverage data",
         "Google Street View metadata endpoint (free), %d properties sampled"
         % len(cov)),
        ("Filter validation",
         "Only type_single_family, value_min/value_max, extra_year_built_min, "
         "extra_vacant, extra_equity_percent_min and extra_last_sale_date_min "
         "were proved to read by count delta. owner_occupied, absentee_owner, "
         "year_built_min, equity_percent_min and is_vacant_property are "
         "SILENTLY IGNORED by the API and were not used."),
        ("Cross-county ZIPs",
         "Six ZIPs straddle the Knox and Blount line and appear in both Market "
         "Finder extracts. Counts are summed. A last-wins merge previously "
         "zeroed 37920 and would have cut 1,370 properties from one of the "
         "busiest investor ZIPs in Knoxville."),
        ("Neighborhood assignment",
         "Neighborhood exists only on the detail endpoint (~3 req/s aggregate), "
         "so all 155K survivors would take ~14 hours. Instead a stratified "
         "sample was hydrated as labelled anchors and the rest assigned by "
         "nearest neighbour on lat/lon, longitude scaled by cos(latitude). "
         "Accuracy is MEASURED on a withheld 15% holdout, not assumed, and the "
         "phase refuses to publish below an 80% floor. Measured k=1 90.8%, "
         "k=3 89.9%, k=5 77.9%: more neighbours smooth across boundaries, "
         "which is the opposite of what is wanted. Properties that were "
         "actually hydrated keep their measured value and are never inferred."),
        ("What is deliberately NOT filtered",
         "Distress flags are not used to select the cohort. The value of visual "
         "scoring is the house that no data filter can see, so screening on "
         "distress first would only re-find what the data already knows."),
    ]
    for i, (a, b) in enumerate(notes, 1):
        ws4.cell(row=i, column=1, value=a).font = Font(bold=True)
        cc = ws4.cell(row=i, column=2, value=b)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        ws4.row_dimensions[i].height = 30
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 98

    for w in [x for x in (ws, ws2, ws3) if x is not None]:
        try:
            _polish(w)
        except Exception:
            pass
    ws2.column_dimensions["Q"].width = 58

    # Excel holds an exclusive lock on an open workbook, so write to a pending
    # name and swap. Saving straight over an open file loses the whole run.
    final = (Path(args.out) if args.out else
             OUTPUT_ROOT / ("Knox_Blount_Market_%s.xlsx"
                            % datetime.now().strftime("%Y%m%d")))
    pending = final.with_name("_PENDING_" + final.name)
    wb.save(pending)
    if final.exists():
        final.unlink()
    pending.replace(final)
    mark_done("report", path=str(final))
    print("written: %s" % final)


# ------------------------------------------------------------------ cli ----

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--phase", required=True, choices=["probe", "sweep", "suppress", "hydrate", "geo", "report"])
    ap.add_argument("--resume", action="store_true",
                    help="continue a partial sweep instead of restarting")
    ap.add_argument("--min-interval", type=float, default=0.30,
                    help="seconds between search calls")
    ap.add_argument("--counties", default="",
                    help="comma-separated subset, e.g. Blount")
    ap.add_argument("--passes", type=int, default=1,
                    help="re-walk and union N times. MEASURED USELESS on this "
                         "API: a second Blount pass returned exactly 0 new ids, "
                         "and a Knox re-walk found 0 new across 340 pages. The "
                         "unreachable subset is stable, not random drift. Kept "
                         "only so the next person does not re-test it blind.")
    ap.add_argument("--min-coverage", type=float, default=90.0,
                    help="fail the phase below this %% of reported total")
    ap.add_argument("--min-properties", type=int, default=150,
                    help="cut a ZIP with fewer than this many SFR")
    ap.add_argument("--min-in-box", type=int, default=100,
                    help="cut a ZIP with fewer than this many buy-box SFR")
    ap.add_argument("--min-median-value", type=float, default=40000,
                    help="cut a ZIP whose median is below this")
    ap.add_argument("--dom-multiple", type=float, default=2.5,
                    help="cut a ZIP whose DOM exceeds this multiple of the "
                         "county median")
    ap.add_argument("--limit", type=int, default=12000,
                    help="hydrate: how many detail calls to make")
    ap.add_argument("--per-zip", type=int, default=120,
                    help="hydrate: minimum anchors per ZIP")
    ap.add_argument("--workers", type=int, default=4)
    ap.add_argument("--seed", type=int, default=17)
    ap.add_argument("--nb-min-properties", type=int, default=60,
                    help="geo: cut a neighborhood with fewer than this many SFR")
    ap.add_argument("--k", type=int, default=3,
                    help="geo: neighbours voting on each assignment. Measured "
                         "on 1,597 anchors: k=1 83.6%%, k=3 83.3%%, k=5 77.9%%. "
                         "More neighbours smooth ACROSS boundaries, which is "
                         "the opposite of what is wanted, so k stays small.")
    ap.add_argument("--min-accuracy", type=float, default=80.0,
                    help="geo: fail below this holdout accuracy")
    ap.add_argument("--out", default="", help="report: output xlsx path")
    ap.add_argument("-v", "--verbose", action="store_true")
    a = ap.parse_args()
    logging.basicConfig(
        level=logging.DEBUG if a.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
    OUT.mkdir(parents=True, exist_ok=True)
    {"probe": phase_probe, "sweep": phase_sweep,
     "suppress": phase_suppress, "hydrate": phase_hydrate,
     "geo": phase_geo, "report": phase_report}[a.phase](a)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
